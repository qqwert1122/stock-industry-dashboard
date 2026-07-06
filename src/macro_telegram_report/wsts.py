from __future__ import annotations

from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Section
from .utils import fmt_number, fmt_pct, pct_change, to_float

MONTH_NAMES = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def collect_wsts(config: dict[str, Any], session: requests.Session) -> Section | None:
    wsts_config = config.get("wsts", {})
    if not wsts_config.get("enabled", True):
        return None

    xlsx_url = wsts_config.get("download_url") or find_wsts_xlsx_url(
        str(wsts_config["page_url"]), session
    )
    workbook_response = session.get(xlsx_url, timeout=60)
    workbook_response.raise_for_status()

    workbook = load_workbook(BytesIO(workbook_response.content), data_only=True, read_only=True)
    regions = [str(region) for region in wsts_config.get("regions", ["Worldwide"])]
    lines = summarize_sheet(workbook["Monthly Data"], regions, "월간")

    if wsts_config.get("include_3mma", True) and "3MMA" in workbook.sheetnames:
        lines.extend(summarize_sheet(workbook["3MMA"], regions, "3MMA"))

    lines.append(f"- 파일: {xlsx_url}")
    return Section("WSTS", lines)


def find_wsts_xlsx_url(page_url: str, session: requests.Session) -> str:
    response = session.get(page_url, timeout=30)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    candidates: list[str] = []
    for link in soup.find_all("a", href=True):
        href = str(link["href"])
        text = link.get_text(" ", strip=True).lower()
        lower_href = href.lower()
        if lower_href.endswith(".xlsx") or ".xlsx" in lower_href:
            score_text = f"{text} {lower_href}"
            if "historical" in score_text or "billing" in score_text or "wsts" in score_text:
                candidates.append(urljoin(page_url, href))

    if not candidates:
        raise ValueError("WSTS 페이지에서 XLSX 다운로드 링크를 찾지 못했습니다.")
    return candidates[0]


def summarize_sheet(sheet: Worksheet, regions: list[str], label: str) -> list[str]:
    series = parse_wsts_sheet(sheet)
    lines: list[str] = []

    for region in regions:
        points = sorted(series.get(region, []), key=lambda point: point[0])
        if not points:
            lines.append(f"- {label} {region}: 데이터 없음")
            continue

        latest_date, latest_value = points[-1]
        previous_value = points[-2][1] if len(points) >= 2 else None
        yoy_value = next(
            (
                value
                for observed_date, value in points
                if observed_date.year == latest_date.year - 1
                and observed_date.month == latest_date.month
            ),
            None,
        )
        latest_billions = latest_value / 1_000_000
        previous_billions = previous_value / 1_000_000 if previous_value is not None else None
        yoy_billions = yoy_value / 1_000_000 if yoy_value is not None else None

        mom = pct_change(latest_billions, previous_billions)
        yoy = pct_change(latest_billions, yoy_billions)
        lines.append(
            "- "
            f"{label} {region}: {latest_date.year}.{latest_date.month:02d} "
            f"${fmt_number(latest_billions)}B "
            f"(MoM {fmt_pct(mom)}, YoY {fmt_pct(yoy)})"
        )

    return lines


def parse_wsts_sheet(sheet: Worksheet) -> dict[str, list[tuple[date, float]]]:
    month_columns: dict[int, int] = {}
    current_year: int | None = None
    series: dict[str, list[tuple[date, float]]] = defaultdict(list)

    for row in sheet.iter_rows():
        first_value = row[0].value if row else None
        first_text = str(first_value).strip() if first_value is not None else ""

        if first_text.isdigit() and len(first_text) == 4:
            current_year = int(first_text)
            continue

        if not month_columns:
            for index, cell in enumerate(row):
                cell_text = str(cell.value).strip().lower() if cell.value is not None else ""
                if cell_text in MONTH_NAMES:
                    month_columns[index] = MONTH_NAMES[cell_text]
            continue

        if current_year is None or not first_text:
            continue

        for index, month in month_columns.items():
            if index >= len(row):
                continue
            value = to_float(row[index].value)
            if value is None or value <= 0:
                continue
            series[first_text].append((date(current_year, month, 1), value))

    return dict(series)
