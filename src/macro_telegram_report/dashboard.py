from __future__ import annotations

import csv
import os
import re
import time
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urljoin
from xml.etree import ElementTree
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .briefing import (
    briefing_date_key,
    build_briefing_card,
    increment_gemini_usage,
    load_briefing_index,
    load_gemini_usage,
    load_latest_briefing_card,
    load_recent_briefing_cards,
    update_intraday_track,
    write_briefing_outputs,
)
from .dashboard_briefing import (
    GEMINI_DAILY_CALL_LIMIT,
    GEMINI_DEFAULT_MODEL,
    GEMINI_GENERATE_URL,
    benchmark_change_summary,
    brief_metric,
    briefing_generation_decision,
    briefing_metric_changes,
    briefing_metric_snapshot,
    briefing_session_context,
    build_morning_briefing,
    compact_stock_market_narrative,
    consecutive_low_signal_count,
    daily_move_significance,
    equity_lead_rows,
    equity_market_for_metric,
    extract_gemini_text,
    gemini_morning_briefing_prompt,
    industry_signal_rows,
    is_persistent_market_metric,
    is_representative_stock,
    load_industry_narratives,
    market_narrative_context,
    metric_change_meaning,
    metric_change_summary,
    metric_direction,
    metric_kind_label,
    narrative_context_for_briefing,
    normalize_briefing_bullets,
    normalize_gemini_briefing,
    normalized_metric_name,
    observed_at_progressed,
    parse_json_object,
    persistent_event_threshold,
    persistent_market_events,
    recent_narrative_industries,
    relevant_briefing_industries,
    request_gemini_briefing,
    rule_based_bullets,
    rule_based_morning_briefing,
    rule_based_summary,
    short_text,
    subject_particle,
    top_mover_metrics,
    topic_label,
    topic_particle,
)
from .dashboard_localization import (
    CAPEX_MEANINGS,
    CAPEX_MEANINGS_EN,
    EN_DEPTH_LABELS,
    EN_EXPORT_ITEM_LABELS,
    EN_FLOW_MEASURE_FALLBACKS,
    EN_FREQUENCY_LABELS,
    EN_GROUP_LABELS,
    EN_INDUSTRY_LABELS,
    EN_INVESTOR_FALLBACKS,
    EN_MARKET_FALLBACKS,
    EN_MEANING_LABELS,
    EN_METRIC_NAME_LABELS,
    EN_PHRASE_FALLBACKS,
    EN_UNIT_LABELS,
    HANGUL_RE,
    WSTS_3MMA_MEANING,
    WSTS_3MMA_MEANING_EN,
    WSTS_REGION_MEANINGS,
    WSTS_REGION_MEANINGS_EN,
    clean_display_text,
    contains_hangul,
    english_depth,
    english_export_item,
    english_frequency,
    english_generic_text,
    english_group,
    english_industry,
    english_metric_meaning,
    english_metric_name,
    english_sentence_fallback,
    english_unit,
)
from .dashboard_metrics import (
    DEFAULT_INDUSTRIES,
    assign_market_navigation_fields,
    assign_metric_country_fields,
    compact_date_label,
    configured_industries,
    export_meaning,
    find_yoy_value,
    format_abs_change,
    format_value,
    infer_export_industry,
    infer_flow_metric_meaning,
    infer_metric_country,
    infer_metric_depth,
    infer_metric_group,
    infer_metric_meaning,
    korea_fear_greed_meaning,
    make_metric,
    next_update_label,
    normalize_market_categories,
    period_label,
    sec_capex_meaning,
    set_market_category,
    stablecoin_meaning,
    status_label,
    status_to_automation,
    visible_dashboard_metrics,
    vkospi_meaning,
    wsts_metric_meaning,
)
from .history_store import (
    HistoryStore,
    downsample_history,
    parse_stored_points,
    percentile_stats,
)
from .interpretation import apply_interpretations
from .fetch_log import (
    FetchLogger,
    append_fetch_log_run,
    current_logger,
    load_fetch_log_history,
    sanitize_message,
    save_fetch_log_history,
    use_fetch_logger,
)
from .event_calendar import build_event_calendar, write_event_calendar
from .future_timeline import build_future_timeline, write_future_timeline
from .korea_exports import build_data_go_kr_url, fetch_itemtrade_records, first_text, local_name
from .alerts import process_alerts
from .market_gauges import build_market_gauges
from .market_flows import (
    FLOW_MEASURES,
    fetch_krx_futures_flow_rows,
    fetch_krx_main_investor_flow_rows,
    fetch_krx_stock_flow_rows,
    investor_slug,
    load_raw_flow_snapshot,
    raw_flow_investors,
    raw_flow_known_dates,
    raw_flow_series,
    raw_flow_snapshot_path,
    rolling_sum_series,
    store_raw_flow_rows,
)
from .market_valuation import (
    fetch_finra_margin_series,
    fetch_krx_valuation_series,
    fetch_multpl_series,
)
from .market_sentiment import (
    KRX_API_BASE,
    KRX_SOURCE_URL,
    build_korea_fear_greed_score,
    collect_market_snapshot,
    fetch_cnn_fear_greed,
    fetch_vkospi_points,
    merge_existing_and_incoming,
    metric_full_points,
    missing_recent_dates,
)
from .site_output import (
    copy_dashboard_assets,
    copy_signal_log_output,
    load_admin_template,
    load_dashboard_template,
    render_dashboard_html,
    write_admin_html,
    write_dashboard_shell,
)
from .storage import load_json, write_json
from .utils import (
    add_months,
    fmt_pct,
    month_key,
    numeric_values_equal,
    parse_iso_date,
    pct_change,
    to_float,
)
from .wsts import find_wsts_xlsx_url, parse_wsts_sheet

FRED_OBSERVATIONS_URL = "https://api.stlouisfed.org/fred/series/observations"
FISCALDATA_TGA_URL = (
    "https://api.fiscaldata.treasury.gov/services/api/fiscal_service/"
    "v1/accounting/dts/operating_cash_balance"
)
MARKET_GAUGE_HISTORY_FILENAME = "market_gauges_history.json"
MARKET_GAUGE_HISTORY_VERSION = 1
FETCH_LOG_HISTORY_FILENAME = "fetch_log_history.json"
FETCH_LOG_FILENAME = "fetch_log.json"
FETCH_SOURCE_ENDPOINTS = {
    "WSTS": "https://www.wsts.org/76/Recent-News-Release",
    "FRED": FRED_OBSERVATIONS_URL,
    "미국 재무부 DTS": FISCALDATA_TGA_URL,
    "미국 유동성": FRED_OBSERVATIONS_URL,
    "ECOS 신용스프레드": "https://ecos.bok.or.kr/api/",
    "ECOS 매크로": "https://ecos.bok.or.kr/api/",
    "금투협회 증시자금": "https://www.data.go.kr/data/15094809/openapi.do",
    "대표주가/시장지수": "https://query1.finance.yahoo.com/v8/finance/chart/{symbol}",
    "스테이블코인": "https://stablecoins.llama.fi/stablecoins",
    "World Bank 원자재": "https://thedocs.worldbank.org/",
    "SEC CAPEX": "https://data.sec.gov/submissions/{cik}.json",
    "USAspending 방산": "https://api.usaspending.gov/api/v2/search/spending_over_time/",
    "EIA": "https://api.eia.gov/v2/",
    "openFDA": "https://api.fda.gov/",
    "ClinicalTrials.gov": "https://clinicaltrials.gov/api/v2/studies",
    "Launch Library": "https://ll.thespacedevs.com/2.3.0/launches/",
    "AFDC EV 충전": "https://developer.nrel.gov/api/alt-fuel-stations/v1/count.json",
    "KOSIS": "https://kosis.kr/openapi/",
    "한국 수출": "https://apis.data.go.kr/1220000/Itemtrade",
    "밸류에이션/수급": "https://finance.naver.com/",
    "시장 수급": "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd",
    "시장 파생지표": "https://api.upbit.com/v1/ticker",
    "시장 심리": "https://production.dataviz.cnn.io/index/fearandgreed/graphdata/{start_date}",
    "미래 타임라인": "data/future/technologies.yaml",
}
INDUSTRY_ICONS = {
    "반도체": "assets/industry-icons/semiconductor.png",
    "자동차": "assets/industry-icons/auto.png",
    "전기차": "assets/industry-icons/electric-car.png",
    "조선": "assets/industry-icons/shipbuilding.png",
    "철강/소재": "assets/industry-icons/steel-materials.png",
    "화학/정유": "assets/industry-icons/oil-barrel.png",
    "은행/금융": "assets/industry-icons/finance.png",
    "건설/부동산": "assets/industry-icons/construction-real-estate.png",
    "방산": "assets/industry-icons/tank.png",
    "스테이블코인": "assets/industry-icons/bitcoin.png",
    "전력": "assets/industry-icons/power.png",
    "로봇": "assets/industry-icons/robotics.png",
    "우주": "assets/industry-icons/space.png",
    "바이오": "assets/industry-icons/biotech.png",
    "배터리": "assets/industry-icons/battery.png",
    "데이터인프라": "assets/industry-icons/data-infrastructure.png",
    "매크로": "assets/industry-icons/macro-trend.png",
}
INDUSTRY_SUMMARIES = {
    "반도체": "메모리, 파운드리, 장비, AI 인프라 수요를 함께 봅니다.",
    "자동차": "완성차 판매와 승용차 수출로 자동차 수요 사이클을 봅니다.",
    "전기차": "순수 전기차 수출과 EV 판매 흐름으로 전기차 보급 속도를 봅니다.",
    "조선": "운임, 선가, 발주, 선박 수출로 조선 수요와 이익 사이클을 봅니다.",
    "철강/소재": "원자재 가격과 중국 제조업 경기로 소재 수요를 봅니다.",
    "화학/정유": "유가, 원료, 제품 스프레드로 마진 방향을 확인합니다.",
    "은행/금융": "금리, 스프레드, 대출, 연체율로 은행 수익성과 신용위험을 봅니다.",
    "건설/부동산": "착공, 허가, 금리, 가격으로 부동산 선행 흐름을 봅니다.",
    "방산": "수주, 생산, 수출 흐름으로 방산 수요를 확인합니다.",
    "스테이블코인": "온체인 달러 유동성과 결제/거래 수요를 봅니다.",
    "전력": "전력 가격, 생산, 장비 수출로 인프라 수요를 확인합니다.",
    "로봇": "설비투자와 로봇 수출 흐름을 묶어 봅니다.",
    "우주": "우주/항공 장비 생산과 이벤트 수요를 추적합니다.",
    "바이오": "바이오 제품 가격과 수출 흐름으로 업황을 봅니다.",
    "배터리": "배터리 가격, 원재료, 수출 흐름으로 셀/소재 업황을 봅니다.",
    "데이터인프라": "서버와 네트워크 인프라 투자 흐름을 봅니다.",
    "매크로": "환율, 변동성, 금리로 시장 분위기를 빠르게 확인합니다.",
}

def build_dashboard_site(config: dict[str, Any], output_dir: str | Path, session: requests.Session) -> dict[str, Any]:
    output_path = Path(output_dir)
    data_path = output_path / "data"
    data_path.mkdir(parents=True, exist_ok=True)

    timezone_name = str(config.get("timezone") or "Asia/Seoul")
    previous_payload = load_previous_dashboard_payload(data_path / "dashboard.json")
    logger = FetchLogger(run_type="full", timezone_name=timezone_name)
    with use_fetch_logger(logger):
        payload = build_dashboard_payload(config, session, previous_payload)
        build_today = datetime.now(ZoneInfo(timezone_name)).date()
        long_histories = enrich_metrics_with_history(
            payload.get("metrics", []), config, previous_payload
        )
        apply_interpretations(payload.get("metrics", []), config)
        payload["market_gauges"] = build_market_gauges(payload.get("metrics", []))
        payload["calendar"] = build_event_calendar(config, payload, today=build_today)
        write_event_calendar(data_path / "calendar.json", payload["calendar"])
        attach_future_timeline_payload(config, payload, data_path, build_today, logger, session=session)
        market_gauge_history = update_market_gauge_history(
            load_previous_dashboard_payload(data_path / MARKET_GAUGE_HISTORY_FILENAME),
            payload,
            timezone_name,
        )
        payload["collection_issues"] = annotate_metric_freshness(
            payload.get("metrics", []), build_today
        )
        payload["freshness_summary"] = build_freshness_summary(
            payload.get("metrics", []), str(payload.get("generated_at") or ""), build_today
        )
        annotate_dashboard_updates(payload, previous_payload)
        briefing_generated_at = str(payload.get("generated_at") or datetime.now(ZoneInfo(timezone_name)).isoformat(timespec="seconds"))
        briefing_time = datetime.fromisoformat(briefing_generated_at.replace("Z", "+00:00"))
        briefing_session = briefing_session_context(briefing_time)
        recent_cards = load_recent_briefing_cards(data_path, limit=4)
        previous_card = load_latest_briefing_card(data_path)
        morning_context = {
            "card_type": "morning",
            "session": briefing_session,
            "low_signal": False,
            "changed_metrics": (payload.get("daily_changes") or {}).get("metrics", []),
            "recent_topic_history": [
                {
                    "generated_at": str(card.get("generated_at") or ""),
                    "narrative_topics": card.get("narrative_topics") or [],
                }
                for card in recent_cards
                if isinstance(card, dict)
            ],
        }
        changed_ids = {
            str(item.get("id") or "")
            for item in morning_context["changed_metrics"]
            if isinstance(item, dict) and item.get("id")
        }
        morning_context["market_narrative"] = market_narrative_context(
            payload, briefing_session, recent_cards, changed_ids=changed_ids or None
        )
        morning_context["persistent_events"] = persistent_market_events(
            payload,
            previous_card.get("metric_snapshot") if isinstance(previous_card, dict) else None,
            recent_cards,
        )
        base_briefing = build_morning_briefing(payload, session, briefing_context=morning_context)
        trajectory = update_intraday_track(data_path, payload, briefing_generated_at)
        payload["morning_briefing"] = build_briefing_card(
            base_briefing,
            card_type="morning",
            generated_at=briefing_generated_at,
            generated_label=str(payload.get("generated_label") or ""),
            trajectory=trajectory,
            low_signal=False,
            session_context=briefing_session,
            gate_reason="일일 전체 빌드",
            metric_snapshot=briefing_metric_snapshot(payload),
        )
        payload["briefing_index"] = write_briefing_outputs(data_path, payload["morning_briefing"])
        try:
            process_alerts(config, payload, session)
        except Exception:  # noqa: BLE001 - 알림 실패가 배포를 막으면 안 됩니다.
            pass
        copy_signal_log_output(config, data_path)
    fetch_log_history = write_fetch_log_outputs(data_path, logger)
    payload["fetch_log_summary"] = fetch_log_history.get("runs", [])[-1].get("summary", {}) if fetch_log_history.get("runs") else {}
    write_json(data_path / "dashboard.json", payload)
    write_json(data_path / "long_history.json", long_histories, compact=True)
    write_json(data_path / MARKET_GAUGE_HISTORY_FILENAME, market_gauge_history, compact=True)
    write_dashboard_shell(output_path, payload)
    return payload


def refresh_prices_site(
    config: dict[str, Any], output_dir: str | Path, session: requests.Session
) -> dict[str, Any]:
    """장중 시세만 갱신하는 경량 빌드.

    이전 전체 빌드 결과(dashboard.json)를 기반으로 대표주/시장지수 지표만
    다시 수집해 교체하고 페이지를 재생성합니다. 이전 결과가 없으면
    전체 빌드로 폴백합니다.
    """
    output_path = Path(output_dir)
    data_path = output_path / "data"
    data_path.mkdir(parents=True, exist_ok=True)

    previous_payload = load_previous_dashboard_payload(data_path / "dashboard.json")
    if not previous_payload or not previous_payload.get("metrics"):
        return build_dashboard_site(config, output_dir, session)

    timezone_name = str(config.get("timezone") or "Asia/Seoul")
    now = datetime.now(ZoneInfo(timezone_name))
    previous_by_key = previous_metric_index(previous_payload)
    logger = FetchLogger(run_type="prices", timezone_name=timezone_name)
    with use_fetch_logger(logger):
        started_at, started_monotonic = logger.source_started()
        scoped_config, scope_label = intraday_price_config(config, now)
        fresh_metrics = collect_equity_price_metrics(
            scoped_config,
            session,
            now.date(),
            intraday_now=now,
        )
        ok_count = sum(1 for item in fresh_metrics if item.get("status") == "ok")
        message = f"{scope_label} {ok_count}/{len(fresh_metrics)}개 지표 자동 수집"
        apply_fetch_metadata(fresh_metrics, now.isoformat(timespec="seconds"), previous_by_key)
        record_fetch_result(
            "대표주가/시장지수",
            fresh_metrics,
            previous_by_key,
            started_at,
            started_monotonic,
            message,
        )
        fresh_long = enrich_metrics_with_history(fresh_metrics, config, previous_payload)

    payload = previous_payload
    metrics = payload.get("metrics", [])
    fresh_by_id = {
        str(metric.get("id")): metric
        for metric in fresh_metrics
        if metric.get("status") == "ok"
    }
    fresh_history_keys = {
        str(metric.get("history_key") or "")
        for metric in fresh_by_id.values()
        if metric.get("history_key")
    }
    metrics[:] = [
        metric
        for metric in metrics
        if str(metric.get("history_key") or "") not in RETIRED_HISTORY_KEYS
        and INTRADAY_REPLACEMENTS.get(str(metric.get("history_key") or ""))
        not in fresh_history_keys
    ]
    replaced = 0
    for index, metric in enumerate(metrics):
        fresh = fresh_by_id.get(str(metric.get("id")))
        if fresh is not None:
            # 일일 변경 배지 상태는 아침 전체 빌드 기준을 유지합니다.
            for keep_field in ("daily_status", "is_new", "is_updated_today",
                               "previous_run_observed_at", "previous_run_value",
                               "previous_run_display_value"):
                if keep_field in metric:
                    fresh[keep_field] = metric[keep_field]
            metrics[index] = fresh
            replaced += 1

    existing_ids = {str(metric.get("id") or "") for metric in metrics if isinstance(metric, dict)}
    for metric_id, fresh in fresh_by_id.items():
        if metric_id in existing_ids:
            continue
        metrics.append(fresh)
        existing_ids.add(metric_id)
        replaced += 1

    payload["generated_at"] = now.isoformat(timespec="seconds")
    payload["generated_label"] = now.strftime("%Y-%m-%d %H:%M %Z")
    payload["market_gauges"] = build_market_gauges(metrics)
    payload["calendar"] = build_event_calendar(config, payload, today=now.date())
    write_event_calendar(data_path / "calendar.json", payload["calendar"])
    attach_future_timeline_payload(config, payload, data_path, now.date(), logger, session=session)
    payload["collection_issues"] = annotate_metric_freshness(metrics, now.date())
    payload["freshness_summary"] = build_freshness_summary(
        metrics, str(payload.get("generated_at") or ""), now.date()
    )
    apply_interpretations(metrics, config)
    payload["prices_refreshed_at"] = now.isoformat(timespec="seconds")

    try:
        process_alerts(config, payload, session, include_weekly=False)
    except Exception:  # noqa: BLE001
        pass
    copy_signal_log_output(config, data_path)

    long_history_path = data_path / "long_history.json"
    long_histories: dict[str, Any] = {}
    loaded_long_histories = load_json(long_history_path, None)
    if isinstance(loaded_long_histories, dict):
        long_histories = loaded_long_histories
    long_histories.update(fresh_long)

    fetch_log_history = write_fetch_log_outputs(data_path, logger)
    payload["fetch_log_summary"] = fetch_log_history.get("runs", [])[-1].get("summary", {}) if fetch_log_history.get("runs") else {}
    write_json(data_path / "dashboard.json", payload)
    write_json(long_history_path, long_histories, compact=True)
    write_dashboard_shell(output_path, payload)
    payload["_prices_refreshed_count"] = replaced
    return payload


def write_briefing_site_outputs(
    output_path: Path,
    data_path: Path,
    payload: dict[str, Any],
    logger: FetchLogger,
) -> None:
    fetch_log_history = write_fetch_log_outputs(data_path, logger)
    payload["fetch_log_summary"] = fetch_log_history.get("runs", [])[-1].get("summary", {}) if fetch_log_history.get("runs") else {}
    write_json(data_path / "dashboard.json", payload)
    write_dashboard_shell(output_path, payload)


def refresh_briefing_site(
    config: dict[str, Any],
    output_dir: str | Path,
    session: requests.Session,
    card_type: str,
) -> dict[str, Any]:
    """Generate a briefing card from the latest published dashboard payload."""
    output_path = Path(output_dir)
    data_path = output_path / "data"
    data_path.mkdir(parents=True, exist_ok=True)

    payload = load_previous_dashboard_payload(data_path / "dashboard.json")
    if not payload or not payload.get("metrics"):
        return build_dashboard_site(config, output_dir, session)

    timezone_name = str(config.get("timezone") or payload.get("timezone") or "Asia/Seoul")
    now = datetime.now(ZoneInfo(timezone_name))
    logger = FetchLogger(run_type=f"briefing-{card_type}", timezone_name=timezone_name)
    with use_fetch_logger(logger):
        started_at, started_monotonic = logger.source_started()
        previous_card = load_latest_briefing_card(data_path)
        recent_cards = load_recent_briefing_cards(data_path, limit=4)
        decision = briefing_generation_decision(payload, previous_card, recent_cards, now)
        payload["briefing_index"] = load_briefing_index(data_path)
        if decision["skip"]:
            if previous_card:
                payload["morning_briefing"] = previous_card
            attach_future_timeline_payload(config, payload, data_path, now.date(), logger, session=session)
            logger.record(
                source="AI 요약",
                endpoint=GEMINI_GENERATE_URL.format(model=os.getenv("GEMINI_MODEL", GEMINI_DEFAULT_MODEL)),
                status="no_new_data",
                message=f"카드 생성 스킵: {decision['reason']}",
                metric_count=int((decision.get("changes") or {}).get("changed_count") or 0),
                new_data_count=0,
                started_at=started_at,
                started_monotonic=started_monotonic,
            )
            write_briefing_site_outputs(output_path, data_path, payload, logger)
            return payload

    payload["generated_at"] = now.isoformat(timespec="seconds")
    payload["generated_label"] = now.strftime("%Y-%m-%d %H:%M %Z")
    usage = load_gemini_usage(data_path, briefing_date_key(payload["generated_at"]))
    gemini_allowed = int(usage.get("count") or 0) < GEMINI_DAILY_CALL_LIMIT
    gemini_guard_message = ""
    if not gemini_allowed:
        gemini_guard_message = f"Gemini 일일 호출 {GEMINI_DAILY_CALL_LIMIT}회 도달; 룰 기반 폴백"
    briefing_context = {
        "card_type": card_type,
        "session": decision["session"],
        "low_signal": bool(decision["low_signal"]),
        "reason": decision["reason"],
        "significant": bool(decision["significant"]),
        "benchmark_drivers": (decision.get("benchmark") or {}).get("drivers", []),
        "daily_move_drivers": (decision.get("daily_move") or {}).get("drivers", []),
        "changed_metrics": (decision.get("changes") or {}).get("changes", []),
        "persistent_events": persistent_market_events(
            payload,
            previous_card.get("metric_snapshot") if isinstance(previous_card, dict) else None,
            recent_cards,
        ),
        "recent_topic_history": [
            {
                "generated_at": str(card.get("generated_at") or ""),
                "card_type": str(card.get("card_type") or ""),
                "narrative_topics": card.get("narrative_topics") or [],
            }
            for card in recent_cards[:6]
            if isinstance(card, dict)
        ],
        "gemini_daily_count_before": int(usage.get("count") or 0),
    }
    if card_type in {"close", "us_close"}:
        briefing_context["day_card_flow"] = [
            {
                "generated_at": str(card.get("generated_at") or ""),
                "card_type": str(card.get("card_type") or ""),
                "headline": str(card.get("headline") or ""),
                "narrative_topics": card.get("narrative_topics") or [],
            }
            for card in recent_cards[:6]
            if isinstance(card, dict)
        ]
    changed_ids = {
        str(metric_id)
        for metric_id in (decision.get("changes") or {}).get("changed_ids", [])
        if metric_id
    }
    briefing_context["market_narrative"] = market_narrative_context(
        payload,
        decision["session"],
        recent_cards,
        changed_ids=changed_ids or None,
    )
    base_briefing = build_morning_briefing(
        payload,
        session,
        briefing_context=briefing_context,
        gemini_allowed=gemini_allowed,
        disabled_message=gemini_guard_message or None,
    )
    if base_briefing.get("gemini_call_attempted"):
        usage = increment_gemini_usage(data_path, payload["generated_at"])
        base_briefing["gemini_daily_count"] = int(usage.get("count") or 0)
    trajectory = update_intraday_track(data_path, payload, payload["generated_at"])
    card = build_briefing_card(
        base_briefing,
        card_type=card_type,
        generated_at=payload["generated_at"],
        generated_label=payload["generated_label"],
        trajectory=trajectory,
        low_signal=bool(decision["low_signal"]),
        session_context=decision["session"],
        gate_reason=decision["reason"],
        metric_snapshot=decision["metric_snapshot"],
    )
    payload["morning_briefing"] = card
    payload["briefing_index"] = write_briefing_outputs(data_path, card)
    attach_future_timeline_payload(config, payload, data_path, now.date(), logger, session=session)

    status_detail = "Gemini 호출" if card.get("gemini_call_attempted") else "룰 기반 폴백"
    if gemini_guard_message:
        status_detail = gemini_guard_message
    logger.record(
        source="AI 요약",
        endpoint=GEMINI_GENERATE_URL.format(model=os.getenv("GEMINI_MODEL", GEMINI_DEFAULT_MODEL)),
        status="success",
        message=f"카드 생성: {decision['reason']} ({status_detail})",
        metric_count=int((decision.get("changes") or {}).get("changed_count") or 0),
        new_data_count=1,
        started_at=started_at,
        started_monotonic=started_monotonic,
    )
    write_briefing_site_outputs(output_path, data_path, payload, logger)
    return payload


STALE_THRESHOLD_DAYS = [
    ("장중", 2),
    ("일간", 7),
    ("주간", 21),
    # observed_at is the period start, not the release date. Monthly and
    # quarterly series therefore need enough room for normal publication lag.
    ("월간", 105),
    ("분기", 230),
    ("연간", 550),
]

FRESHNESS_SOURCE_AGE_OVERRIDES = {
    "EIA Open Data API": 145,
    "SEC Company Facts API": 245,
    "WSTS Historical Billings Report": 130,
}

INTRADAY_REPLACEMENTS = {
    "fred-DEXKOUS": "equity-KRW=X",
    "fred-VIXCLS": "equity-^VIX",
}

RETIRED_HISTORY_KEYS = {
    "multpl-s-p-500-pe-ratio",
}


def metric_stale_after_days(metric: dict[str, Any]) -> int | None:
    source = str(metric.get("source") or "")
    if source in FRESHNESS_SOURCE_AGE_OVERRIDES:
        return FRESHNESS_SOURCE_AGE_OVERRIDES[source]
    frequency = str(metric.get("frequency") or "")
    return next((days for token, days in STALE_THRESHOLD_DAYS if token in frequency), None)


def annotate_metric_freshness(metrics: list[dict[str, Any]], today: date) -> list[dict[str, Any]]:
    """지표 주기 대비 오래된 데이터에 is_stale 표시를 하고 지연 목록을 반환합니다."""
    stale_items: list[dict[str, Any]] = []
    for metric in metrics:
        metric["is_stale"] = False
        if metric.get("status") != "ok":
            continue
        observed = parse_iso_date(metric.get("observed_at"))
        if observed is None:
            continue
        threshold = metric_stale_after_days(metric)
        if threshold is None:
            continue
        age = (today - observed).days
        metric["freshness_age_days"] = max(0, age)
        metric["freshness_due_at"] = (observed + timedelta(days=threshold)).isoformat()
        if age > threshold:
            metric["is_stale"] = True
            metric["stale_days"] = age
            stale_items.append(
                {
                    "id": metric.get("id"),
                    "name": metric.get("name"),
                    "observed_at": metric.get("observed_at"),
                    "days": age,
                }
            )
    return stale_items


def build_freshness_summary(
    metrics: list[dict[str, Any]], generated_at: str, today: date
) -> dict[str, Any]:
    """Compact user-facing health summary, separate from the admin fetch log."""
    sources: dict[str, dict[str, Any]] = {}
    totals = {
        "total_count": 0,
        "current_count": 0,
        "updated_count": 0,
        "waiting_count": 0,
        "delayed_count": 0,
        "failed_count": 0,
    }
    last_checked_at = ""
    latest_observed_at = ""

    for metric in metrics:
        if not isinstance(metric, dict):
            continue
        totals["total_count"] += 1
        fetch_status = str(metric.get("fetch_status") or "")
        # Some slow official APIs can time out while a valid cached metric is
        # still shown. Keep that visible in source_status, but do not make the
        # user-facing freshness summary look like the data disappeared.
        displayable_fallback = (
            fetch_status == "failed"
            and metric.get("status") == "ok"
            and metric.get("value") is not None
            and bool(metric.get("observed_at"))
        )
        failed = (fetch_status == "failed" or metric.get("status") != "ok") and not displayable_fallback
        delayed = bool(metric.get("is_stale")) and not failed
        if failed:
            state = "failed"
            totals["failed_count"] += 1
        elif delayed:
            state = "delayed"
            totals["delayed_count"] += 1
        else:
            state = "current"
            totals["current_count"] += 1
        if fetch_status == "success":
            totals["updated_count"] += 1
        elif fetch_status == "no_new_data":
            totals["waiting_count"] += 1

        checked_at = str(metric.get("fetched_at") or "")
        observed_at = str(metric.get("observed_at") or "")[:10]
        last_checked_at = max(last_checked_at, checked_at)
        latest_observed_at = max(latest_observed_at, observed_at)
        source_name = str(metric.get("source") or "기타")
        source = sources.setdefault(
            source_name,
            {
                "name": source_name,
                "total_count": 0,
                "current_count": 0,
                "updated_count": 0,
                "waiting_count": 0,
                "delayed_count": 0,
                "failed_count": 0,
                "last_checked_at": "",
                "latest_observed_at": "",
            },
        )
        source["total_count"] += 1
        source[f"{state}_count"] += 1
        if fetch_status == "success":
            source["updated_count"] += 1
        elif fetch_status == "no_new_data":
            source["waiting_count"] += 1
        source["last_checked_at"] = max(str(source["last_checked_at"]), checked_at)
        source["latest_observed_at"] = max(str(source["latest_observed_at"]), observed_at)

    source_rows = sorted(
        sources.values(),
        key=lambda item: (
            -int(item.get("failed_count") or 0),
            -int(item.get("delayed_count") or 0),
            str(item.get("name") or ""),
        ),
    )
    status = "failed" if totals["failed_count"] else "delayed" if totals["delayed_count"] else "current"
    return {
        "generated_at": generated_at,
        "last_checked_at": last_checked_at or generated_at,
        "latest_observed_at": latest_observed_at,
        "status": status,
        "as_of_date": today.isoformat(),
        **totals,
        "sources": source_rows,
    }


def attach_history_store(config: dict[str, Any]) -> HistoryStore | None:
    """config에 히스토리 저장소를 1회 생성해 붙입니다. 수집기와 enrichment가 공유합니다."""
    history_config = config.get("history", {}) or {}
    if not history_config.get("enabled", True):
        return None
    store = config.get("_history_store")
    if not isinstance(store, HistoryStore):
        store = HistoryStore(str(history_config.get("dir") or "data/history"))
        config["_history_store"] = store
    return store


def cached_history_last_date(config: dict[str, Any], key: str) -> date | None:
    """캐시된 마지막 관측일. None이면 최초 백필(전체 기간 요청)이 필요합니다."""
    store = attach_history_store(config)
    if store is None:
        return None
    series = store.series(key)
    return series[-1][0] if series else None


def enrich_metrics_with_history(
    metrics: list[dict[str, Any]],
    config: dict[str, Any],
    previous_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """지표를 장기 히스토리 캐시와 병합하고 백분위 통계와 장기 시리즈를 만듭니다.

    캐시는 repo에 커밋되는 data/history 밑 JSON 파일이므로, 최초 백필 이후에는
    수집기가 최근 구간만 가져와도 전체 기간 시계열이 유지됩니다.
    """
    history_config = config.get("history", {}) or {}
    display_points = int(history_config.get("display_points") or 60)
    store = attach_history_store(config)
    if store is None:
        # 축적이 꺼져 있으면 페이지 무게 보호를 위해 표시 포인트만 남깁니다.
        for metric in metrics:
            history = metric.get("history")
            if isinstance(history, list) and len(history) > display_points:
                metric["history"] = history[-display_points:]
        return {}

    max_long_points = int(history_config.get("long_history_max_points") or 480)

    previous_metrics = (previous_payload or {}).get("metrics", [])
    previous_history_by_id = {
        str(item.get("id")): item.get("history")
        for item in previous_metrics
        if isinstance(item, dict) and item.get("id")
    }

    long_histories: dict[str, Any] = {}
    for metric in metrics:
        if metric.get("status") != "ok":
            continue
        points = parse_stored_points(metric.get("history"))
        if not points:
            continue

        key = str(metric.get("history_key") or "") or f"id-{metric.get('id')}"
        merge_mode = str(metric.get("history_merge") or "full")

        # 이전 배포본의 히스토리를 먼저 병합해 캐시 최초 구축 시 데이터 공백을 줄입니다.
        previous_points = parse_stored_points(previous_history_by_id.get(str(metric.get("id"))))
        if previous_points:
            store.merge(key, previous_points, mode="full")

        full = store.merge(
            key,
            points,
            name=str(metric.get("name") or ""),
            unit=str(metric.get("unit") or ""),
            source=str(metric.get("source") or ""),
            mode=merge_mode,
        )
        if not full:
            continue

        # 스냅샷형 소스는 축적 초기에 캐시가 더 짧을 수 있으므로 긴 쪽을 표시합니다.
        if len(full) >= len(points):
            metric["history"] = [
                {"date": point_date.isoformat(), "value": value}
                for point_date, value in full[-display_points:]
            ]
            metric["period_label"] = period_label(
                metric["history"], str(metric.get("observed_at") or "")
            )

        stats = percentile_stats(full, to_float(metric.get("value")))
        if stats:
            metric["percentiles"] = stats

        if key == US_NET_LIQUIDITY_KEY:
            analysis_cutoff = full[-1][0] - timedelta(days=370)
            metric["_analysis_history"] = [
                {"date": point_date.isoformat(), "value": value}
                for point_date, value in full
                if point_date >= analysis_cutoff
            ]

        if metric.get("yoy_pct") is None and metric.get("value") is not None:
            yoy_value = find_yoy_value(full, full[-1][0])
            yoy_pct = pct_change(to_float(metric.get("value")), yoy_value)
            if yoy_pct is not None:
                metric["yoy_pct"] = yoy_pct
                metric["yoy_pct_label"] = fmt_pct(yoy_pct)

        sampled = downsample_history(full, max_points=max_long_points)
        if len(sampled) > display_points:
            long_histories[str(metric["id"])] = {
                "key": key,
                "unit": str(metric.get("unit") or ""),
                "points": [
                    [point_date.isoformat(), value] for point_date, value in sampled
                ],
            }

    store.save_all()
    return long_histories


def load_previous_dashboard_payload(path: Path) -> dict[str, Any] | None:
    payload = load_json(path, None)
    return payload if isinstance(payload, dict) else None


def metric_identity(metric: dict[str, Any]) -> str:
    return str(metric.get("history_key") or metric.get("id") or "")


def previous_metric_index(previous_payload: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    if not isinstance(previous_payload, dict):
        return {}
    metrics = previous_payload.get("metrics")
    if not isinstance(metrics, list):
        return {}
    indexed: dict[str, dict[str, Any]] = {}
    for metric in metrics:
        if not isinstance(metric, dict):
            continue
        key = metric_identity(metric)
        if key:
            indexed[key] = metric
        metric_id = str(metric.get("id") or "")
        if metric_id:
            indexed[metric_id] = metric
    return indexed


def metric_observed_advanced(metric: dict[str, Any], previous: dict[str, Any] | None) -> bool:
    if metric.get("status") != "ok":
        return False
    observed_at = str(metric.get("observed_at") or "")
    if not observed_at:
        return False
    if not previous:
        return True
    previous_observed_at = str(previous.get("observed_at") or "")
    if not previous_observed_at:
        return True
    observed_date = parse_iso_date(observed_at)
    previous_date = parse_iso_date(previous_observed_at)
    if observed_date is not None and previous_date is not None:
        return observed_date > previous_date
    return observed_at != previous_observed_at


def fetch_status_label(status: str) -> str:
    return {
        "success": "업데이트됨",
        "no_new_data": "신규 데이터 없음",
        "failed": "수집 실패",
    }.get(status, status)


def infer_metric_fetch_status(metric: dict[str, Any], previous: dict[str, Any] | None) -> str:
    if metric.get("fetch_attempt_failed"):
        return "failed"
    if metric.get("status") != "ok":
        return "failed"
    if metric_observed_advanced(metric, previous):
        return "success"
    if previous is not None and metric_was_updated(metric, previous):
        return "success"
    return "no_new_data"


def apply_fetch_metadata(
    metrics: list[dict[str, Any]],
    fetched_at: str,
    previous_by_key: dict[str, dict[str, Any]],
) -> None:
    for metric in metrics:
        if not isinstance(metric, dict):
            continue
        previous = previous_by_key.get(metric_identity(metric)) or previous_by_key.get(str(metric.get("id") or ""))
        status = infer_metric_fetch_status(metric, previous)
        metric["fetched_at"] = fetched_at
        metric["fetch_status"] = status
        metric["fetch_status_label"] = fetch_status_label(status)


def source_new_data_count(
    metrics: list[dict[str, Any]],
    previous_by_key: dict[str, dict[str, Any]],
) -> int:
    count = 0
    for metric in metrics:
        if not isinstance(metric, dict):
            continue
        previous = previous_by_key.get(metric_identity(metric)) or previous_by_key.get(str(metric.get("id") or ""))
        if metric_observed_advanced(metric, previous) or (
            previous is not None and metric_was_updated(metric, previous)
        ):
            count += 1
    return count


def record_fetch_result(
    source_name: str,
    source_metrics: list[dict[str, Any]],
    previous_by_key: dict[str, dict[str, Any]],
    started_at: str,
    started_monotonic: float,
    message: str,
) -> None:
    logger = current_logger()
    if logger is None:
        return
    ok_count = sum(1 for item in source_metrics if isinstance(item, dict) and item.get("status") == "ok")
    attempt_failed = any(
        isinstance(item, dict) and item.get("fetch_attempt_failed") for item in source_metrics
    )
    new_count = source_new_data_count(source_metrics, previous_by_key)
    if not source_metrics:
        status = "no_new_data"
    elif ok_count <= 0:
        status = "failed"
    elif attempt_failed:
        status = "failed"
    elif new_count > 0:
        status = "success"
    else:
        status = "no_new_data"
    logger.record(
        source=source_name,
        endpoint=FETCH_SOURCE_ENDPOINTS.get(source_name, ""),
        status=status,
        message=message,
        metric_count=len(source_metrics),
        new_data_count=new_count,
        started_at=started_at,
        started_monotonic=started_monotonic,
    )


def record_fetch_failure(
    source_name: str,
    started_at: str,
    started_monotonic: float,
    exc: Exception,
) -> None:
    logger = current_logger()
    if logger is None:
        return
    logger.record(
        source=source_name,
        endpoint=FETCH_SOURCE_ENDPOINTS.get(source_name, ""),
        status="failed",
        message=str(exc),
        metric_count=0,
        new_data_count=0,
        started_at=started_at,
        started_monotonic=started_monotonic,
    )


def upsert_source_status(payload: dict[str, Any], name: str, status: str, message: str) -> None:
    records = [item for item in payload.get("source_status", []) if isinstance(item, dict)]
    next_record = {"name": name, "status": status, "message": sanitize_message(message)}
    replaced = False
    for index, record in enumerate(records):
        if str(record.get("name") or "") == name:
            records[index] = next_record
            replaced = True
            break
    if not replaced:
        records.append(next_record)
    payload["source_status"] = records


def attach_future_timeline_payload(
    config: dict[str, Any],
    payload: dict[str, Any],
    data_path: Path,
    today: date,
    logger: FetchLogger | None = None,
    session: requests.Session | None = None,
) -> dict[str, Any]:
    started_at, started_monotonic = logger.source_started() if logger else ("", time.monotonic())
    document = build_future_timeline(config, payload.get("metrics", []), today=today, session=session)
    payload["future"] = document
    write_future_timeline(data_path / "future.json", document)

    warnings = document.get("warnings") or []
    tech_count = int((document.get("summary") or {}).get("technology_count") or 0)
    if warnings:
        message = f"{tech_count}개 기술 타임라인 생성, YAML 경고 {len(warnings)}건"
        source_status = "partial"
        log_status = "no_new_data"
    else:
        message = f"{tech_count}개 기술 타임라인 생성"
        source_status = "ok"
        log_status = "success"
    upsert_source_status(payload, "미래 타임라인", source_status, message)
    if logger:
        logger.record(
            source="미래 타임라인",
            endpoint=FETCH_SOURCE_ENDPOINTS.get("미래 타임라인", ""),
            status=log_status,
            message=message,
            metric_count=tech_count,
            new_data_count=0,
            started_at=started_at,
            started_monotonic=started_monotonic,
        )
    return document


def write_fetch_log_outputs(data_path: Path, logger: FetchLogger) -> dict[str, Any]:
    run = logger.finish()
    history = append_fetch_log_run(
        load_fetch_log_history(data_path / FETCH_LOG_HISTORY_FILENAME),
        run,
    )
    save_fetch_log_history(data_path / FETCH_LOG_HISTORY_FILENAME, history)
    save_fetch_log_history(data_path / FETCH_LOG_FILENAME, history)
    return history


def market_gauge_snapshot_date(payload: dict[str, Any], timezone_name: str) -> str:
    generated_at = str(payload.get("generated_at") or "")
    try:
        instant = datetime.fromisoformat(generated_at.replace("Z", "+00:00"))
        if instant.tzinfo is None:
            instant = instant.replace(tzinfo=ZoneInfo(timezone_name))
        return instant.astimezone(ZoneInfo(timezone_name)).date().isoformat()
    except (TypeError, ValueError):
        return datetime.now(ZoneInfo(timezone_name)).date().isoformat()


def compact_gauge_items(items: Any, fields: tuple[str, ...]) -> list[dict[str, Any]]:
    if not isinstance(items, list):
        return []
    compacted: list[dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        compacted_item = {
            field: item[field]
            for field in fields
            if field in item and item[field] is not None and item[field] != ""
        }
        if compacted_item:
            compacted.append(compacted_item)
    return compacted


def market_gauge_history_snapshot(
    payload: dict[str, Any], timezone_name: str
) -> dict[str, Any] | None:
    gauges = payload.get("market_gauges")
    if not isinstance(gauges, dict) or not gauges:
        return None

    snapshot: dict[str, Any] = {
        "date": market_gauge_snapshot_date(payload, timezone_name),
        "generated_at": str(payload.get("generated_at") or ""),
    }
    thermometer = gauges.get("thermometer")
    if isinstance(thermometer, dict):
        snapshot["thermometer"] = {
            field: thermometer[field]
            for field in ("score", "label", "comment")
            if field in thermometer and thermometer[field] is not None and thermometer[field] != ""
        }
        snapshot["thermometer"]["components"] = compact_gauge_items(
            thermometer.get("components"),
            ("name", "metric_id", "metric_name", "value_label", "percentile", "heat", "basis"),
        )

    recession = gauges.get("recession")
    if isinstance(recession, dict):
        snapshot["recession"] = {
            field: recession[field]
            for field in ("alert_count", "warn_count", "summary")
            if field in recession and recession[field] is not None and recession[field] != ""
        }
        snapshot["recession"]["signals"] = compact_gauge_items(
            recession.get("signals"),
            ("name", "metric_id", "value_label", "status", "description"),
        )

    fear_greed = gauges.get("fear_greed")
    if isinstance(fear_greed, dict):
        snapshot["fear_greed"] = {
            field: fear_greed[field]
            for field in ("comment",)
            if field in fear_greed and fear_greed[field] is not None and fear_greed[field] != ""
        }
        snapshot["fear_greed"]["items"] = compact_gauge_items(
            fear_greed.get("items"),
            ("name", "metric_id", "metric_name", "score", "label", "value_label"),
        )

    if "thermometer" not in snapshot and "recession" not in snapshot and "fear_greed" not in snapshot:
        return None
    return snapshot


def update_market_gauge_history(
    previous_history: dict[str, Any] | None,
    payload: dict[str, Any],
    timezone_name: str,
) -> dict[str, Any]:
    existing = previous_history if isinstance(previous_history, dict) else {}
    raw_snapshots = existing.get("snapshots")
    if not isinstance(raw_snapshots, list):
        raw_snapshots = []
    snapshots = [
        snapshot
        for snapshot in raw_snapshots
        if isinstance(snapshot, dict) and snapshot.get("date")
    ]
    current = market_gauge_history_snapshot(payload, timezone_name)
    if current:
        current_date = str(current["date"])
        snapshots = [snapshot for snapshot in snapshots if str(snapshot.get("date")) != current_date]
        snapshots.append(current)
    snapshots.sort(key=lambda snapshot: str(snapshot.get("date") or ""))

    document: dict[str, Any] = {
        "version": MARKET_GAUGE_HISTORY_VERSION,
        "updated_at": str((current or {}).get("generated_at") or existing.get("updated_at") or ""),
        "count": len(snapshots),
        "snapshots": snapshots,
    }
    if snapshots:
        document["first_date"] = str(snapshots[0].get("date") or "")
        document["last_date"] = str(snapshots[-1].get("date") or "")
    return document


def annotate_dashboard_updates(payload: dict[str, Any], previous_payload: dict[str, Any] | None) -> None:
    metrics = payload.get("metrics", [])
    if not isinstance(metrics, list):
        payload["daily_changes"] = empty_daily_changes(payload)
        return

    previous_metrics = previous_payload.get("metrics", []) if isinstance(previous_payload, dict) else []
    previous_by_id = {
        str(metric.get("id")): metric
        for metric in previous_metrics
        if isinstance(metric, dict) and metric.get("id")
    }
    changed_metrics: list[dict[str, Any]] = []

    for metric in metrics:
        if not isinstance(metric, dict):
            continue
        previous = previous_by_id.get(str(metric.get("id") or ""))
        status = ""
        if previous_payload and previous is None:
            status = "new"
        elif previous_payload and previous is not None and metric_was_updated(metric, previous):
            status = "updated"

        metric["daily_status"] = status
        metric["is_new"] = status == "new"
        metric["is_updated_today"] = status == "updated"
        if previous:
            metric["previous_run_observed_at"] = previous.get("observed_at", "")
            metric["previous_run_value"] = previous.get("value")
            metric["previous_run_display_value"] = previous.get("display_value", "")
        if status:
            changed_metrics.append(daily_change_item(metric, previous, status))

    payload["daily_changes"] = {
        "date": str(payload.get("generated_at") or "")[:10],
        "has_previous": bool(previous_payload),
        "updated_count": sum(1 for item in changed_metrics if item["status"] == "updated"),
        "new_count": sum(1 for item in changed_metrics if item["status"] == "new"),
        "metrics": changed_metrics,
    }


def empty_daily_changes(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "date": str(payload.get("generated_at") or "")[:10],
        "has_previous": False,
        "updated_count": 0,
        "new_count": 0,
        "metrics": [],
    }


def metric_was_updated(metric: dict[str, Any], previous: dict[str, Any]) -> bool:
    if str(metric.get("observed_at") or "") != str(previous.get("observed_at") or ""):
        return True
    if not numeric_values_equal(metric.get("value"), previous.get("value")):
        return True

    current_latest = latest_history_point(metric)
    previous_latest = latest_history_point(previous)
    if current_latest or previous_latest:
        if str((current_latest or {}).get("date") or "") != str((previous_latest or {}).get("date") or ""):
            return True
        if not numeric_values_equal(
            (current_latest or {}).get("value"),
            (previous_latest or {}).get("value"),
        ):
            return True
    return False


def latest_history_point(metric: dict[str, Any]) -> dict[str, Any] | None:
    history = metric.get("history")
    if not isinstance(history, list) or not history:
        return None
    latest = history[-1]
    return latest if isinstance(latest, dict) else None


def daily_change_item(
    metric: dict[str, Any], previous: dict[str, Any] | None, status: str
) -> dict[str, Any]:
    return {
        "id": metric.get("id", ""),
        "status": status,
        "industry": metric.get("industry", ""),
        "industry_en": metric.get("industry_en", ""),
        "group": metric.get("group", ""),
        "group_en": metric.get("group_en", ""),
        "name": metric.get("name", ""),
        "name_en": metric.get("name_en", ""),
        "display_value": metric.get("display_value", ""),
        "observed_label": metric.get("observed_label", ""),
        "change_pct_label": metric.get("change_pct_label", ""),
        "previous_display_value": previous.get("display_value", "") if previous else "",
        "previous_observed_label": previous.get("observed_label", "") if previous else "",
    }


def run_dashboard_collector(
    *,
    source_name: str,
    collector: Callable[[], list[dict[str, Any]]],
    metrics: list[dict[str, Any]],
    source_status: list[dict[str, str]],
    previous_by_key: dict[str, dict[str, Any]],
    fetched_at: str,
    status_requires_nonempty: bool = True,
    status_requires_clean_metrics: bool = True,
    include_issue_summary: bool = True,
    error_metric_overrides: dict[str, Any] | None = None,
) -> None:
    """Run one source without allowing its failure to stop the dashboard build."""
    before = len(metrics)
    logger = current_logger()
    started_at, started_monotonic = (
        logger.source_started() if logger else (fetched_at, time.monotonic())
    )
    try:
        source_metrics = collector()
        metrics.extend(source_metrics)
        ok_count = sum(1 for item in source_metrics if item.get("status") == "ok")
        issue_summary = metric_issue_summary(source_metrics) if include_issue_summary else ""
        message = f"{ok_count}/{len(source_metrics)}개 지표 자동 수집"
        if issue_summary:
            message = f"{message} ({issue_summary})"

        complete = ok_count == len(source_metrics)
        if status_requires_nonempty:
            complete = complete and bool(source_metrics)
        if status_requires_clean_metrics:
            complete = complete and not issue_summary
        source_status.append(
            {
                "name": source_name,
                "status": "ok" if complete else "partial",
                "message": sanitize_message(message),
            }
        )
        record_fetch_result(
            source_name,
            source_metrics,
            previous_by_key,
            started_at,
            started_monotonic,
            message,
        )
    except Exception as exc:  # noqa: BLE001 - source failures are isolated by design.
        source_status.append(
            {
                "name": source_name,
                "status": "error",
                "message": sanitize_message(str(exc)),
            }
        )
        record_fetch_failure(source_name, started_at, started_monotonic, exc)
        if len(metrics) != before:
            return

        error_metric = {
            "industry": "매크로",
            "name": f"{source_name} 수집 상태",
            "source": source_name,
            "source_url": "",
            "frequency": "",
            "automation": "부분 자동화 가능",
            "status": "error",
            "note": str(exc),
        }
        error_metric.update(error_metric_overrides or {})
        metrics.append(make_metric(**error_metric))


def build_dashboard_payload(
    config: dict[str, Any],
    session: requests.Session,
    previous_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    timezone = str(config.get("timezone") or "Asia/Seoul")
    now = datetime.now(ZoneInfo(timezone))
    fetched_at = now.isoformat(timespec="seconds")
    previous_by_key = previous_metric_index(previous_payload)

    source_status: list[dict[str, str]] = []
    metrics: list[dict[str, Any]] = []

    collectors = [
        ("WSTS", collect_wsts_metrics),
        ("FRED", collect_fred_metrics),
        ("ECOS 신용스프레드", collect_ecos_credit_spread_metrics),
        ("ECOS 매크로", collect_ecos_series_metrics),
        ("금투협회 증시자금", collect_kofia_capital_market_metrics),
        ("대표주가/시장지수", collect_equity_price_metrics),
        ("스테이블코인", collect_stablecoin_metrics),
        ("World Bank 원자재", collect_world_bank_commodity_metrics),
        ("SEC CAPEX", collect_sec_capex_metrics),
        ("USAspending 방산", collect_usaspending_metrics),
        ("EIA", collect_eia_metrics),
        ("openFDA", collect_openfda_metrics),
        ("ClinicalTrials.gov", collect_clinical_trials_metrics),
        ("Launch Library", collect_launch_library_metrics),
        ("AFDC EV 충전", collect_afdc_metrics),
        ("KOSIS", collect_kosis_metrics),
        ("한국 수출", collect_korea_export_metrics),
        ("밸류에이션/수급", collect_valuation_metrics),
        ("시장 수급", collect_market_flow_metrics),
    ]
    for source_name, collector in collectors:
        run_dashboard_collector(
            source_name=source_name,
            collector=lambda collector=collector: collector(config, session, now.date()),
            metrics=metrics,
            source_status=source_status,
            previous_by_key=previous_by_key,
            fetched_at=fetched_at,
        )

    run_dashboard_collector(
        source_name="미국 유동성",
        collector=lambda: collect_us_liquidity_metrics(config, session, now.date(), metrics),
        metrics=metrics,
        source_status=source_status,
        previous_by_key=previous_by_key,
        fetched_at=fetched_at,
        status_requires_clean_metrics=False,
        error_metric_overrides={
            "automation": "무료로 안정적으로 자동화 가능",
            "group": "미국 유동성",
            "section": "market",
            "market_category": US_LIQUIDITY_CATEGORY,
        },
    )
    run_dashboard_collector(
        source_name=GLOBAL_LIQUIDITY_SOURCE_NAME,
        collector=lambda: collect_global_liquidity_metrics(config, session, now.date()),
        metrics=metrics,
        source_status=source_status,
        previous_by_key=previous_by_key,
        fetched_at=fetched_at,
        status_requires_clean_metrics=False,
        error_metric_overrides={
            "name": "국제 유동성 수집 상태",
            "automation": "무료 공식 API 자동 수집",
            "group": GLOBAL_LIQUIDITY_SOURCE_NAME,
            "section": "market",
            "market_category": US_LIQUIDITY_CATEGORY,
        },
    )
    run_dashboard_collector(
        source_name="시장 심리",
        collector=lambda: collect_market_sentiment_metrics(config, session, now.date(), metrics),
        metrics=metrics,
        source_status=source_status,
        previous_by_key=previous_by_key,
        fetched_at=fetched_at,
        status_requires_clean_metrics=False,
        error_metric_overrides={"group": "공포탐욕"},
    )
    run_dashboard_collector(
        source_name="시장 파생지표",
        collector=lambda: collect_market_derived_metrics(config, session, now.date(), metrics),
        metrics=metrics,
        source_status=source_status,
        previous_by_key=previous_by_key,
        fetched_at=fetched_at,
        status_requires_nonempty=False,
        status_requires_clean_metrics=False,
        include_issue_summary=False,
        error_metric_overrides={
            "group": "시장 분위기",
            "section": "market",
            "market_category": "심리·변동성",
        },
    )

    metrics.extend(collect_reference_metrics(config))
    apply_fetch_metadata(metrics, fetched_at, previous_by_key)
    assign_market_navigation_fields(metrics)
    assign_metric_country_fields(metrics)
    metrics = visible_dashboard_metrics(metrics)
    industries = configured_industries(config, metrics)

    return {
        "title": "산업별 지표 대시보드",
        "generated_at": now.isoformat(timespec="seconds"),
        "generated_label": now.strftime("%Y-%m-%d %H:%M %Z"),
        "timezone": timezone,
        "industries": industries,
        "industry_labels_en": {industry: english_industry(industry) for industry in industries},
        "industry_icons": INDUSTRY_ICONS,
        "search_aliases": config.get("search_aliases", {}) or {},
        "source_status": source_status,
        "metrics": metrics,
    }


def metric_issue_summary(metrics: list[dict[str, Any]], limit: int = 3) -> str:
    issues: list[str] = []
    warning_tokens = ("이전 저장값 표시", "응답 실패", "관측값 없음", "API_KEY 없음")
    for metric in metrics:
        note = str(metric.get("note") or metric.get("status_label") or metric.get("status") or "").strip()
        has_warning_note = any(token in note for token in warning_tokens)
        if metric.get("status") == "ok" and not has_warning_note:
            continue
        name = str(metric.get("name") or "지표")
        compact = f"{name}: {note}" if note else name
        if compact not in issues:
            issues.append(compact)
    if not issues:
        return ""
    shown = issues[:limit]
    suffix = f" 외 {len(issues) - limit}개" if len(issues) > limit else ""
    return " / ".join(shown) + suffix


def collect_fred_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    dashboard_config = config.get("dashboard", {})
    fred_config = config.get("fred", {})
    if not fred_config.get("enabled", True):
        return []

    series_config = dashboard_config.get("fred_series") or fred_config.get("series", [])
    api_key = os.getenv("FRED_API_KEY", "").strip()

    if not api_key:
        return [
            make_metric(
                industry=str(series.get("industry") or "매크로"),
                name=str(series.get("name") or series.get("id")),
                source="FRED API",
                source_url=f"https://fred.stlouisfed.org/series/{str(series.get('id', '')).strip()}",
                frequency=str(series.get("frequency") or "FRED"),
                automation="무료로 안정적으로 자동화 가능",
                status="needs_key",
                note="GitHub Secrets에 FRED_API_KEY 등록 필요",
                group=str(series.get("group") or ""),
                depth=str(series.get("depth") or ""),
                meaning=str(series.get("meaning") or ""),
                section=str(series.get("section") or ""),
                market_category=str(series.get("market_category") or ""),
                also_market_category=series.get("also_market_category") or "",
            )
            for series in series_config
            if series.get("id")
        ]

    metrics: list[dict[str, Any]] = []
    for series in series_config:
        series_id = str(series.get("id", "")).strip()
        if not series_id:
            continue

        name = str(series.get("name") or series_id)
        unit = str(series.get("unit") or "")
        industry = str(series.get("industry") or "매크로")
        frequency = str(series.get("frequency") or "FRED")
        source_url = f"https://fred.stlouisfed.org/series/{series_id}"

        history_key = f"fred-{series_id}"
        cached_last = cached_history_last_date(config, history_key)
        # 최초 1회만 전체 기간을 백필하고, 이후에는 개정치 반영을 위해 최근 450일만 다시 받습니다.
        observation_start = (
            (cached_last - timedelta(days=450)).isoformat() if cached_last else ""
        )

        try:
            points, source_label = fetch_fred_history(
                session=session,
                series_id=series_id,
                api_key=api_key,
                observation_start=observation_start,
            )
            if not points:
                metrics.append(
                    make_metric(
                        industry=industry,
                        name=name,
                        source="FRED",
                        source_url=source_url,
                        frequency=frequency,
                        automation="무료로 안정적으로 자동화 가능",
                        status="error",
                        note="관측값 없음",
                        depth=str(series.get("depth") or ""),
                    )
                )
                continue

            scale = to_float(series.get("scale")) or 1.0
            if scale != 1.0:
                points = [(point_date, value * scale) for point_date, value in points]
            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            yoy_value = find_yoy_value(points, latest_date)
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source=source_label,
                    source_url=source_url,
                    frequency=frequency,
                    automation="무료로 안정적으로 자동화 가능",
                    status="ok",
                    value=latest_value,
                    unit=unit,
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=points,
                    note=str(series.get("note") or ""),
                    group=str(series.get("group") or ""),
                    depth=str(series.get("depth") or ""),
                    meaning=str(series.get("meaning") or ""),
                    history_key=history_key,
                    section=str(series.get("section") or ""),
                    market_category=str(series.get("market_category") or ""),
                    also_market_category=series.get("also_market_category") or "",
                )
            )
        except Exception as exc:  # noqa: BLE001 - keep each card independent.
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="FRED",
                    source_url=source_url,
                    frequency=frequency,
                    automation="무료로 안정적으로 자동화 가능",
                    status="error",
                    note=str(exc),
                )
            )

    return metrics


def fetch_fred_history(
    session: requests.Session,
    series_id: str,
    api_key: str,
    observation_start: str = "",
) -> tuple[list[tuple[date, float]], str]:
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
        "sort_order": "asc",
    }
    if observation_start:
        params["observation_start"] = observation_start
    response = session.get(FRED_OBSERVATIONS_URL, params=params, timeout=(5, 30))
    response.raise_for_status()
    payload = response.json()
    points = []
    for item in payload.get("observations", []):
        value = to_float(item.get("value"))
        if value is None:
            continue
        points.append((date.fromisoformat(str(item["date"])), value))
    points.sort(key=lambda point: point[0])
    return points, "FRED API"


US_LIQUIDITY_GROUP = "미국 유동성"
US_LIQUIDITY_CATEGORY = "유동성"
US_NET_LIQUIDITY_KEY = "us-net-liquidity"
US_TGA_DAILY_KEY = "fiscaldata-tga"
US_RRP_KEY = "fred-RRPONTSYD"


def collect_us_liquidity_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
    existing_metrics: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    liquidity_config = config.get("us_liquidity", {}) or {}
    if liquidity_config.get("enabled", True) is False:
        return []

    api_key = os.getenv("FRED_API_KEY", "").strip()
    walcl_points = component_points_from_metric_or_store(config, existing_metrics, "fred-WALCL")
    walcl_metric = metric_by_history_key(existing_metrics, "fred-WALCL")
    if not walcl_points and api_key:
        walcl_points, _ = fetch_scaled_fred_component(
            config=config,
            session=session,
            series_id="WALCL",
            api_key=api_key,
            history_key="fred-WALCL",
            scale=0.001,
        )

    tga_metric, tga_points = collect_tga_component_metric(config, session, api_key, today)
    rrp_metric, rrp_points = collect_rrp_component_metric(config, session, api_key)
    net_metric = build_net_liquidity_metric(walcl_points, tga_points, rrp_points)
    walcl_market_metric = build_walcl_market_metric(walcl_metric, walcl_points)

    ordered = [net_metric, walcl_market_metric, tga_metric, rrp_metric]
    return [metric for metric in ordered if metric]


def metric_by_history_key(
    metrics: list[dict[str, Any]], history_key: str
) -> dict[str, Any] | None:
    for metric in metrics:
        if isinstance(metric, dict) and str(metric.get("history_key") or "") == history_key:
            return metric
    return None


def component_points_from_metric_or_store(
    config: dict[str, Any],
    metrics: list[dict[str, Any]],
    history_key: str,
) -> list[tuple[date, float]]:
    metric = metric_by_history_key(metrics, history_key)
    incoming = parse_stored_points(metric.get("history") if isinstance(metric, dict) else None)
    return merged_component_points(config, history_key, incoming)


def merged_component_points(
    config: dict[str, Any],
    history_key: str,
    incoming: list[tuple[date, float]],
) -> list[tuple[date, float]]:
    merged: dict[date, float] = {}
    store = attach_history_store(config)
    if store is not None:
        for point_date, value in store.series(history_key):
            merged[point_date] = value
    for point_date, value in incoming:
        merged[point_date] = value
    return sorted(merged.items(), key=lambda item: item[0])


def fetch_scaled_fred_component(
    *,
    config: dict[str, Any],
    session: requests.Session,
    series_id: str,
    api_key: str,
    history_key: str,
    scale: float = 1.0,
) -> tuple[list[tuple[date, float]], str]:
    cached_last = cached_history_last_date(config, history_key)
    observation_start = (cached_last - timedelta(days=450)).isoformat() if cached_last else ""
    incoming, source_label = fetch_fred_history(
        session=session,
        series_id=series_id,
        api_key=api_key,
        observation_start=observation_start,
    )
    if scale != 1.0:
        incoming = [(point_date, value * scale) for point_date, value in incoming]
    return merged_component_points(config, history_key, incoming), source_label


def collect_rrp_component_metric(
    config: dict[str, Any],
    session: requests.Session,
    api_key: str,
) -> tuple[dict[str, Any], list[tuple[date, float]]]:
    name = "미국 역레포"
    source_url = "https://fred.stlouisfed.org/series/RRPONTSYD"
    if not api_key:
        return (
            make_metric(
                industry="매크로",
                name=name,
                source="FRED API",
                source_url=source_url,
                frequency="일간",
                automation="무료로 안정적으로 자동화 가능",
                status="needs_key",
                note="GitHub Secrets에 FRED_API_KEY 등록 필요",
                group=US_LIQUIDITY_GROUP,
                meaning=US_RRP_MEANING,
                metric_id="us-rrp",
                section="market",
                market_category=US_LIQUIDITY_CATEGORY,
                history_key=US_RRP_KEY,
            ),
            [],
        )
    try:
        points, source_label = fetch_scaled_fred_component(
            config=config,
            session=session,
            series_id="RRPONTSYD",
            api_key=api_key,
            history_key=US_RRP_KEY,
            scale=0.001,
        )
        if not points:
            raise ValueError("관측값 없음")
        latest_date, latest_value = points[-1]
        previous_value = points[-2][1] if len(points) > 1 else None
        yoy_value = find_yoy_value(points, latest_date)
        return (
            make_metric(
                industry="매크로",
                name=name,
                source=source_label,
                source_url=source_url,
                frequency="일간",
                automation="무료로 안정적으로 자동화 가능",
                status="ok",
                value=latest_value,
                unit="$B",
                observed_at=latest_date.isoformat(),
                previous_value=previous_value,
                yoy_value=yoy_value,
                history=points,
                group=US_LIQUIDITY_GROUP,
                meaning=US_RRP_MEANING,
                metric_id="us-rrp",
                section="market",
                market_category=US_LIQUIDITY_CATEGORY,
                history_key=US_RRP_KEY,
            ),
            points,
        )
    except Exception as exc:  # noqa: BLE001
        return (
            make_metric(
                industry="매크로",
                name=name,
                source="FRED",
                source_url=source_url,
                frequency="일간",
                automation="무료로 안정적으로 자동화 가능",
                status="error",
                note=str(exc),
                group=US_LIQUIDITY_GROUP,
                meaning=US_RRP_MEANING,
                metric_id="us-rrp",
                section="market",
                market_category=US_LIQUIDITY_CATEGORY,
                history_key=US_RRP_KEY,
            ),
            [],
        )


US_NET_LIQUIDITY_MEANING = (
    "연준이 공급한 돈에서 재무부 금고(TGA)와 역레포에 잠긴 돈을 뺀, 실제로 금융시장에 돌고 있는 달러 "
    "유동성입니다. 증가하면 위험자산에 우호적, 감소하면 부담이 되는 흐름으로 해석합니다. 계산은 WALCL, "
    "TGA, 역레포의 관측일이 다를 때 각 날짜 이전의 가장 최근 값을 사용합니다."
)
US_TGA_MEANING = (
    "재무부가 연준에 맡겨둔 현금입니다. TGA가 늘면 시중 유동성이 흡수되고, 줄면 방출됩니다. "
    "부채한도 협상 국면에서 크게 출렁입니다."
)
US_RRP_MEANING = (
    "시중 자금이 연준에 하루짜리로 파킹된 규모입니다. 줄어들면 그만큼 시장에 유동성이 풀려나오는 효과가 있습니다."
)


def collect_tga_component_metric(
    config: dict[str, Any],
    session: requests.Session,
    api_key: str,
    today: date,
) -> tuple[dict[str, Any], list[tuple[date, float]]]:
    try:
        points = fetch_fiscaldata_tga_history(config, session)
        if not points:
            raise ValueError("FiscalData TGA 관측값 없음")
        return build_tga_metric(
            points,
            source="FiscalData DTS API",
            source_url=FISCALDATA_TGA_URL,
            frequency="일간",
            history_key=US_TGA_DAILY_KEY,
            note="",
        ), points
    except Exception as fiscal_exc:  # noqa: BLE001 - fallback to weekly FRED TGA.
        if not api_key:
            return (
                make_metric(
                    industry="매크로",
                    name="미국 TGA",
                    source="FiscalData DTS API",
                    source_url=FISCALDATA_TGA_URL,
                    frequency="일간",
                    automation="무료로 안정적으로 자동화 가능",
                    status="error",
                    note=f"FiscalData 실패: {fiscal_exc}",
                    group=US_LIQUIDITY_GROUP,
                    meaning=US_TGA_MEANING,
                    metric_id="us-tga",
                    section="market",
                    market_category=US_LIQUIDITY_CATEGORY,
                    history_key=US_TGA_DAILY_KEY,
                ),
                [],
            )
        try:
            points, source_label = fetch_scaled_fred_component(
                config=config,
                session=session,
                series_id="WTREGEN",
                api_key=api_key,
                history_key="fred-WTREGEN",
                scale=0.001,
            )
            if not points:
                raise ValueError("FRED WTREGEN 관측값 없음")
            return build_tga_metric(
                points,
                source=source_label,
                source_url="https://fred.stlouisfed.org/series/WTREGEN",
                frequency="주간",
                history_key="fred-WTREGEN",
                note=f"FiscalData 일간 TGA 실패로 FRED WTREGEN 주간값 사용: {fiscal_exc}",
            ), points
        except Exception as fred_exc:  # noqa: BLE001
            return (
                make_metric(
                    industry="매크로",
                    name="미국 TGA",
                    source="FiscalData/FRED",
                    source_url=FISCALDATA_TGA_URL,
                    frequency="일간/주간",
                    automation="무료로 안정적으로 자동화 가능",
                    status="error",
                    note=f"FiscalData 실패: {fiscal_exc}; FRED 실패: {fred_exc}",
                    group=US_LIQUIDITY_GROUP,
                    meaning=US_TGA_MEANING,
                    metric_id="us-tga",
                    section="market",
                    market_category=US_LIQUIDITY_CATEGORY,
                    history_key=US_TGA_DAILY_KEY,
                ),
                [],
            )


def fetch_fiscaldata_tga_history(
    config: dict[str, Any],
    session: requests.Session,
) -> list[tuple[date, float]]:
    cached_last = cached_history_last_date(config, US_TGA_DAILY_KEY)
    params: dict[str, str] = {
        "fields": "record_date,account_type,open_today_bal",
        "filter": "account_type:eq:Treasury General Account (TGA) Closing Balance",
        "sort": "record_date",
        "page[size]": "10000",
    }
    if cached_last:
        start = (cached_last - timedelta(days=45)).isoformat()
        params["filter"] = (
            "account_type:eq:Treasury General Account (TGA) Closing Balance,"
            f"record_date:gte:{start}"
        )
    response = session.get(FISCALDATA_TGA_URL, params=params, timeout=(10, 45))
    response.raise_for_status()
    payload = response.json()
    incoming: list[tuple[date, float]] = []
    for item in payload.get("data", []):
        if not isinstance(item, dict):
            continue
        point_date_text = str(item.get("record_date") or "")
        value = to_float(item.get("open_today_bal"))
        if value is None:
            continue
        try:
            point_date = date.fromisoformat(point_date_text)
        except ValueError:
            continue
        # FiscalData DTS balances are reported in millions of dollars.
        incoming.append((point_date, value / 1000.0))
    incoming.sort(key=lambda point: point[0])
    return merged_component_points(config, US_TGA_DAILY_KEY, incoming)


def build_tga_metric(
    points: list[tuple[date, float]],
    *,
    source: str,
    source_url: str,
    frequency: str,
    history_key: str,
    note: str,
) -> dict[str, Any]:
    latest_date, latest_value = points[-1]
    previous_value = points[-2][1] if len(points) > 1 else None
    yoy_value = find_yoy_value(points, latest_date)
    return make_metric(
        industry="매크로",
        name="미국 TGA",
        source=source,
        source_url=source_url,
        frequency=frequency,
        automation="무료로 안정적으로 자동화 가능",
        status="ok",
        value=latest_value,
        unit="$B",
        observed_at=latest_date.isoformat(),
        previous_value=previous_value,
        yoy_value=yoy_value,
        history=points,
        note=note,
        group=US_LIQUIDITY_GROUP,
        meaning=US_TGA_MEANING,
        metric_id="us-tga",
        section="market",
        market_category=US_LIQUIDITY_CATEGORY,
        history_key=history_key,
    )


def build_walcl_market_metric(
    walcl_metric: dict[str, Any] | None,
    walcl_points: list[tuple[date, float]],
) -> dict[str, Any] | None:
    if not walcl_points:
        return None
    latest_date, latest_value = walcl_points[-1]
    previous_value = walcl_points[-2][1] if len(walcl_points) > 1 else None
    yoy_value = find_yoy_value(walcl_points, latest_date)
    source = str((walcl_metric or {}).get("source") or "FRED API")
    meaning = str((walcl_metric or {}).get("meaning") or "")
    if not meaning:
        meaning = "연준 대차대조표 규모로 양적완화/긴축 방향을 보여줍니다. 글로벌 유동성의 큰 물줄기를 확인하는 지표입니다."
    return make_metric(
        industry="매크로",
        name="미국 연준 총자산",
        source=source,
        source_url="https://fred.stlouisfed.org/series/WALCL",
        frequency="주간",
        automation="무료로 안정적으로 자동화 가능",
        status="ok",
        value=latest_value,
        unit="$B",
        observed_at=latest_date.isoformat(),
        previous_value=previous_value,
        yoy_value=yoy_value,
        history=walcl_points,
        group=US_LIQUIDITY_GROUP,
        meaning=meaning,
        history_key="fred-WALCL",
        metric_id="us-liquidity-walcl",
        section="market",
        market_category=US_LIQUIDITY_CATEGORY,
    )


def build_net_liquidity_metric(
    walcl_points: list[tuple[date, float]],
    tga_points: list[tuple[date, float]],
    rrp_points: list[tuple[date, float]],
) -> dict[str, Any] | None:
    net_points = calculate_us_net_liquidity(walcl_points, tga_points, rrp_points)
    if not net_points:
        return make_metric(
            industry="매크로",
            name="미국 순유동성",
            source="FRED/FiscalData",
            source_url=FISCALDATA_TGA_URL,
            frequency="일간",
            automation="무료로 안정적으로 자동화 가능",
            status="error",
            note="WALCL, TGA, 역레포 중 계산에 필요한 시계열이 부족합니다.",
            group=US_LIQUIDITY_GROUP,
            meaning=US_NET_LIQUIDITY_MEANING,
            history_key=US_NET_LIQUIDITY_KEY,
            metric_id="us-net-liquidity",
            section="market",
            market_category=US_LIQUIDITY_CATEGORY,
        )
    latest_date, latest_value = net_points[-1]
    previous_value = net_points[-2][1] if len(net_points) > 1 else None
    yoy_value = find_yoy_value(net_points, latest_date)
    return make_metric(
        industry="매크로",
        name="미국 순유동성",
        source="FRED/FiscalData",
        source_url=FISCALDATA_TGA_URL,
        frequency="일간",
        automation="무료로 안정적으로 자동화 가능",
        status="ok",
        value=latest_value,
        unit="$B",
        observed_at=latest_date.isoformat(),
        previous_value=previous_value,
        yoy_value=yoy_value,
        history=net_points,
        group=US_LIQUIDITY_GROUP,
        meaning=US_NET_LIQUIDITY_MEANING,
        history_key=US_NET_LIQUIDITY_KEY,
        metric_id="us-net-liquidity",
        section="market",
        market_category=US_LIQUIDITY_CATEGORY,
    )


def calculate_us_net_liquidity(
    walcl_points: list[tuple[date, float]],
    tga_points: list[tuple[date, float]],
    rrp_points: list[tuple[date, float]],
) -> list[tuple[date, float]]:
    components = [walcl_points, tga_points, rrp_points]
    if any(not points for points in components):
        return []
    start = max(points[0][0] for points in components)
    end = max(points[-1][0] for points in components)
    if start > end:
        return []

    aligned: list[tuple[date, float]] = []
    indexes = [0, 0, 0]
    current = start
    while current <= end:
        values: list[float] = []
        for component_index, points in enumerate(components):
            while indexes[component_index] + 1 < len(points) and points[indexes[component_index] + 1][0] <= current:
                indexes[component_index] += 1
            if points[indexes[component_index]][0] > current:
                values = []
                break
            values.append(points[indexes[component_index]][1])
        if len(values) == 3:
            walcl, tga, rrp = values
            aligned.append((current, walcl - tga - rrp))
        current += timedelta(days=1)
    return aligned


BOJ_DATA_CODE_URL = "https://www.stat-search.boj.or.jp/api/v1/getDataCode"
ECB_DATA_API_BASE = "https://data-api.ecb.europa.eu/service/data"
GLOBAL_LIQUIDITY_SOURCE_NAME = "국제 유동성"

KOREA_LIQUIDITY_ITEMS: list[dict[str, Any]] = [
    {
        "name": "한국은행 총자산",
        "group": "한국 유동성",
        "stat_code": "103Y002",
        "item_code": "BCAA1",
        "history_key": "ecos-liquidity-103Y002-BCAA1",
        "metric_id": "korea-liquidity-bok-assets",
        "meaning": "한국은행 대차대조표의 자산 총액입니다. 중앙은행이 시장에 공급한 유동성의 큰 방향을 볼 때 기준으로 씁니다.",
    },
    {
        "name": "한국 본원통화",
        "group": "한국 유동성",
        "stat_code": "102Y004",
        "item_code": "ABA1",
        "history_key": "ecos-liquidity-102Y004-ABA1",
        "metric_id": "korea-liquidity-monetary-base",
        "meaning": "현금통화와 금융기관의 중앙은행 예치금을 합친 돈의 바탕입니다. 은행 시스템에 공급된 기본 유동성이 늘고 줄어드는지 확인합니다.",
    },
    {
        "name": "한국 M2",
        "group": "한국 유동성",
        "stat_code": "161Y005",
        "item_code": "BBHS00",
        "history_key": "ecos-liquidity-161Y005-BBHS00",
        "metric_id": "korea-liquidity-m2",
        "meaning": "현금, 요구불예금, 수시입출식 예금 등 비교적 바로 쓸 수 있는 돈을 넓게 묶은 통화량입니다. 가계와 기업의 자금 여유를 볼 때 핵심으로 봅니다.",
    },
    {
        "name": "한국 Lf",
        "group": "한국 유동성",
        "stat_code": "171Y003",
        "item_code": "LAS0000",
        "history_key": "ecos-liquidity-171Y003-LAS0000",
        "metric_id": "korea-liquidity-lf",
        "meaning": "M2보다 더 넓게 금융기관이 공급한 유동성을 보는 지표입니다. 은행권 밖까지 포함한 자금 여건의 폭을 확인합니다.",
    },
    {
        "name": "한국 L",
        "group": "한국 유동성",
        "stat_code": "172Y001",
        "item_code": "XS00000",
        "history_key": "ecos-liquidity-172Y001-XS00000",
        "metric_id": "korea-liquidity-l",
        "meaning": "국채, 회사채 같은 시장성 금융상품까지 포함한 가장 넓은 유동성 지표입니다. 한국 경제 전체의 돈의 양이 얼마나 넓게 풀려 있는지 볼 때 씁니다.",
    },
]

JAPAN_LIQUIDITY_ITEMS: list[dict[str, Any]] = [
    {
        "name": "일본 BOJ 총자산",
        "group": "일본 유동성",
        "db": "BS01",
        "code": "MABJMTA",
        "history_key": "boj-BS01-MABJMTA",
        "metric_id": "japan-liquidity-boj-assets",
        "meaning": "일본은행 대차대조표의 자산 총액입니다. 일본 중앙은행이 시장에 공급한 유동성의 큰 물줄기를 보여줍니다.",
    },
    {
        "name": "일본 본원통화",
        "group": "일본 유동성",
        "db": "MD01",
        "code": "MABS1AN11",
        "history_key": "boj-MD01-MABS1AN11",
        "metric_id": "japan-liquidity-monetary-base",
        "meaning": "일본의 현금통화와 일본은행 당좌예금 등을 합친 기본 통화량입니다. BOJ 정책이 실제 유동성으로 얼마나 남아 있는지 볼 때 봅니다.",
    },
    {
        "name": "일본 M2",
        "group": "일본 유동성",
        "db": "MD02",
        "code": "MAM1NAM2M2MO",
        "history_key": "boj-MD02-MAM1NAM2M2MO",
        "metric_id": "japan-liquidity-m2",
        "meaning": "일본 경제 안에서 가계와 기업이 비교적 쉽게 쓸 수 있는 돈의 규모입니다. 민간 유동성이 늘고 줄어드는지 확인합니다.",
    },
    {
        "name": "일본 BOJ 당좌예금",
        "group": "일본 유동성",
        "db": "MD08",
        "code": "MACAB2201",
        "history_key": "boj-MD08-MACAB2201",
        "metric_id": "japan-liquidity-current-account",
        "meaning": "금융기관이 일본은행에 맡겨둔 당좌예금 잔액입니다. 은행 시스템 안에 남아 있는 초과 유동성의 크기를 볼 때 참고합니다.",
    },
]

EUROPE_LIQUIDITY_ITEMS: list[dict[str, Any]] = [
    {
        "name": "유럽 초과유동성",
        "group": "유럽 유동성",
        "series_key": "ILM.D.U2.C.EXLIQ.U2.EUR",
        "history_key": "ecb-ILM-D-U2-C-EXLIQ-U2-EUR",
        "metric_id": "europe-liquidity-excess",
        "frequency": "일간",
        "meaning": "유로존 은행 시스템에 필요한 지급준비를 넘어서 남아 있는 유동성입니다. 숫자가 클수록 은행권에 여유자금이 많이 남아 있다는 뜻입니다.",
    },
    {
        "name": "유럽 유로시스템 총자산",
        "group": "유럽 유동성",
        "series_key": "BSI.M.U2.N.C.T00.A.1.Z5.0000.Z01.E",
        "history_key": "ecb-BSI-M-U2-N-C-T00-A-1-Z5-0000-Z01-E",
        "metric_id": "europe-liquidity-eurosystem-assets",
        "frequency": "월간",
        "meaning": "ECB와 유로존 중앙은행들의 총자산입니다. 양적완화와 긴축으로 중앙은행 유동성이 커지는지 줄어드는지 보여줍니다.",
    },
    {
        "name": "유럽 M3",
        "group": "유럽 유동성",
        "series_key": "BSI.M.U2.Y.V.M30.X.1.U2.2300.Z01.E",
        "history_key": "ecb-BSI-M-U2-Y-V-M30-X-1-U2-2300-Z01-E",
        "metric_id": "europe-liquidity-m3",
        "frequency": "월간",
        "meaning": "유로존의 넓은 통화량입니다. 가계와 기업, 금융기관에 풀린 돈의 규모와 민간 유동성 흐름을 볼 때 씁니다.",
    },
    {
        "name": "유럽 본원통화",
        "group": "유럽 유동성",
        "series_key": "ILM.M.U2.C.LT00001MP.Z5.EUR",
        "history_key": "ecb-ILM-M-U2-C-LT00001MP-Z5-EUR",
        "metric_id": "europe-liquidity-base-money",
        "frequency": "월간",
        "meaning": "유로존의 현금과 중앙은행 예치금을 합친 기본 통화량의 지급준비 유지기간 평균입니다. ECB 정책이 은행 시스템 안의 돈의 바탕을 얼마나 크게 만들고 있는지 보여줍니다.",
    },
]


def collect_global_liquidity_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
) -> list[dict[str, Any]]:
    liquidity_config = config.get("global_liquidity", {}) or {}
    if liquidity_config.get("enabled", True) is False:
        return []

    metrics: list[dict[str, Any]] = []
    metrics.extend(collect_korea_liquidity_metrics(config, session, today))
    metrics.extend(collect_japan_liquidity_metrics(config, session, today))
    metrics.extend(collect_europe_liquidity_metrics(config, session, today))
    return metrics


def collect_korea_liquidity_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
) -> list[dict[str, Any]]:
    ecos_config = config.get("ecos", {}) or {}
    api_key = os.getenv("ECOS_API_KEY", "").strip()
    source_url = str(ecos_config.get("source_url") or "https://ecos.bok.or.kr/api/")
    if not api_key:
        return [
            liquidity_placeholder_metric(
                item,
                source="한국은행 ECOS API",
                source_url=source_url,
                note="GitHub Secrets에 ECOS_API_KEY 등록 필요",
                status="needs_key",
                unit="조원",
            )
            for item in KOREA_LIQUIDITY_ITEMS
        ]

    metrics: list[dict[str, Any]] = []
    fetch_days = int(ecos_config.get("series_fetch_days", 1100))
    backfill_days = int(ecos_config.get("backfill_days", 9200))
    base_url = str(ecos_config.get("endpoint") or "https://ecos.bok.or.kr/api")
    for item in KOREA_LIQUIDITY_ITEMS:
        history_key = str(item["history_key"])
        start = today - timedelta(days=backfill_days if cached_history_last_date(config, history_key) is None else fetch_days)
        try:
            points = fetch_ecos_points(
                session=session,
                base_url=base_url,
                api_key=api_key,
                stat_code=str(item["stat_code"]),
                period="M",
                start=start,
                end=today,
                item_code=str(item["item_code"]),
                row_count=20000 if cached_history_last_date(config, history_key) is None else 2000,
            )
            # ECOS liquidity series used here are reported in billion KRW.
            points = [(point_date, value * 0.001) for point_date, value in points]
            points = merged_component_points(config, history_key, points)
            metrics.append(
                build_liquidity_metric(
                    item,
                    points,
                    source="한국은행 ECOS API",
                    source_url=source_url,
                    frequency="월간",
                    unit="조원",
                )
            )
        except Exception as exc:  # noqa: BLE001
            metrics.append(
                liquidity_placeholder_metric(
                    item,
                    source="한국은행 ECOS API",
                    source_url=source_url,
                    note=str(exc),
                    status="error",
                    unit="조원",
                )
            )
    return metrics


def collect_japan_liquidity_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
) -> list[dict[str, Any]]:
    metrics: list[dict[str, Any]] = []
    for item in JAPAN_LIQUIDITY_ITEMS:
        history_key = str(item["history_key"])
        try:
            points = fetch_boj_points(
                config=config,
                session=session,
                db=str(item["db"]),
                code=str(item["code"]),
                history_key=history_key,
            )
            # BOJ series used here are reported in 100 million yen; convert to trillion yen.
            points = [(point_date, value * 0.0001) for point_date, value in points]
            points = merged_component_points(config, history_key, points)
            metrics.append(
                build_liquidity_metric(
                    item,
                    points,
                    source="Bank of Japan API",
                    source_url=BOJ_DATA_CODE_URL,
                    frequency="월간",
                    unit="¥T",
                )
            )
        except Exception as exc:  # noqa: BLE001
            metrics.append(
                liquidity_placeholder_metric(
                    item,
                    source="Bank of Japan API",
                    source_url=BOJ_DATA_CODE_URL,
                    note=str(exc),
                    status="error",
                    unit="¥T",
                )
            )
    return metrics


def collect_europe_liquidity_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
) -> list[dict[str, Any]]:
    metrics: list[dict[str, Any]] = []
    for item in EUROPE_LIQUIDITY_ITEMS:
        history_key = str(item["history_key"])
        try:
            points = fetch_ecb_points(
                config=config,
                session=session,
                series_key=str(item["series_key"]),
                history_key=history_key,
            )
            # ECB BSI/ILM stock series used here are reported in million euro.
            points = [(point_date, value / 1000.0) for point_date, value in points]
            points = merged_component_points(config, history_key, points)
            metrics.append(
                build_liquidity_metric(
                    item,
                    points,
                    source="ECB Data Portal API",
                    source_url=ecb_series_url(str(item["series_key"])),
                    frequency=str(item.get("frequency") or "월간"),
                    unit="€B",
                )
            )
        except Exception as exc:  # noqa: BLE001
            metrics.append(
                liquidity_placeholder_metric(
                    item,
                    source="ECB Data Portal API",
                    source_url=ecb_series_url(str(item["series_key"])),
                    note=str(exc),
                    status="error",
                    unit="€B",
                )
            )
    return metrics


def liquidity_placeholder_metric(
    item: dict[str, Any],
    *,
    source: str,
    source_url: str,
    note: str,
    status: str,
    unit: str,
) -> dict[str, Any]:
    return make_metric(
        industry="매크로",
        name=str(item.get("name") or "유동성 지표"),
        source=source,
        source_url=source_url,
        frequency=str(item.get("frequency") or "월간"),
        automation="무료 공식 API 자동 수집",
        status=status,
        unit=unit,
        note=note,
        group=str(item.get("group") or GLOBAL_LIQUIDITY_SOURCE_NAME),
        meaning=str(item.get("meaning") or ""),
        history_key=str(item.get("history_key") or ""),
        metric_id=str(item.get("metric_id") or ""),
        section="market",
        market_category=US_LIQUIDITY_CATEGORY,
    )


def build_liquidity_metric(
    item: dict[str, Any],
    points: list[tuple[date, float]],
    *,
    source: str,
    source_url: str,
    frequency: str,
    unit: str,
) -> dict[str, Any]:
    if not points:
        return liquidity_placeholder_metric(
            item,
            source=source,
            source_url=source_url,
            note="관측값 없음",
            status="error",
            unit=unit,
        )
    latest_date, latest_value = points[-1]
    previous_value = points[-2][1] if len(points) > 1 else None
    yoy_value = find_yoy_value(points, latest_date)
    return make_metric(
        industry="매크로",
        name=str(item.get("name") or "유동성 지표"),
        source=source,
        source_url=source_url,
        frequency=frequency,
        automation="무료 공식 API 자동 수집",
        status="ok",
        value=latest_value,
        unit=unit,
        observed_at=latest_date.isoformat(),
        previous_value=previous_value,
        yoy_value=yoy_value,
        history=points,
        group=str(item.get("group") or GLOBAL_LIQUIDITY_SOURCE_NAME),
        meaning=str(item.get("meaning") or ""),
        history_key=str(item.get("history_key") or ""),
        metric_id=str(item.get("metric_id") or ""),
        section="market",
        market_category=US_LIQUIDITY_CATEGORY,
    )


def fetch_boj_points(
    *,
    config: dict[str, Any],
    session: requests.Session,
    db: str,
    code: str,
    history_key: str,
) -> list[tuple[date, float]]:
    params: dict[str, str] = {"format": "json", "lang": "en", "db": db, "code": code}
    cached_last = cached_history_last_date(config, history_key)
    if cached_last is not None:
        params["startDate"] = (cached_last - timedelta(days=450)).strftime("%Y%m")
    response = session.get(BOJ_DATA_CODE_URL, params=params, timeout=(10, 45))
    response.raise_for_status()
    return parse_boj_points(response.json())


def parse_boj_points(payload: dict[str, Any]) -> list[tuple[date, float]]:
    if int(payload.get("STATUS") or 0) != 200:
        raise ValueError(str(payload.get("MESSAGE") or payload.get("MESSAGEID") or "BOJ API error"))
    resultset = payload.get("RESULTSET") or []
    points: list[tuple[date, float]] = []
    for result in resultset:
        values = (result or {}).get("VALUES") or {}
        dates = values.get("SURVEY_DATES") or []
        observations = values.get("VALUES") or []
        for raw_date, raw_value in zip(dates, observations):
            point_date = parse_api_period(str(raw_date))
            value = to_float(raw_value)
            if point_date is not None and value is not None:
                points.append((point_date, value))
    points.sort(key=lambda point: point[0])
    return points


def fetch_ecb_points(
    *,
    config: dict[str, Any],
    session: requests.Session,
    series_key: str,
    history_key: str,
) -> list[tuple[date, float]]:
    params: dict[str, str] = {"detail": "dataonly"}
    cached_last = cached_history_last_date(config, history_key)
    frequency = ecb_frequency_from_key(series_key)
    if cached_last is not None:
        start = cached_last - timedelta(days=450)
        params["startPeriod"] = start.strftime("%Y-%m") if frequency == "M" else start.isoformat()
    response = session.get(
        ecb_series_url(series_key),
        params=params,
        headers={"Accept": "text/csv"},
        timeout=(10, 45),
    )
    response.raise_for_status()
    return parse_ecb_csv_points(response.text)


def ecb_series_url(series_key: str) -> str:
    parts = series_key.split(".", 1)
    if len(parts) != 2:
        raise ValueError(f"ECB series_key 형식 오류: {series_key}")
    flow, key = parts
    return f"{ECB_DATA_API_BASE}/{flow}/{key}"


def ecb_frequency_from_key(series_key: str) -> str:
    parts = series_key.split(".")
    return parts[1].upper() if len(parts) > 1 else ""


def parse_ecb_csv_points(text: str) -> list[tuple[date, float]]:
    rows = csv.DictReader(StringIO(text))
    points: list[tuple[date, float]] = []
    for row in rows:
        point_date = parse_api_period(str(row.get("TIME_PERIOD") or ""))
        value = to_float(row.get("OBS_VALUE"))
        if point_date is not None and value is not None:
            points.append((point_date, value))
    points.sort(key=lambda point: point[0])
    return points


def parse_api_period(value: str) -> date | None:
    text = value.strip()
    if not text:
        return None
    try:
        if "-W" in text:
            year_text, week_text = text.split("-W", 1)
            return date.fromisocalendar(int(year_text), int(week_text[:2]), 5)
        if len(text) >= 10 and text[4] == "-" and text[7] == "-":
            return date.fromisoformat(text[:10])
        if len(text) >= 7 and text[4] == "-":
            return date(int(text[:4]), int(text[5:7]), 1)
        if "Q" in text.upper():
            year_text, quarter_text = text.upper().split("Q", 1)
            return date(int(year_text), (int(quarter_text[:1]) - 1) * 3 + 1, 1)
        if len(text) >= 8 and text[:8].isdigit():
            return date(int(text[:4]), int(text[4:6]), int(text[6:8]))
        if len(text) >= 6 and text[:6].isdigit():
            return date(int(text[:4]), int(text[4:6]), 1)
        if len(text) >= 4 and text[:4].isdigit():
            return date(int(text[:4]), 1, 1)
    except ValueError:
        return None
    return None


def collect_ecos_credit_spread_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    ecos_config = config.get("ecos", {})
    if not ecos_config.get("enabled", True):
        return []

    items = ecos_config.get("credit_spreads", [])
    if not items:
        return []

    api_key = os.getenv("ECOS_API_KEY", "").strip()
    fetch_days = int(ecos_config.get("fetch_days", 730))
    backfill_days = int(ecos_config.get("backfill_days", 9200))
    row_count = int(ecos_config.get("row_count", 1000))
    if api_key == "sample":
        fetch_days = min(fetch_days, 14)
        backfill_days = min(backfill_days, 14)
        row_count = min(row_count, 10)
    source_url = str(ecos_config.get("source_url") or "https://ecos.bok.or.kr/api/")

    if not api_key:
        return [
            make_metric(
                industry=str(item.get("industry") or "은행/금융"),
                name=str(item.get("name") or "한국 신용 스프레드"),
                source="한국은행 ECOS API",
                source_url=source_url,
                frequency=str(item.get("frequency") or "일간"),
                automation="무료로 안정적으로 자동화 가능",
                status="needs_key",
                note="GitHub Secrets에 ECOS_API_KEY 등록 필요",
                group=str(item.get("group") or "신용 스프레드"),
                meaning=str(item.get("meaning") or ""),
            )
            for item in items
        ]

    metrics: list[dict[str, Any]] = []
    for item in items:
        name = str(item.get("name") or "한국 신용 스프레드")
        industry = str(item.get("industry") or "은행/금융")
        frequency = str(item.get("frequency") or "일간")
        group = str(item.get("group") or "신용 스프레드")
        meaning = str(item.get("meaning") or "")
        stat_code = str(item.get("stat_code") or "817Y002")
        corporate_code = str(item.get("corporate_item_code") or "")
        treasury_code = str(item.get("treasury_item_code") or "")
        history_key = f"ecos-spread-{stat_code}-{corporate_code}-{treasury_code}"
        # 최초 1회는 ECOS가 제공하는 과거 구간을 넓게 백필하고, 이후엔 최근 구간만 갱신합니다.
        if cached_history_last_date(config, history_key) is None:
            item_fetch_start = today - timedelta(days=backfill_days)
            item_row_count = max(row_count, 20000)
        else:
            item_fetch_start = today - timedelta(days=fetch_days)
            item_row_count = row_count
        try:
            corporate_points = fetch_ecos_points(
                session=session,
                base_url=str(ecos_config.get("endpoint") or "https://ecos.bok.or.kr/api"),
                api_key=api_key,
                stat_code=stat_code,
                period=str(item.get("period") or "D"),
                start=item_fetch_start,
                end=today,
                item_code=corporate_code,
                row_count=item_row_count,
            )
            treasury_points = fetch_ecos_points(
                session=session,
                base_url=str(ecos_config.get("endpoint") or "https://ecos.bok.or.kr/api"),
                api_key=api_key,
                stat_code=stat_code,
                period=str(item.get("period") or "D"),
                start=item_fetch_start,
                end=today,
                item_code=treasury_code,
                row_count=item_row_count,
            )
            points = compute_spread_points(corporate_points, treasury_points)
            if not points:
                metrics.append(
                    make_metric(
                        industry=industry,
                        name=name,
                        source="한국은행 ECOS API",
                        source_url=source_url,
                        frequency=frequency,
                        automation="무료로 안정적으로 자동화 가능",
                        status="error",
                        note="관측값 없음",
                        group=group,
                        meaning=meaning,
                    )
                )
                continue

            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            yoy_value = find_yoy_value(points, latest_date)
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="한국은행 ECOS API",
                    source_url=source_url,
                    frequency=frequency,
                    automation="무료로 안정적으로 자동화 가능",
                    status="ok",
                    value=latest_value,
                    unit=str(item.get("unit") or "%"),
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=points,
                    note=str(item.get("note") or ""),
                    group=group,
                    meaning=meaning,
                    history_key=history_key,
                )
            )
        except Exception as exc:  # noqa: BLE001 - one ECOS item should not break the dashboard.
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="한국은행 ECOS API",
                    source_url=source_url,
                    frequency=frequency,
                    automation="무료로 안정적으로 자동화 가능",
                    status="error",
                    note=str(exc),
                    group=group,
                    meaning=meaning,
                )
            )
    return metrics


def collect_ecos_series_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    """ECOS 단일 통계 시리즈(소비자심리지수, 선행지수 등) 수집기."""
    ecos_config = config.get("ecos", {})
    if not ecos_config.get("enabled", True):
        return []

    items = ecos_config.get("series", [])
    if not items:
        return []

    api_key = os.getenv("ECOS_API_KEY", "").strip()
    source_url = str(ecos_config.get("source_url") or "https://ecos.bok.or.kr/api/")
    fetch_days = int(ecos_config.get("series_fetch_days", 1100))
    backfill_days = int(ecos_config.get("backfill_days", 9200))

    if not api_key:
        return [
            make_metric(
                industry=str(item.get("industry") or "매크로"),
                name=str(item.get("name") or "ECOS 지표"),
                source="한국은행 ECOS API",
                source_url=source_url,
                frequency=str(item.get("frequency") or "월간"),
                automation="무료로 안정적으로 자동화 가능",
                status="needs_key",
                note="GitHub Secrets에 ECOS_API_KEY 등록 필요",
                group=str(item.get("group") or ""),
                meaning=str(item.get("meaning") or ""),
            )
            for item in items
        ]

    metrics: list[dict[str, Any]] = []
    for item in items:
        name = str(item.get("name") or "ECOS 지표")
        industry = str(item.get("industry") or "매크로")
        frequency = str(item.get("frequency") or "월간")
        group = str(item.get("group") or "")
        meaning = str(item.get("meaning") or "")
        stat_code = str(item.get("stat_code") or "")
        period = str(item.get("period") or "M")
        item_code = str(item.get("item_code") or "")
        item_code2 = str(item.get("item_code2") or "")
        if not stat_code or not item_code:
            continue

        history_key = f"ecos-{stat_code}-{item_code}{('-' + item_code2) if item_code2 else ''}"
        if cached_history_last_date(config, history_key) is None:
            start = today - timedelta(days=backfill_days)
            row_count = 20000
        else:
            start = today - timedelta(days=fetch_days)
            row_count = 2000

        try:
            points = fetch_ecos_points(
                session=session,
                base_url=str(ecos_config.get("endpoint") or "https://ecos.bok.or.kr/api"),
                api_key=api_key,
                stat_code=stat_code,
                period=period,
                start=start,
                end=today,
                item_code=item_code,
                item_code2=item_code2,
                row_count=row_count,
            )
            scale = to_float(item.get("scale")) or 1.0
            if scale != 1.0:
                points = [(point_date, value * scale) for point_date, value in points]
            if not points:
                metrics.append(
                    make_metric(
                        industry=industry,
                        name=name,
                        source="한국은행 ECOS API",
                        source_url=source_url,
                        frequency=frequency,
                        automation="무료로 안정적으로 자동화 가능",
                        status="error",
                        note="관측값 없음",
                        group=group,
                        meaning=meaning,
                    )
                )
                continue

            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            yoy_value = find_yoy_value(points, latest_date)
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="한국은행 ECOS API",
                    source_url=source_url,
                    frequency=frequency,
                    automation="무료로 안정적으로 자동화 가능",
                    status="ok",
                    value=latest_value,
                    unit=str(item.get("unit") or ""),
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=points,
                    note=str(item.get("note") or ""),
                    group=group,
                    meaning=meaning,
                    history_key=history_key,
                )
            )
        except Exception as exc:  # noqa: BLE001 - one ECOS item should not break the dashboard.
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="한국은행 ECOS API",
                    source_url=source_url,
                    frequency=frequency,
                    automation="무료로 안정적으로 자동화 가능",
                    status="error",
                    note=str(exc),
                    group=group,
                    meaning=meaning,
                )
            )
    return metrics


def fetch_ecos_points(
    *,
    session: requests.Session,
    base_url: str,
    api_key: str,
    stat_code: str,
    period: str,
    start: date,
    end: date,
    item_code: str,
    item_code2: str = "",
    row_count: int = 1000,
) -> list[tuple[date, float]]:
    if not item_code:
        return []
    period_upper = period.upper()
    if period_upper == "M":
        start_text, end_text = start.strftime("%Y%m"), end.strftime("%Y%m")
    elif period_upper == "Q":
        start_text = f"{start.year}Q{(start.month - 1) // 3 + 1}"
        end_text = f"{end.year}Q{(end.month - 1) // 3 + 1}"
    elif period_upper == "A":
        start_text, end_text = str(start.year), str(end.year)
    else:
        start_text, end_text = start.strftime("%Y%m%d"), end.strftime("%Y%m%d")
    url = (
        f"{base_url.rstrip('/')}/StatisticSearch/{api_key}/json/kr/1/{row_count}/"
        f"{stat_code}/{period}/{start_text}/{end_text}/{item_code}"
    )
    if item_code2:
        url = f"{url}/{item_code2}"
    response = session.get(url, timeout=(5, 20))
    response.raise_for_status()
    return parse_ecos_points(response.json(), period)


def parse_ecos_points(payload: dict[str, Any], period: str = "D") -> list[tuple[date, float]]:
    result = payload.get("RESULT") or {}
    code = str(result.get("CODE") or "")
    if code and code != "INFO-200":
        raise ValueError(str(result.get("MESSAGE") or code))

    rows = (payload.get("StatisticSearch") or {}).get("row") or []
    points: list[tuple[date, float]] = []
    for row in rows:
        observed_at = parse_ecos_period(str(row.get("TIME") or ""), period)
        value = to_float(row.get("DATA_VALUE"))
        if observed_at is None or value is None:
            continue
        points.append((observed_at, value))
    points.sort(key=lambda point: point[0])
    return points


def parse_ecos_period(value: str, period: str = "D") -> date | None:
    if not value:
        return None
    compact_period = period.upper()
    try:
        if compact_period == "D" and len(value) >= 8:
            return date(int(value[:4]), int(value[4:6]), int(value[6:8]))
        if compact_period == "M" and len(value) >= 6:
            return date(int(value[:4]), int(value[4:6]), 1)
        if compact_period == "Q" and "Q" in value.upper():
            year_text, quarter_text = value.upper().split("Q", 1)
            return date(int(year_text), (int(quarter_text) - 1) * 3 + 1, 1)
        if len(value) >= 4:
            return date(int(value[:4]), 1, 1)
    except ValueError:
        return None
    return None


KOFIA_STATISTICS_ENDPOINT = "https://apis.data.go.kr/1160100/service/GetKofiaStatisticsInfoService"
KOFIA_SOURCE_URL = "https://www.data.go.kr/data/15094809/openapi.do"

KOFIA_DAILY_ITEMS = [
    {
        "name": "한국 투자자예탁금",
        "operation": "getSecuritiesMarketTotalCapitalInfo",
        "value_field": "invrDpsgAmt",
        "history_key": "kofia-securities-market-investor-deposits",
        "unit": "조원",
        "scale": 0.000000000001,
        "meaning": "증권계좌에 대기 중인 현금입니다. 늘어나면 주식시장으로 들어올 수 있는 대기 자금이 많아졌다는 뜻으로 봅니다.",
    },
    {
        "name": "한국 신용융자 잔고",
        "operation": "getGrantingOfCreditBalanceInfo",
        "value_field": "crdTrFingWhl",
        "history_key": "kofia-credit-financing-balance",
        "unit": "조원",
        "scale": 0.000000000001,
        "meaning": "투자자가 빚을 내서 주식을 산 잔고입니다. 빠르게 늘면 과열 신호가 될 수 있고, 급감하면 반대매매 압력을 의심할 수 있습니다.",
    },
    {
        "name": "한국 CMA 잔고",
        "operation": "getCMAStatus",
        "value_field": "actBal",
        "history_key": "kofia-cma-balance",
        "unit": "조원",
        "scale": 0.000000000001,
        "row_filter": {"mngInvTgt": "합계", "invrCtg": "합계"},
        "meaning": "CMA 계좌에 머무는 단기 대기자금입니다. 증시 주변의 현금 여력과 단기자금 선호가 커지는지 확인할 때 봅니다.",
    },
]


def collect_kofia_capital_market_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    del today
    kofia_config = config.get("kofia", {}) or {}
    if not kofia_config.get("enabled", True):
        return []

    service_key = os.getenv("DATA_GO_KR_SERVICE_KEY", "").strip()
    endpoint = str(kofia_config.get("endpoint") or KOFIA_STATISTICS_ENDPOINT).rstrip("/")
    source_url = str(kofia_config.get("source_url") or KOFIA_SOURCE_URL)
    items = kofia_config.get("daily_items") or KOFIA_DAILY_ITEMS
    num_rows = int(kofia_config.get("num_rows") or 10000)

    if not service_key:
        metrics: list[dict[str, Any]] = []
        for item in items:
            operation = str(item.get("operation") or "")
            value_field = str(item.get("value_field") or "")
            history_key = str(item.get("history_key") or f"kofia-{operation}-{value_field}")
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=str(item.get("name") or "금투협회 증시자금"),
                    source="금융투자협회 종합통계",
                    source_url=source_url,
                    frequency="일간",
                    automation="공공데이터 API 자동 수집",
                    status="needs_key",
                    note="GitHub Secrets에 DATA_GO_KR_SERVICE_KEY 등록 필요",
                    group="수급 과열",
                    section="market",
                    market_category="신용·예탁금",
                    meaning=str(item.get("meaning") or ""),
                    history_key=history_key,
                    metric_id=history_key,
                )
            )
        return metrics

    metrics: list[dict[str, Any]] = []
    for item in items:
        name = str(item.get("name") or "금투협회 증시자금")
        operation = str(item.get("operation") or "")
        value_field = str(item.get("value_field") or "")
        history_key = str(item.get("history_key") or f"kofia-{operation}-{value_field}")
        if not operation or not value_field:
            continue
        try:
            rows = fetch_kofia_rows(
                session=session,
                endpoint=endpoint,
                operation=operation,
                service_key=service_key,
                num_rows=num_rows,
            )
            points = kofia_points_from_rows(
                rows,
                value_field=value_field,
                row_filter=item.get("row_filter") if isinstance(item.get("row_filter"), dict) else None,
                scale=to_float(item.get("scale")) or 1.0,
            )
            if not points:
                metrics.append(
                    make_metric(
                        industry="매크로",
                        name=name,
                        source="금융투자협회 종합통계",
                        source_url=source_url,
                        frequency="일간",
                        automation="공공데이터 API 자동 수집",
                        status="error",
                        note="관측값 없음",
                        group="수급 과열",
                        section="market",
                        market_category="신용·예탁금",
                        meaning=str(item.get("meaning") or ""),
                    )
                )
                continue

            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=name,
                    source="금융투자협회 종합통계",
                    source_url=source_url,
                    frequency="일간",
                    automation="공공데이터 API 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit=str(item.get("unit") or "조원"),
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=find_yoy_value(points, latest_date),
                    history=points,
                    group="수급 과열",
                    section="market",
                    market_category="신용·예탁금",
                    meaning=str(item.get("meaning") or ""),
                    history_key=history_key,
                    metric_id=history_key,
                )
            )
        except Exception as exc:  # noqa: BLE001 - 무료 공공 API 장애는 소스 단위 soft-fail.
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=name,
                    source="금융투자협회 종합통계",
                    source_url=source_url,
                    frequency="일간",
                    automation="공공데이터 API 자동 수집",
                    status="error",
                    note=f"금투협회 API 응답 실패: {exc}",
                    group="수급 과열",
                    section="market",
                    market_category="신용·예탁금",
                    meaning=str(item.get("meaning") or ""),
                    history_key=history_key,
                    metric_id=history_key,
                )
            )
    return metrics


def fetch_kofia_rows(
    session: requests.Session,
    endpoint: str,
    operation: str,
    service_key: str,
    num_rows: int = 10000,
) -> list[dict[str, Any]]:
    url = build_data_go_kr_url(
        f"{endpoint.rstrip('/')}/{operation}",
        service_key,
        {
            "pageNo": "1",
            "numOfRows": str(num_rows),
            "_type": "json",
            "resultType": "json",
        },
    )
    response = session.get(url, timeout=30)
    response.raise_for_status()
    text = response.text.strip()
    if text.startswith("{") or text.startswith("["):
        return parse_kofia_json_rows(response.json())
    return parse_kofia_xml_rows(text)


def parse_kofia_json_rows(payload: object) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    response = payload.get("response", payload)
    if not isinstance(response, dict):
        return []
    header = response.get("header")
    if isinstance(header, dict):
        result_code = str(header.get("resultCode") or header.get("returnReasonCode") or "")
        result_msg = str(header.get("resultMsg") or header.get("returnAuthMsg") or "")
        if result_code and result_code not in {"00", "0", "NORMAL_CODE"}:
            raise ValueError(f"공공데이터 API 오류 {result_code}: {result_msg or 'unknown'}")
    body = response.get("body", response)
    if not isinstance(body, dict):
        return []
    items = body.get("items", [])
    if isinstance(items, dict):
        items = items.get("item", [])
    if isinstance(items, dict):
        items = [items]
    if not isinstance(items, list):
        return []
    return [item for item in items if isinstance(item, dict)]


def parse_kofia_xml_rows(xml_text: str) -> list[dict[str, Any]]:
    root = ElementTree.fromstring(xml_text)
    result_code = first_text(root, "resultCode") or first_text(root, "returnReasonCode")
    result_message = (
        first_text(root, "resultMsg")
        or first_text(root, "returnAuthMsg")
        or first_text(root, "errMsg")
    )
    if result_code and result_code not in {"00", "0", "NORMAL_CODE"}:
        raise ValueError(f"공공데이터 API 오류 {result_code}: {result_message or 'unknown'}")
    rows: list[dict[str, Any]] = []
    for item in root.iter():
        if local_name(item.tag) != "item":
            continue
        row = {
            local_name(child.tag): (child.text or "").strip()
            for child in list(item)
            if child.text is not None
        }
        if row:
            rows.append(row)
    return rows


def kofia_points_from_rows(
    rows: list[dict[str, Any]],
    value_field: str,
    row_filter: dict[str, Any] | None = None,
    scale: float = 1.0,
) -> list[tuple[date, float]]:
    by_date: dict[date, float] = {}
    for row in rows:
        if row_filter and any(str(row.get(key) or "").strip() != str(value).strip() for key, value in row_filter.items()):
            continue
        observed_at = parse_kofia_date(row.get("basDt"))
        value = to_float(row.get(value_field))
        if observed_at is None or value is None:
            continue
        by_date[observed_at] = value * scale
    return sorted(by_date.items())


def parse_kofia_date(value: object) -> date | None:
    text = str(value or "").strip().replace("-", "")
    if not re.fullmatch(r"\d{8}", text):
        return parse_iso_date(value)
    try:
        return date(int(text[:4]), int(text[4:6]), int(text[6:8]))
    except ValueError:
        return None


def compute_spread_points(
    corporate_points: list[tuple[date, float]],
    treasury_points: list[tuple[date, float]],
) -> list[tuple[date, float]]:
    treasury_by_date = dict(treasury_points)
    spreads = [
        (observed_at, round(corporate_value - treasury_by_date[observed_at], 4))
        for observed_at, corporate_value in corporate_points
        if observed_at in treasury_by_date
    ]
    spreads.sort(key=lambda point: point[0])
    return spreads


def price_item_market(item: dict[str, Any]) -> str:
    symbol = str(item.get("symbol") or "").upper()
    if symbol.endswith((".KS", ".KQ")) or symbol in {"^KS11", "^KQ11", "^KS200"}:
        return "korea"
    if str(item.get("refresh_scope") or "") == "intraday":
        return "continuous"
    return "us"


def intraday_price_config(config: dict[str, Any], now: datetime) -> tuple[dict[str, Any], str]:
    equities = dict(config.get("equities", {}) or {})
    items = [item for item in equities.get("items", []) if isinstance(item, dict)]
    kst = now.astimezone(ZoneInfo("Asia/Seoul")) if now.tzinfo else now.replace(tzinfo=ZoneInfo("Asia/Seoul"))
    minute = kst.hour * 60 + kst.minute
    if kst.weekday() >= 5:
        markets = {"continuous"}
        label = "주말 연속시장"
    elif 8 * 60 + 30 <= minute < 16 * 60:
        markets = {"korea", "continuous"}
        label = "한국장"
    elif minute >= 16 * 60 or minute < 7 * 60:
        markets = {"us", "continuous"}
        label = "미국 프리·정규·애프터장"
    else:
        markets = {"continuous"}
        label = "세션 외 연속시장"
    equities["items"] = [item for item in items if price_item_market(item) in markets]
    scoped = dict(config)
    scoped["equities"] = equities
    return scoped, label


def is_us_equity_symbol(symbol: str) -> bool:
    upper = symbol.upper()
    if upper.startswith("^") or "=" in upper or upper.endswith((".KS", ".KQ", "-USD")):
        return False
    return bool(re.fullmatch(r"[A-Z][A-Z0-9.-]{0,14}", upper))


def classify_yahoo_trading_period(meta: dict[str, Any], timestamp: int) -> str:
    periods = meta.get("currentTradingPeriod")
    if not isinstance(periods, dict):
        return ""
    for key, label in (("pre", "premarket"), ("post", "afterhours"), ("regular", "regular")):
        period = periods.get(key)
        if not isinstance(period, dict):
            continue
        start = int(period.get("start") or 0)
        end = int(period.get("end") or 0)
        if start <= timestamp <= end:
            return label
    return ""


def fetch_yahoo_extended_quote(
    session: requests.Session, url: str
) -> dict[str, Any] | None:
    response = session.get(
        url,
        params={"range": "1d", "interval": "5m", "includePrePost": "true"},
        headers={"User-Agent": "Mozilla/5.0 stock-industry-dashboard/1.0"},
        timeout=(5, 30),
    )
    response.raise_for_status()
    result = ((response.json().get("chart") or {}).get("result") or [None])[0]
    if not isinstance(result, dict):
        return None
    timestamps = result.get("timestamp") or []
    quote = ((result.get("indicators") or {}).get("quote") or [{}])[0]
    closes = quote.get("close") or []
    values = [
        (int(timestamp), value)
        for timestamp, raw in zip(timestamps, closes)
        if (value := to_float(raw)) is not None
    ]
    if not values:
        return None
    timestamp, value = values[-1]
    meta = result.get("meta") if isinstance(result.get("meta"), dict) else {}
    market_session = classify_yahoo_trading_period(meta, timestamp)
    if market_session not in {"premarket", "afterhours"}:
        return None
    regular_market_time = int(meta.get("regularMarketTime") or 0)
    if regular_market_time and timestamp <= regular_market_time:
        return None
    observed = datetime.fromtimestamp(timestamp, tz=yahoo_exchange_timezone(meta))
    return {
        "value": value,
        "observed": observed,
        "market_session": market_session,
    }


def collect_equity_price_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
    *,
    intraday_now: datetime | None = None,
) -> list[dict[str, Any]]:
    del today
    equities_config = config.get("equities", {})
    if not equities_config.get("enabled", True):
        return []

    items = equities_config.get("items", [])
    if not items:
        return []

    endpoint_template = str(
        equities_config.get("endpoint")
        or "https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
    )
    source_url = str(equities_config.get("source_url") or "https://finance.yahoo.com/")
    metrics: list[dict[str, Any]] = []

    for item in items:
        symbol = str(item.get("symbol") or "").strip()
        if not symbol:
            continue
        name = str(item.get("name") or symbol)
        industry = str(item.get("industry") or "매크로")
        url = endpoint_template.format(symbol=symbol)
        quote_url = f"{source_url.rstrip('/')}/quote/{symbol}"
        history_key = f"equity-{symbol}"
        # 캐시가 없으면 상장 이후 전체(max)를 1회 백필하고, 이후에는 최근 3개월만 갱신합니다.
        fetch_range = "max" if cached_history_last_date(config, history_key) is None else str(
            item.get("range") or "3mo"
        )

        try:
            points, currency = fetch_equity_history_with_fallback(
                session, url, fetch_range, symbol
            )
            if not points:
                metrics.append(
                    make_metric(
                        industry=industry,
                        name=name,
                        source="Yahoo Finance chart API",
                        source_url=quote_url,
                        frequency="일간",
                        automation="무료 공개 JSON 자동 수집",
                        status="error",
                        note="관측값 없음",
                        group=str(item.get("group") or "대표주가"),
                        depth=str(item.get("depth") or ""),
                        meaning=str(item.get("meaning") or equity_price_meaning(name)),
                        section=str(item.get("section") or ""),
                        market_category=str(item.get("market_category") or ""),
                        also_market_category=item.get("also_market_category") or "",
                        refresh_scope=str(item.get("refresh_scope") or ""),
                        chart_style=str(item.get("chart_style") or ""),
                    )
                )
                continue

            extended_quote = None
            if intraday_now is not None and is_us_equity_symbol(symbol):
                try:
                    extended_quote = fetch_yahoo_extended_quote(session, url)
                except Exception:  # noqa: BLE001 - daily close remains a valid fallback.
                    extended_quote = None

            unit = str(
                item.get("unit")
                or ("원" if currency == "KRW" else "$" if currency == "USD" else currency)
            )
            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            price_session = ""
            price_observed_at = ""
            if extended_quote:
                quote_time = extended_quote["observed"]
                quote_date = quote_time.date()
                if quote_date >= latest_date:
                    close_baseline = latest_value
                    points_by_date = dict(points)
                    points_by_date[quote_date] = float(extended_quote["value"])
                    points = sorted(points_by_date.items())
                    latest_date, latest_value = points[-1]
                    previous_value = close_baseline
                    price_session = str(extended_quote["market_session"])
                    price_observed_at = quote_time.isoformat(timespec="seconds")
            yoy_value = find_yoy_value(points, latest_date)
            metric = make_metric(
                    industry=industry,
                    name=name,
                    source="Yahoo Finance chart API",
                    source_url=quote_url,
                    frequency="일간",
                    automation="무료 공개 JSON 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit=unit,
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=points,
                    note=str(item.get("note") or ""),
                    group=str(item.get("group") or "대표주가"),
                    depth=str(item.get("depth") or ""),
                    meaning=str(item.get("meaning") or equity_price_meaning(name)),
                    history_key=history_key,
                    section=str(item.get("section") or ""),
                    market_category=str(item.get("market_category") or ""),
                    also_market_category=item.get("also_market_category") or "",
                    refresh_scope=str(item.get("refresh_scope") or ""),
                    chart_style=str(item.get("chart_style") or ""),
                )
            if price_session:
                metric["price_session"] = price_session
                metric["price_observed_at"] = price_observed_at
            metrics.append(metric)
        except Exception as exc:  # noqa: BLE001 - one ticker should not break the dashboard.
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="Yahoo Finance chart API",
                    source_url=quote_url,
                    frequency="일간",
                    automation="무료 공개 JSON 자동 수집",
                    status="error",
                    note=str(exc),
                    group=str(item.get("group") or "대표주가"),
                    depth=str(item.get("depth") or ""),
                    meaning=str(item.get("meaning") or equity_price_meaning(name)),
                    section=str(item.get("section") or ""),
                    market_category=str(item.get("market_category") or ""),
                    also_market_category=item.get("also_market_category") or "",
                    refresh_scope=str(item.get("refresh_scope") or ""),
                    chart_style=str(item.get("chart_style") or ""),
                )
            )
    return metrics


def parse_yahoo_chart_points(payload: dict[str, Any]) -> tuple[list[tuple[date, float]], str]:
    chart = payload.get("chart", {})
    if chart.get("error"):
        raise ValueError(str(chart.get("error")))
    result = (chart.get("result") or [None])[0]
    if not isinstance(result, dict):
        return [], ""

    timestamps = result.get("timestamp") or []
    quote = ((result.get("indicators") or {}).get("quote") or [{}])[0]
    closes = quote.get("close") or []
    adjclose_item = ((result.get("indicators") or {}).get("adjclose") or [{}])[0]
    adjclose = adjclose_item.get("adjclose") or []
    meta = result.get("meta") or {}
    currency = str(meta.get("currency") or "")
    exchange_timezone = yahoo_exchange_timezone(meta)
    by_date: dict[date, float] = {}
    for index, timestamp in enumerate(timestamps):
        value = None
        if index < len(closes):
            value = to_float(closes[index])
        if value is None and index < len(adjclose):
            value = to_float(adjclose[index])
        if value is None:
            continue
        observed_at = datetime.fromtimestamp(int(timestamp), tz=exchange_timezone).date()
        by_date[observed_at] = value
    return sorted(by_date.items()), currency


def yahoo_exchange_timezone(meta: dict[str, Any]) -> timezone | ZoneInfo:
    timezone_name = str(
        meta.get("exchangeTimezoneName") or meta.get("timezone") or ""
    ).strip()
    if timezone_name:
        try:
            return ZoneInfo(timezone_name)
        except Exception:  # noqa: BLE001 - fall back to UTC when Yahoo returns an alias.
            pass
    return timezone.utc


def fetch_equity_history_with_fallback(
    session: requests.Session, url: str, fetch_range: str, symbol: str
) -> tuple[list[tuple[date, float]], str]:
    """Yahoo 차트 API를 우선 사용하고, 한국 종목/지수는 실패 시 네이버로 폴백합니다."""
    try:
        response = session.get(
            url,
            params={"range": fetch_range, "interval": "1d"},
            headers={"User-Agent": "Mozilla/5.0 stock-industry-dashboard/1.0"},
            timeout=(5, 30),
        )
        response.raise_for_status()
        points, currency = parse_yahoo_chart_points(response.json())
        if points:
            return points, currency
        raise ValueError("Yahoo 관측값 없음")
    except Exception:
        naver_symbol = naver_fallback_symbol(symbol)
        if not naver_symbol:
            raise
        count = 8000 if fetch_range == "max" else 90
        return fetch_naver_chart_points(session, naver_symbol, count), "KRW"


def naver_fallback_symbol(symbol: str) -> str:
    """네이버 fchart에서 쓸 심볼. 한국 종목/지수만 지원합니다."""
    if symbol == "^KS11":
        return "KOSPI"
    if symbol == "^KQ11":
        return "KOSDAQ"
    match = re.match(r"^(\d{6})\.(KS|KQ)$", symbol)
    if match:
        return match.group(1)
    return ""


def fetch_naver_chart_points(
    session: requests.Session, symbol: str, count: int
) -> list[tuple[date, float]]:
    response = session.get(
        "https://fchart.stock.naver.com/sise.nhn",
        params={
            "symbol": symbol,
            "timeframe": "day",
            "count": count,
            "requestType": "0",
        },
        headers={"User-Agent": "Mozilla/5.0 stock-industry-dashboard/1.0"},
        timeout=(5, 30),
    )
    response.raise_for_status()
    points: list[tuple[date, float]] = []
    for match in re.finditer(r'data="(\d{8})\|[^|]*\|[^|]*\|[^|]*\|([0-9.]+)\|', response.text):
        date_text, close_text = match.group(1), match.group(2)
        try:
            observed_at = date(int(date_text[:4]), int(date_text[4:6]), int(date_text[6:8]))
            points.append((observed_at, float(close_text)))
        except ValueError:
            continue
    points.sort(key=lambda point: point[0])
    return points


def equity_price_meaning(name: str) -> str:
    return f"{name} 주가는 시장이 해당 기업의 성장성과 위험을 어떻게 평가하는지 보여줍니다."


def collect_stablecoin_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    stablecoin_config = config.get("stablecoins", {})
    if not stablecoin_config.get("enabled", True):
        return []

    endpoint = str(
        stablecoin_config.get("endpoint")
        or "https://stablecoins.llama.fi/stablecoins?includePrices=true"
    )
    source_url = str(stablecoin_config.get("source_url") or "https://defillama.com/stablecoins")
    response = session.get(endpoint, timeout=(5, 20))
    response.raise_for_status()
    payload = response.json()
    assets = [
        asset
        for asset in payload.get("peggedAssets", [])
        if isinstance(asset, dict) and not asset.get("delisted")
    ]

    configured_assets = stablecoin_config.get("assets") or [
        {"symbol": "TOTAL", "name": "전체 스테이블코인 유통량"},
        {"symbol": "USDT", "name": "USDT 유통량"},
        {"symbol": "USDC", "name": "USDC 유통량"},
    ]

    metrics: list[dict[str, Any]] = []
    for item in configured_assets:
        symbol = str(item.get("symbol") or "").upper()
        asset_id = str(item.get("id") or "")
        name = str(item.get("name") or symbol or asset_id or "스테이블코인 유통량")

        if symbol == "TOTAL":
            values = stablecoin_total_values(assets)
        else:
            asset = find_stablecoin_asset(assets, symbol=symbol, asset_id=asset_id)
            if asset is None:
                metrics.append(
                    make_metric(
                        industry="스테이블코인",
                        name=name,
                        source="DefiLlama Stablecoins API",
                        source_url=source_url,
                        frequency="일간",
                        automation="무료로 안정적으로 자동화 가능",
                        status="error",
                        note=f"{symbol or asset_id} 자산을 찾을 수 없음",
                        group=str(item.get("group") or "유통량"),
                        meaning=str(item.get("meaning") or stablecoin_meaning()),
                    )
                )
                continue
            values = stablecoin_asset_values(asset)
            if not item.get("name"):
                name = f"{asset.get('name', symbol)}({asset.get('symbol', symbol)}) 유통량"

        current = values.get("current")
        if current is None:
            metrics.append(
                make_metric(
                    industry="스테이블코인",
                    name=name,
                    source="DefiLlama Stablecoins API",
                    source_url=source_url,
                    frequency="일간",
                    automation="무료로 안정적으로 자동화 가능",
                    status="error",
                    note="유통량 데이터 없음",
                    group=str(item.get("group") or "유통량"),
                    meaning=str(item.get("meaning") or stablecoin_meaning()),
                )
            )
            continue

        history_key = f"stablecoin-{symbol or asset_id or name}"
        history: list[tuple[date, float]] = []
        history_merge = "latest"
        # 최초 1회는 DefiLlama 히스토리 엔드포인트로 전체 기간(2017~)을 백필합니다.
        if cached_history_last_date(config, history_key) is None:
            try:
                resolved_id = asset_id
                if not resolved_id and symbol != "TOTAL":
                    matched = find_stablecoin_asset(assets, symbol=symbol, asset_id="")
                    resolved_id = str((matched or {}).get("id") or "")
                history = fetch_stablecoin_chart_history(
                    session,
                    str(stablecoin_config.get("charts_endpoint") or "https://stablecoins.llama.fi/stablecoincharts/all"),
                    resolved_id if symbol != "TOTAL" else "",
                )
                if history:
                    history_merge = "full"
            except Exception:  # noqa: BLE001 - 백필 실패는 스냅샷 축적으로 대체합니다.
                history = []
        if not history:
            history = stablecoin_history_points(values, today)
        metrics.append(
            make_metric(
                industry="스테이블코인",
                name=name,
                source="DefiLlama Stablecoins API",
                source_url=source_url,
                frequency="일간",
                automation="무료 공개 API 자동 수집",
                status="ok",
                value=current / 1_000_000_000,
                unit="$B",
                observed_at=today.isoformat(),
                previous_value=stablecoin_billions(values.get("prev_day")),
                yoy_value=None,
                history=history,
                note=str(item.get("note") or ""),
                group=str(item.get("group") or "유통량"),
                meaning=str(item.get("meaning") or stablecoin_meaning()),
                history_key=history_key,
                history_merge=history_merge,
            )
        )
    return metrics


def fetch_stablecoin_chart_history(
    session: requests.Session, charts_endpoint: str, asset_id: str
) -> list[tuple[date, float]]:
    """DefiLlama 차트 엔드포인트에서 유통량 전체 히스토리를 $B 단위로 가져옵니다."""
    params = {"stablecoin": asset_id} if asset_id else None
    response = session.get(charts_endpoint, params=params, timeout=(5, 30))
    response.raise_for_status()
    rows = response.json()
    points: list[tuple[date, float]] = []
    if not isinstance(rows, list):
        return points
    for row in rows:
        if not isinstance(row, dict):
            continue
        timestamp = to_float(row.get("date"))
        circulating = row.get("totalCirculatingUSD") or row.get("totalCirculating") or {}
        value = None
        if isinstance(circulating, dict):
            value = to_float(circulating.get("peggedUSD"))
        if timestamp is None or value is None:
            continue
        observed_at = datetime.fromtimestamp(int(timestamp), tz=timezone.utc).date()
        points.append((observed_at, value / 1_000_000_000))
    points.sort(key=lambda point: point[0])
    return points


def find_stablecoin_asset(
    assets: list[dict[str, Any]], *, symbol: str, asset_id: str
) -> dict[str, Any] | None:
    for asset in assets:
        if asset_id and str(asset.get("id") or "") == asset_id:
            return asset
        if symbol and str(asset.get("symbol") or "").upper() == symbol:
            return asset
    return None


def stablecoin_asset_values(asset: dict[str, Any]) -> dict[str, float | None]:
    return {
        "current": stablecoin_supply_value(asset, "circulating"),
        "prev_day": stablecoin_supply_value(asset, "circulatingPrevDay"),
        "prev_week": stablecoin_supply_value(asset, "circulatingPrevWeek"),
        "prev_month": stablecoin_supply_value(asset, "circulatingPrevMonth"),
    }


def stablecoin_total_values(assets: list[dict[str, Any]]) -> dict[str, float | None]:
    values: dict[str, float | None] = {}
    for key, source_key in [
        ("current", "circulating"),
        ("prev_day", "circulatingPrevDay"),
        ("prev_week", "circulatingPrevWeek"),
        ("prev_month", "circulatingPrevMonth"),
    ]:
        total = 0.0
        found = False
        for asset in assets:
            if str(asset.get("pegType") or "") != "peggedUSD":
                continue
            value = stablecoin_supply_value(asset, source_key)
            if value is None:
                continue
            total += value
            found = True
        values[key] = total if found else None
    return values


def stablecoin_supply_value(asset: dict[str, Any], key: str) -> float | None:
    value = asset.get(key)
    if isinstance(value, dict):
        return to_float(value.get("peggedUSD"))
    return to_float(value)


def stablecoin_history_points(values: dict[str, float | None], today: date) -> list[tuple[date, float]]:
    points = [
        (today - timedelta(days=30), stablecoin_billions(values.get("prev_month"))),
        (today - timedelta(days=7), stablecoin_billions(values.get("prev_week"))),
        (today - timedelta(days=1), stablecoin_billions(values.get("prev_day"))),
        (today, stablecoin_billions(values.get("current"))),
    ]
    return [(observed_at, value) for observed_at, value in points if value is not None]


def stablecoin_billions(value: float | None) -> float | None:
    return value / 1_000_000_000 if value is not None else None


def collect_world_bank_commodity_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    del today
    commodity_config = config.get("world_bank_commodities", {})
    if not commodity_config.get("enabled", True):
        return []

    items = commodity_config.get("items", [])
    if not items:
        return []

    page_url = str(
        commodity_config.get("page_url")
        or "https://www.worldbank.org/en/research/commodity-markets"
    )
    xlsx_url = str(
        commodity_config.get("download_url") or find_world_bank_monthly_xlsx_url(page_url, session)
    )
    response = session.get(xlsx_url, timeout=(5, 60))
    response.raise_for_status()
    workbook = load_workbook(BytesIO(response.content), data_only=True, read_only=True)
    sheet_name = str(commodity_config.get("sheet") or "Monthly Prices")
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    price_table = parse_world_bank_monthly_prices(
        sheet,
        header_row=int(commodity_config.get("header_row", 5)),
        data_start_row=int(commodity_config.get("data_start_row", 7)),
    )

    metrics: list[dict[str, Any]] = []
    for item in items:
        column = str(item.get("column") or "").strip()
        name = str(item.get("name") or column or "World Bank 원자재 가격")
        industry = str(item.get("industry") or "철강/소재")
        unit = str(item.get("unit") or "")
        group = str(item.get("group") or "원자재 가격")
        meaning = str(item.get("meaning") or infer_metric_meaning(industry, name))

        points = world_bank_column_points(price_table, column)
        if not points:
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="World Bank Commodity Markets Pink Sheet",
                    source_url=page_url,
                    frequency="월간",
                    automation="무료 공개 엑셀 자동 수집",
                    status="error",
                    note=f"{column} 열을 찾을 수 없음",
                    group=group,
                    meaning=meaning,
                )
            )
            continue

        scale = to_float(item.get("scale")) or 1.0
        scaled_points = [(observed_month, value * scale) for observed_month, value in points]
        latest_month, latest_value = scaled_points[-1]
        previous_value = scaled_points[-2][1] if len(scaled_points) > 1 else None
        yoy_value = find_yoy_value(scaled_points, latest_month)
        metrics.append(
            make_metric(
                industry=industry,
                name=name,
                source="World Bank Commodity Markets Pink Sheet",
                source_url=xlsx_url,
                frequency="월간",
                automation="무료 공개 엑셀 자동 수집",
                status="ok",
                value=latest_value,
                unit=unit,
                observed_at=latest_month.isoformat(),
                previous_value=previous_value,
                yoy_value=yoy_value,
                history=scaled_points,
                note=str(item.get("note") or ""),
                group=group,
                meaning=meaning,
                history_key=f"worldbank-{column}",
            )
        )
    return metrics


def find_world_bank_monthly_xlsx_url(page_url: str, session: requests.Session) -> str:
    response = session.get(page_url, timeout=(5, 20))
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    candidates: list[str] = []
    for anchor in soup.find_all("a", href=True):
        href = str(anchor["href"])
        lower_href = href.lower()
        if lower_href.endswith(".xlsx") and "monthly" in lower_href:
            candidates.append(urljoin(page_url, href))
        elif "cmo-historical-data-monthly" in lower_href and ".xlsx" in lower_href:
            candidates.append(urljoin(page_url, href))
    if not candidates:
        raise ValueError("World Bank 월간 원자재 엑셀 링크를 찾을 수 없음")
    return candidates[0]


def parse_world_bank_monthly_prices(
    sheet: Any, *, header_row: int, data_start_row: int
) -> dict[str, list[tuple[date, float]]]:
    columns: dict[int, str] = {}
    values: dict[str, list[tuple[date, float]]] = {}
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_index == header_row:
            for column_index, header in enumerate(row[1:], start=1):
                if header is None:
                    continue
                name = str(header).strip()
                if name:
                    columns[column_index] = name
                    values[name] = []
            continue

        if row_index < data_start_row or not columns:
            continue

        observed_month = parse_world_bank_month(row[0] if row else None)
        if observed_month is None:
            continue
        for column_index, name in columns.items():
            if column_index >= len(row):
                continue
            value = to_float(row[column_index])
            if value is not None:
                values[name].append((observed_month, value))
    return {name: points for name, points in values.items() if points}


def world_bank_column_points(
    price_table: dict[str, list[tuple[date, float]]], column_name: str
) -> list[tuple[date, float]]:
    if column_name in price_table:
        return price_table[column_name]

    normalized_target = normalize_lookup_text(column_name)
    for name, points in price_table.items():
        if normalize_lookup_text(name) == normalized_target:
            return points
    return []


def parse_world_bank_month(value: object) -> date | None:
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)

    text = str(value or "").strip()
    if not text:
        return None
    if "M" in text:
        year_text, month_text = text.split("M", 1)
        try:
            return date(int(year_text), int(month_text), 1)
        except ValueError:
            return None
    try:
        parsed = datetime.strptime(text[:10], "%Y-%m-%d")
    except ValueError:
        return None
    return date(parsed.year, parsed.month, 1)


def collect_sec_capex_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    del today
    capex_config = config.get("sec_capex", {})
    if not capex_config.get("enabled", True):
        return []

    companies = capex_config.get("companies", [])
    if not companies:
        return []

    source_url = str(
        capex_config.get("source_url")
        or "https://www.sec.gov/search-filings/edgar-application-programming-interfaces"
    )
    user_agent = str(
        os.getenv("SEC_USER_AGENT")
        or capex_config.get("user_agent")
        or "stock-industry-dashboard/0.1 contact@example.com"
    )
    metrics: list[dict[str, Any]] = []

    for company in companies:
        raw_cik = str(company.get("cik") or "").strip()
        if not raw_cik:
            continue
        cik = raw_cik.zfill(10)
        ticker = str(company.get("ticker") or cik)
        name = str(company.get("name") or ticker or "CAPEX")
        metric_name = str(company.get("metric_name") or name)
        configured_tags = capex_config.get("tags")
        tags = [str(tag) for tag in configured_tags] if isinstance(configured_tags, list) else None

        api_url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
        try:
            response = session.get(
                api_url,
                headers={
                    "User-Agent": user_agent,
                    "Accept": "application/json",
                    "Accept-Encoding": "gzip, deflate",
                },
                timeout=(5, 30),
            )
            response.raise_for_status()
            points = sec_capex_points(response.json(), tags)
            if not points:
                metrics.append(
                    make_metric(
                        industry="데이터인프라",
                        name=metric_name,
                        source="SEC Company Facts API",
                        source_url=source_url,
                        frequency="분기",
                        automation="무료 공식 API 자동 수집",
                        status="error",
                        note="CAPEX 태그 관측값 없음",
                        group="CAPEX",
                        meaning=sec_capex_meaning(name),
                    )
                )
                continue

            billion_points = [(observed_at, value / 1_000_000_000) for observed_at, value in points]
            latest_date, latest_value = billion_points[-1]
            previous_value = billion_points[-2][1] if len(billion_points) > 1 else None
            yoy_value = find_yoy_value(billion_points, latest_date)
            metrics.append(
                make_metric(
                    industry="데이터인프라",
                    name=metric_name,
                    source="SEC Company Facts API",
                    source_url=api_url,
                    frequency="분기",
                    automation="무료 공식 API 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit="$B",
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=billion_points,
                    group="CAPEX",
                    meaning=sec_capex_meaning(name),
                    history_key=f"sec-capex-{ticker}",
                )
            )
        except Exception as exc:  # noqa: BLE001 - keep company cards independent.
            metrics.append(
                make_metric(
                    industry="데이터인프라",
                    name=metric_name,
                    source="SEC Company Facts API",
                    source_url=source_url,
                    frequency="분기",
                    automation="무료 공식 API 자동 수집",
                    status="error",
                    note=str(exc),
                    group="CAPEX",
                    meaning=sec_capex_meaning(name),
                )
            )
    return metrics


def sec_capex_points(payload: dict[str, Any], tags: list[str] | None = None) -> list[tuple[date, float]]:
    tag_candidates = tags or [
        "PaymentsToAcquirePropertyPlantAndEquipment",
        "PaymentsToAcquirePropertyAndEquipment",
        "PaymentsToAcquireProductiveAssets",
        "CapitalExpenditures",
    ]
    us_gaap = payload.get("facts", {}).get("us-gaap", {})
    best_points: list[tuple[date, float]] = []
    for tag in tag_candidates:
        fact = us_gaap.get(tag, {})
        rows = fact.get("units", {}).get("USD", [])
        if not rows:
            continue

        by_period: dict[tuple[date, date], tuple[str, float]] = {}
        for row in rows:
            if str(row.get("form") or "") not in {"10-Q", "10-K"}:
                continue
            start_date = parse_iso_date(row.get("start"))
            end_date = parse_iso_date(row.get("end"))
            value = to_float(row.get("val"))
            if start_date is None or end_date is None or value is None:
                continue
            duration = (end_date - start_date).days + 1
            if not 70 <= duration <= 380:
                continue
            filed = str(row.get("filed") or "")
            key = (start_date, end_date)
            if key not in by_period or filed >= by_period[key][0]:
                by_period[key] = (filed, abs(value))

        direct_by_end: dict[date, tuple[str, float]] = {}
        cumulative_by_start: dict[date, list[tuple[date, str, float]]] = defaultdict(list)
        for (start_date, end_date), (filed, value) in by_period.items():
            if is_quarter_duration(start_date, end_date):
                current = direct_by_end.get(end_date)
                if current is None or filed >= current[0]:
                    direct_by_end[end_date] = (filed, value)
            cumulative_by_start[start_date].append((end_date, filed, value))

        derived_by_end: dict[date, tuple[str, float]] = {}
        for cumulative_rows in cumulative_by_start.values():
            cumulative_rows.sort(key=lambda item: item[0])
            previous_end: date | None = None
            previous_value: float | None = None
            for end_date, filed, value in cumulative_rows:
                increment_days = (end_date - previous_end).days if previous_end is not None else 0
                if (
                    previous_value is not None
                    and 70 <= increment_days <= 110
                    and end_date not in direct_by_end
                ):
                    derived = value - previous_value
                    if derived >= 0:
                        derived_by_end[end_date] = (filed, derived)
                previous_end = end_date
                previous_value = value

        quarterly = {**derived_by_end, **direct_by_end}
        points = sorted((end_date, value) for end_date, (_, value) in quarterly.items())
        if points and (not best_points or points[-1][0] > best_points[-1][0]):
            best_points = points
    return best_points


def is_quarter_duration(start_date: date, end_date: date) -> bool:
    days = (end_date - start_date).days + 1
    return 70 <= days <= 110


def collect_usaspending_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    spending_config = config.get("usaspending", {})
    if not spending_config.get("enabled", True):
        return []

    items = spending_config.get("items", [])
    if not items:
        return []

    endpoint = str(
        spending_config.get("endpoint")
        or "https://api.usaspending.gov/api/v2/search/spending_over_time/"
    )
    metrics: list[dict[str, Any]] = []

    for index, item in enumerate(items):
        history_key = f"usaspending-{item.get('key') or index}"
        months_back = int(item.get("months_back") or spending_config.get("months_back") or 18)
        # 최초 1회는 과거 구간을 넓게 백필합니다. 한 번의 응답에 전체 기간이 담기므로 추가 호출 부담이 없습니다.
        if cached_history_last_date(config, history_key) is None:
            months_back = max(
                months_back, int(spending_config.get("backfill_months") or 144)
            )
        current_month = date(today.year, today.month, 1)
        end_month = add_months(current_month, -1)
        end_date = current_month - timedelta(days=1)
        start_month = add_months(end_month, -months_back + 1)
        filters: dict[str, Any] = {
            "time_period": [
                {"start_date": start_month.isoformat(), "end_date": end_date.isoformat()}
            ],
        }
        if item.get("naics_codes"):
            filters["naics_codes"] = [str(code) for code in item.get("naics_codes", [])]
        if item.get("award_type_codes"):
            filters["award_type_codes"] = [str(code) for code in item.get("award_type_codes", [])]
        if item.get("toptier_agencies"):
            filters["toptier_agencies"] = [str(code) for code in item.get("toptier_agencies", [])]
        if item.get("awarding_agency_codes"):
            filters["awarding_agency_codes"] = [
                str(code) for code in item.get("awarding_agency_codes", [])
            ]
        if item.get("agencies"):
            filters["agencies"] = item.get("agencies", [])

        payload = {
            "group": "month",
            "subawards": False,
            "filters": filters,
        }
        name = str(item.get("name") or "미국 방산 계약 의무액")
        industry = str(item.get("industry") or "방산")
        group = str(item.get("group") or "국방 계약")
        meaning = str(
            item.get("meaning")
            or "미국 연방 방산 계약 의무액으로 방산 수주와 예산 집행 모멘텀을 확인합니다."
        )

        def cached_metric(note: str) -> dict[str, Any] | None:
            store = attach_history_store(config)
            cached_points = store.series(history_key) if store is not None else []
            if not cached_points:
                return None
            latest_month, latest_value = cached_points[-1]
            previous_value = cached_points[-2][1] if len(cached_points) > 1 else None
            metric = make_metric(
                industry=industry,
                name=name,
                source="USAspending API",
                source_url=endpoint,
                frequency="월간",
                automation="무료 공식 API 자동 수집",
                status="ok",
                value=latest_value,
                unit="$B",
                observed_at=latest_month.isoformat(),
                previous_value=previous_value,
                yoy_value=find_yoy_value(cached_points, latest_month),
                history=cached_points,
                note=f"{note}; 이전 저장값 표시",
                group=group,
                meaning=meaning,
                history_key=history_key,
            )
            metric["fetch_attempt_failed"] = True
            return metric

        try:
            response = session.post(endpoint, json=payload, timeout=(5, 30))
            response.raise_for_status()
            points = parse_usaspending_monthly_amounts(response.json())
            if not points:
                metrics.append(
                    cached_metric("관측값 없음")
                    or
                    make_metric(
                        industry=industry,
                        name=name,
                        source="USAspending API",
                        source_url=endpoint,
                        frequency="월간",
                        automation="무료 공식 API 자동 수집",
                        status="error",
                        note="관측값 없음",
                        group=group,
                        meaning=meaning,
                    )
                )
                continue

            billion_points = [(observed_month, value / 1_000_000_000) for observed_month, value in points]
            latest_month, latest_value = billion_points[-1]
            previous_value = billion_points[-2][1] if len(billion_points) > 1 else None
            yoy_value = find_yoy_value(billion_points, latest_month)
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="USAspending API",
                    source_url=endpoint,
                    frequency="월간",
                    automation="무료 공식 API 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit="$B",
                    observed_at=latest_month.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=billion_points,
                    group=group,
                    meaning=meaning,
                    history_key=history_key,
                )
            )
        except Exception as exc:  # noqa: BLE001 - one spending item should not break the page.
            metrics.append(
                cached_metric(f"API 응답 실패: {exc}")
                or
                make_metric(
                    industry=industry,
                    name=name,
                    source="USAspending API",
                    source_url=endpoint,
                    frequency="월간",
                    automation="무료 공식 API 자동 수집",
                    status="error",
                    note=str(exc),
                    group=group,
                    meaning=meaning,
                )
            )
    return metrics


def parse_usaspending_monthly_amounts(payload: dict[str, Any]) -> list[tuple[date, float]]:
    points: list[tuple[date, float]] = []
    for row in payload.get("results", []):
        period = row.get("time_period", {})
        fiscal_year = period.get("fiscal_year")
        fiscal_month = period.get("month")
        value = to_float(row.get("aggregated_amount"))
        if fiscal_year is None or fiscal_month is None or value is None:
            continue
        try:
            observed_month = fiscal_month_to_calendar_date(fiscal_year, fiscal_month)
        except ValueError:
            continue
        points.append((observed_month, value))
    return sorted(points)


def fiscal_month_to_calendar_date(fiscal_year: object, fiscal_month: object) -> date:
    year = int(fiscal_year)
    month = int(fiscal_month)
    if not 1 <= month <= 12:
        raise ValueError(f"Invalid fiscal month: {fiscal_month}")
    if month <= 3:
        return date(year - 1, month + 9, 1)
    return date(year, month - 3, 1)


def collect_eia_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    del today
    eia_config = config.get("eia", {})
    if not eia_config.get("enabled", True):
        return []

    series_config = eia_config.get("series", [])
    if not series_config:
        return []

    api_key = os.getenv("EIA_API_KEY", "").strip()
    source_url = str(eia_config.get("source_url") or "https://www.eia.gov/opendata/")
    if not api_key:
        return [
            make_metric(
                industry=str(series.get("industry") or "매크로"),
                name=str(series.get("name") or series.get("series_id")),
                source="EIA Open Data API",
                source_url=source_url,
                frequency=str(series.get("frequency") or ""),
                automation="무료 공식 API 자동 수집",
                status="needs_key",
                note="GitHub Secrets에 EIA_API_KEY 등록 필요",
                group=str(series.get("group") or ""),
                meaning=str(series.get("meaning") or ""),
            )
            for series in series_config
            if series.get("series_id")
        ]

    metrics: list[dict[str, Any]] = []
    for series in series_config:
        series_id = str(series.get("series_id") or "").strip()
        if not series_id:
            continue

        name = str(series.get("name") or series_id)
        industry = str(series.get("industry") or "매크로")
        unit = str(series.get("unit") or "")
        frequency = str(series.get("frequency") or "")
        value_field = str(series.get("value_field") or "")
        api_url = f"https://api.eia.gov/v2/seriesid/{series_id}"
        history_key = f"eia-{series_id}"
        # 최초 백필은 EIA 1회 응답 최대치(5000행), 이후에는 최근 구간만 갱신합니다.
        fetch_limit = 5000 if cached_history_last_date(config, history_key) is None else 120

        try:
            response = session.get(
                api_url,
                params={
                    "api_key": api_key,
                    "sort[0][column]": "period",
                    "sort[0][direction]": "desc",
                    "length": fetch_limit,
                },
                timeout=(5, 30),
            )
            response.raise_for_status()
            points = parse_eia_points(response.json(), value_field)
            if not points:
                metrics.append(
                    make_metric(
                        industry=industry,
                        name=name,
                        source="EIA Open Data API",
                        source_url=source_url,
                        frequency=frequency,
                        automation="무료 공식 API 자동 수집",
                        status="error",
                        note="관측값 없음",
                        group=str(series.get("group") or ""),
                        meaning=str(series.get("meaning") or ""),
                    )
                )
                continue

            scale = to_float(series.get("scale")) or 1.0
            scaled_points = [(observed_at, value * scale) for observed_at, value in points]
            latest_date, latest_value = scaled_points[-1]
            previous_value = scaled_points[-2][1] if len(scaled_points) > 1 else None
            yoy_value = find_yoy_value(scaled_points, latest_date)
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="EIA Open Data API",
                    source_url=api_url,
                    frequency=frequency,
                    automation="무료 공식 API 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit=unit,
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=scaled_points,
                    group=str(series.get("group") or ""),
                    meaning=str(series.get("meaning") or ""),
                    history_key=history_key,
                )
            )
        except Exception as exc:  # noqa: BLE001 - one EIA series should not break the page.
            metrics.append(
                make_metric(
                    industry=industry,
                    name=name,
                    source="EIA Open Data API",
                    source_url=source_url,
                    frequency=frequency,
                    automation="무료 공식 API 자동 수집",
                    status="error",
                    note=str(exc),
                    group=str(series.get("group") or ""),
                    meaning=str(series.get("meaning") or ""),
                )
            )
    return metrics


def parse_eia_points(payload: dict[str, Any], value_field: str = "") -> list[tuple[date, float]]:
    rows = payload.get("response", {}).get("data", [])
    points: list[tuple[date, float]] = []
    for row in rows:
        observed_at = parse_eia_period(row.get("period"))
        value = eia_row_value(row, value_field)
        if observed_at is not None and value is not None:
            points.append((observed_at, value))
    return sorted(points)


def eia_row_value(row: dict[str, Any], value_field: str = "") -> float | None:
    if value_field:
        return to_float(row.get(value_field))
    for field in ["value", "price", "sales", "generation", "revenue", "customers"]:
        value = to_float(row.get(field))
        if value is not None:
            return value
    return None


def parse_eia_period(value: object) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    if len(text) == 4 and text.isdigit():
        return date(int(text), 1, 1)
    if len(text) == 7 and text[4] == "-":
        try:
            return date(int(text[:4]), int(text[5:7]), 1)
        except ValueError:
            return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def collect_openfda_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    openfda_config = config.get("openfda", {})
    if not openfda_config.get("enabled", True):
        return []

    items = openfda_config.get("items", [])
    if not items:
        return []

    endpoint = str(openfda_config.get("endpoint") or "https://api.fda.gov/drug/drugsfda.json")
    source_url = str(openfda_config.get("source_url") or "https://open.fda.gov/apis/drug/drugsfda/")
    api_key = os.getenv("OPENFDA_API_KEY", "").strip()
    metrics: list[dict[str, Any]] = []

    for index, item in enumerate(items):
        name = str(item.get("name") or "FDA 의약품 승인 활동")
        months_back = int(item.get("months_back") or openfda_config.get("months_back") or 18)
        history_key = f"openfda-{item.get('key') or index}"
        fetch_months = months_needing_fetch(
            config,
            history_key,
            today,
            months_back,
            backfill_months=int(openfda_config.get("backfill_months") or 240),
            max_backfill_per_run=int(openfda_config.get("backfill_per_run") or 24),
        )
        base_search = str(
            item.get("search")
            or "submissions.submission_status:AP"
        )
        points: list[tuple[date, float]] = []
        try:
            for month in fetch_months:
                start_text, end_text = openfda_month_range(month)
                search = f"{base_search} AND submissions.submission_status_date:[{start_text} TO {end_text}]"
                params = {"search": search, "limit": 1}
                if api_key:
                    params["api_key"] = api_key
                response = session.get(endpoint, params=params, timeout=(5, 20))
                if response.status_code == 404:
                    total = 0
                else:
                    response.raise_for_status()
                    total = int(response.json().get("meta", {}).get("results", {}).get("total") or 0)
                points.append((month, float(total)))

            metrics.append(
                event_count_metric(
                    points=points,
                    history_key=history_key,
                    industry=str(item.get("industry") or "바이오"),
                    name=name,
                    source="openFDA Drugs@FDA API",
                    source_url=source_url,
                    frequency="월간",
                    group=str(item.get("group") or "승인 이벤트"),
                    meaning=str(
                        item.get("meaning")
                        or "FDA 의약품 승인 관련 기록 수로 바이오 규제 이벤트와 신약 모멘텀을 확인합니다."
                    ),
                )
            )
        except Exception as exc:  # noqa: BLE001 - one event source should not break the page.
            metrics.append(
                make_metric(
                    industry=str(item.get("industry") or "바이오"),
                    name=name,
                    source="openFDA Drugs@FDA API",
                    source_url=source_url,
                    frequency="월간",
                    automation="무료 공개 API 자동 수집",
                    status="error",
                    note=str(exc),
                    group=str(item.get("group") or "승인 이벤트"),
                    meaning=str(item.get("meaning") or ""),
                )
            )
    return metrics


def collect_clinical_trials_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    trials_config = config.get("clinical_trials", {})
    if not trials_config.get("enabled", True):
        return []

    items = trials_config.get("items", [])
    if not items:
        return []

    endpoint = str(trials_config.get("endpoint") or "https://clinicaltrials.gov/api/v2/studies")
    source_url = str(trials_config.get("source_url") or "https://clinicaltrials.gov/data-api")
    metrics: list[dict[str, Any]] = []

    for index, item in enumerate(items):
        name = str(item.get("name") or "글로벌 임상 시작 건수")
        months_back = int(item.get("months_back") or trials_config.get("months_back") or 18)
        history_key = f"clinicaltrials-{item.get('key') or index}"
        fetch_months = months_needing_fetch(
            config,
            history_key,
            today,
            months_back,
            backfill_months=int(trials_config.get("backfill_months") or 240),
            max_backfill_per_run=int(trials_config.get("backfill_per_run") or 24),
        )
        extra_query = str(item.get("query") or "")
        points: list[tuple[date, float]] = []
        try:
            for month in fetch_months:
                start_date, end_date = month_date_range(month)
                query = (
                    f"AREA[StartDate]RANGE[{start_date.isoformat()},{end_date.isoformat()}]"
                )
                if extra_query:
                    query = f"{query} AND {extra_query}"
                response = session.get(
                    endpoint,
                    params={
                        "format": "json",
                        "pageSize": 1,
                        "countTotal": "true",
                        "query.term": query,
                    },
                    timeout=(5, 20),
                )
                response.raise_for_status()
                points.append((month, float(response.json().get("totalCount") or 0)))

            metrics.append(
                event_count_metric(
                    points=points,
                    history_key=history_key,
                    industry=str(item.get("industry") or "바이오"),
                    name=name,
                    source="ClinicalTrials.gov API",
                    source_url=source_url,
                    frequency="월간",
                    group=str(item.get("group") or "임상 이벤트"),
                    meaning=str(
                        item.get("meaning")
                        or "새로 시작되는 임상시험 수로 바이오 업계의 파이프라인 활동성을 확인합니다."
                    ),
                )
            )
        except Exception as exc:  # noqa: BLE001 - keep cards independent.
            metrics.append(
                make_metric(
                    industry=str(item.get("industry") or "바이오"),
                    name=name,
                    source="ClinicalTrials.gov API",
                    source_url=source_url,
                    frequency="월간",
                    automation="무료 공식 API 자동 수집",
                    status="error",
                    note=str(exc),
                    group=str(item.get("group") or "임상 이벤트"),
                    meaning=str(item.get("meaning") or ""),
                )
            )
    return metrics


def collect_launch_library_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    launch_config = config.get("launch_library", {})
    if not launch_config.get("enabled", True):
        return []

    items = launch_config.get("items", [])
    if not items:
        return []

    endpoint = str(launch_config.get("endpoint") or "https://ll.thespacedevs.com/2.3.0/launches/")
    source_url = str(launch_config.get("source_url") or "https://thespacedevs.com/llapi")
    metrics: list[dict[str, Any]] = []

    for index, item in enumerate(items):
        name = str(item.get("name") or "글로벌 우주 발사 건수")
        months_back = int(item.get("months_back") or launch_config.get("months_back") or 18)
        history_key = f"launchlibrary-{item.get('key') or index}"
        # Launch Library 무료 티어는 시간당 15회 제한이라 백필 폭을 보수적으로 잡습니다.
        if cached_history_last_date(config, history_key) is None:
            months_back = max(months_back, int(launch_config.get("backfill_months") or 48))
        try:
            months = completed_months(today, months_back)
            points = launch_library_monthly_counts(session, endpoint, months)
            metrics.append(
                event_count_metric(
                    points=points,
                    history_key=history_key,
                    industry=str(item.get("industry") or "우주"),
                    name=name,
                    source="The Space Devs Launch Library 2 API",
                    source_url=source_url,
                    frequency="월간",
                    group=str(item.get("group") or "발사 이벤트"),
                    meaning=str(
                        item.get("meaning")
                        or "글로벌 발사 건수로 우주 산업 활동성과 위성 인프라 수요를 확인합니다."
                    ),
                )
            )
        except Exception as exc:  # noqa: BLE001 - keep cards independent.
            metrics.append(
                make_metric(
                    industry=str(item.get("industry") or "우주"),
                    name=name,
                    source="The Space Devs Launch Library 2 API",
                    source_url=source_url,
                    frequency="월간",
                    automation="무료 공개 API 자동 수집",
                    status="error",
                    note=str(exc),
                    group=str(item.get("group") or "발사 이벤트"),
                    meaning=str(item.get("meaning") or ""),
                )
            )
    return metrics


def launch_library_monthly_counts(
    session: requests.Session, endpoint: str, months: list[date]
) -> list[tuple[date, float]]:
    if not months:
        return []

    start_date, _ = month_date_range(months[0])
    _, end_date = month_date_range(months[-1])
    counts: dict[date, float] = {month: 0.0 for month in months}
    offset = 0
    limit = 100
    while True:
        params = {
            "format": "json",
            "mode": "list",
            "limit": limit,
            "offset": offset,
            "ordering": "net",
            "net__gte": f"{start_date.isoformat()}T00:00:00Z",
            "net__lte": f"{end_date.isoformat()}T23:59:59Z",
        }
        response = get_with_rate_limit_retry(session, endpoint, params=params)
        response.raise_for_status()
        payload = response.json()
        for launch in payload.get("results", []):
            launch_date = parse_iso_date(str(launch.get("net") or "")[:10])
            if launch_date is None:
                continue
            launch_month = date(launch_date.year, launch_date.month, 1)
            if launch_month in counts:
                counts[launch_month] += 1

        offset += len(payload.get("results", []))
        total = int(payload.get("count") or 0)
        if offset >= total or not payload.get("next"):
            break
    return [(month, counts[month]) for month in months]


def get_with_rate_limit_retry(
    session: requests.Session, url: str, *, params: dict[str, Any], attempts: int = 1
) -> requests.Response:
    response: requests.Response | None = None
    for attempt in range(attempts):
        response = session.get(url, params=params, timeout=(5, 20))
        if response.status_code != 429 or attempt == attempts - 1:
            return response
        retry_after = to_float(response.headers.get("Retry-After")) or 10.0
        time.sleep(min(max(retry_after, 2.0), 20.0))
    if response is None:
        raise RuntimeError("request was not attempted")
    return response


def collect_afdc_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    afdc_config = config.get("afdc", {})
    if not afdc_config.get("enabled", True):
        return []

    api_key = os.getenv("NREL_API_KEY", "").strip() or os.getenv("NLR_API_KEY", "").strip()
    source_url = str(
        afdc_config.get("source_url")
        or "https://developer.nlr.gov/docs/transportation/alt-fuel-stations-v1/"
    )
    endpoint = str(
        afdc_config.get("endpoint")
        or "https://developer.nlr.gov/api/alt-fuel-stations/v1.json"
    )
    last_updated_endpoint = str(
        afdc_config.get("last_updated_endpoint")
        or "https://developer.nlr.gov/api/alt-fuel-stations/v1/last-updated.json"
    )
    if not api_key:
        return [
            make_metric(
                industry="전기차",
                name=str(item.get("name") or "미국 EV 충전 인프라"),
                source="NLR Alternative Fuel Stations API",
                source_url=source_url,
                frequency="일간",
                automation="무료 공식 API 자동 수집",
                status="needs_key",
                note="GitHub Secrets에 NREL_API_KEY 등록 필요",
                group=str(item.get("group") or "충전 인프라"),
                meaning=str(item.get("meaning") or ""),
            )
            for item in afdc_config.get("items", [])
        ]

    response = session.get(
        endpoint,
        params={
            "api_key": api_key,
            "fuel_type": "ELEC",
            "country": str(afdc_config.get("country") or "US"),
            "limit": 1,
        },
        timeout=(5, 20),
    )
    response.raise_for_status()
    payload = response.json()
    observed_at = today
    try:
        updated_response = session.get(
            last_updated_endpoint, params={"api_key": api_key}, timeout=(5, 20)
        )
        updated_response.raise_for_status()
        updated_at = str(updated_response.json().get("last_updated") or "")
        parsed = parse_iso_date(updated_at[:10])
        if parsed is not None:
            observed_at = parsed
    except Exception:
        observed_at = today

    counts = {
        "stations": to_float(payload.get("total_results")),
        "ports": to_float(
            payload.get("station_counts", {})
            .get("fuels", {})
            .get("ELEC", {})
            .get("total")
        ),
    }
    metrics: list[dict[str, Any]] = []
    for item in afdc_config.get("items", []):
        key = str(item.get("field") or "stations")
        value = counts.get(key)
        if value is None:
            continue
        metrics.append(
            make_metric(
                industry=str(item.get("industry") or "전기차"),
                name=str(item.get("name") or "미국 EV 충전 인프라"),
                source="NLR Alternative Fuel Stations API",
                source_url=source_url,
                frequency="일간",
                automation="무료 공식 API 자동 수집",
                status="ok",
                value=value,
                unit=str(item.get("unit") or ""),
                observed_at=observed_at.isoformat(),
                previous_value=None,
                yoy_value=None,
                history=[(observed_at, value)],
                history_key=f"afdc-{key}",
                history_merge="latest",
                group=str(item.get("group") or "충전 인프라"),
                meaning=str(
                    item.get("meaning")
                    or "미국 EV 충전 인프라 규모로 전기차를 이용하기 쉬워지고 있는지 확인합니다."
                ),
            )
        )
    return metrics


def collect_kosis_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    del today
    kosis_config = config.get("kosis", {})
    if not kosis_config.get("enabled", True):
        return []

    items = kosis_config.get("items", [])
    if not items:
        return []

    api_key = os.getenv("KOSIS_API_KEY", "").strip()
    endpoint = str(
        kosis_config.get("endpoint")
        or "https://kosis.kr/openapi/Param/statisticsParameterData.do"
    )
    source_url = str(kosis_config.get("source_url") or "https://kosis.kr/openapi/")
    history_limit = int(config.get("dashboard", {}).get("history_points", 48))

    def cached_kosis_metric(item: dict[str, Any], history_key: str, note: str) -> dict[str, Any] | None:
        store = attach_history_store(config)
        if store is None:
            return None
        points = store.series(history_key)
        if not points:
            return None
        latest_date, latest_value = points[-1]
        previous_value = points[-2][1] if len(points) > 1 else None
        yoy_value = find_yoy_value(points, latest_date)
        return make_metric(
            industry=str(item.get("industry") or "건설/부동산"),
            name=str(item.get("name") or item.get("tbl_id") or "KOSIS 지표"),
            source="KOSIS OpenAPI",
            source_url=source_url,
            frequency=str(item.get("frequency") or "월간"),
            automation="무료 공식 API 자동 수집",
            status="ok",
            value=latest_value,
            unit=str(item.get("unit") or ""),
            observed_at=latest_date.isoformat(),
            previous_value=previous_value,
            yoy_value=yoy_value,
            history=points[-history_limit:],
            note=f"{note}; 이전 저장값 표시",
            group=str(item.get("group") or "국내 주택"),
            meaning=str(item.get("meaning") or ""),
            history_key=history_key,
        )

    if not api_key:
        metrics: list[dict[str, Any]] = []
        for item in items:
            if not item.get("tbl_id") or not item.get("item_id"):
                continue
            org_id = str(item.get("org_id") or item.get("orgId") or "").strip()
            tbl_id = str(item.get("tbl_id") or item.get("tblId") or "").strip()
            item_id = item.get("item_id") or item.get("itmId")
            history_key = f"kosis-{org_id}-{tbl_id}-{kosis_code_param(item_id).rstrip('+')}"
            cached = cached_kosis_metric(item, history_key, "KOSIS_API_KEY 없음")
            metrics.append(
                cached
                or make_metric(
                    industry=str(item.get("industry") or "건설/부동산"),
                    name=str(item.get("name") or item.get("tbl_id") or "KOSIS 지표"),
                    source="KOSIS OpenAPI",
                    source_url=source_url,
                    frequency=str(item.get("frequency") or "월간"),
                    automation="무료 공식 API 자동 수집",
                    status="needs_key",
                    note="GitHub Secrets에 KOSIS_API_KEY 등록 필요",
                    group=str(item.get("group") or "국내 주택"),
                    meaning=str(item.get("meaning") or ""),
                )
            )
        return metrics

    metrics: list[dict[str, Any]] = []
    for item in items:
        name = str(item.get("name") or item.get("tbl_id") or "KOSIS 지표")
        org_id = str(item.get("org_id") or item.get("orgId") or "").strip()
        tbl_id = str(item.get("tbl_id") or item.get("tblId") or "").strip()
        item_id = item.get("item_id") or item.get("itmId")
        prd_se = str(item.get("prd_se") or item.get("prdSe") or "M").strip()
        if not org_id or not tbl_id or not item_id:
            continue

        history_key = f"kosis-{org_id}-{tbl_id}-{kosis_code_param(item_id).rstrip('+')}"
        # 최초 백필은 통계표 최대 기간(600기), 이후에는 최근 기간만 갱신합니다.
        if cached_history_last_date(config, history_key) is None:
            fetch_periods = int(item.get("backfill_points") or 600)
        else:
            fetch_periods = int(item.get("history_points") or history_limit)

        params: dict[str, Any] = {
            "method": "getList",
            "apiKey": api_key,
            "format": "json",
            "jsonVD": "Y",
            "prdSe": prd_se,
            "newEstPrdCnt": fetch_periods,
            "prdInterval": int(item.get("prd_interval") or 1),
            "orgId": org_id,
            "tblId": tbl_id,
            "itmId": kosis_code_param(item_id),
        }
        for index in range(1, 9):
            key = f"objL{index}"
            value = item.get(key)
            params[key] = kosis_code_param(value) if value is not None else ""

        try:
            response = session.get(
                endpoint,
                params=params,
                headers={"User-Agent": "Mozilla/5.0 stock-industry-dashboard/1.0"},
                timeout=(10, 60),
            )
            response.raise_for_status()
            payload = response.json()
            points = parse_kosis_points(payload, prd_se)
            if not points:
                cached = cached_kosis_metric(item, history_key, "KOSIS 관측값 없음")
                metrics.append(
                    cached
                    or
                    make_metric(
                        industry=str(item.get("industry") or "건설/부동산"),
                        name=name,
                        source="KOSIS OpenAPI",
                        source_url=source_url,
                        frequency=str(item.get("frequency") or "월간"),
                        automation="무료 공식 API 자동 수집",
                        status="error",
                        note="관측값 없음",
                        group=str(item.get("group") or "국내 주택"),
                        meaning=str(item.get("meaning") or ""),
                    )
                )
                continue

            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            yoy_value = find_yoy_value(points, latest_date)
            metrics.append(
                make_metric(
                    industry=str(item.get("industry") or "건설/부동산"),
                    name=name,
                    source="KOSIS OpenAPI",
                    source_url=source_url,
                    frequency=str(item.get("frequency") or "월간"),
                    automation="무료 공식 API 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit=str(item.get("unit") or ""),
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=points,
                    group=str(item.get("group") or "국내 주택"),
                    meaning=str(item.get("meaning") or ""),
                    history_key=history_key,
                )
            )
        except Exception as exc:  # noqa: BLE001 - one KOSIS table should not break the page.
            cached = cached_kosis_metric(item, history_key, f"KOSIS 응답 실패: {exc}")
            metrics.append(
                cached
                or
                make_metric(
                    industry=str(item.get("industry") or "건설/부동산"),
                    name=name,
                    source="KOSIS OpenAPI",
                    source_url=source_url,
                    frequency=str(item.get("frequency") or "월간"),
                    automation="무료 공식 API 자동 수집",
                    status="error",
                    note=str(exc),
                    group=str(item.get("group") or "국내 주택"),
                    meaning=str(item.get("meaning") or ""),
                )
            )
    return metrics


def kosis_code_param(value: object) -> str:
    if isinstance(value, list):
        codes = [str(item).strip() for item in value if str(item).strip()]
    else:
        codes = [part.strip() for part in str(value).replace(",", "+").split("+") if part.strip()]
    return "+".join(codes) + ("+" if codes else "")


def parse_kosis_points(payload: object, prd_se: str) -> list[tuple[date, float]]:
    if isinstance(payload, dict):
        error_code = payload.get("err") or payload.get("errCd")
        if error_code:
            raise ValueError(str(payload.get("errMsg") or payload))
        rows = payload.get("data") or payload.get("result") or []
    else:
        rows = payload

    period_values: dict[date, float] = {}
    if not isinstance(rows, list):
        return []

    for row in rows:
        if not isinstance(row, dict):
            continue
        observed_at = parse_kosis_period(row.get("PRD_DE") or row.get("prdDe"), prd_se)
        value = to_float(row.get("DT") or row.get("dt"))
        if observed_at is None or value is None:
            continue
        period_values[observed_at] = period_values.get(observed_at, 0.0) + value
    return sorted(period_values.items())


def parse_kosis_period(value: object, prd_se: str = "M") -> date | None:
    text = re.sub(r"[^0-9]", "", str(value or ""))
    if not text:
        return None
    period = prd_se.upper()
    try:
        if period == "D" and len(text) >= 8:
            return date(int(text[:4]), int(text[4:6]), int(text[6:8]))
        if period == "M" and len(text) >= 6:
            return date(int(text[:4]), int(text[4:6]), 1)
        if period == "Q" and len(text) >= 5:
            quarter = int(text[4])
            return date(int(text[:4]), (quarter - 1) * 3 + 1, 1)
        if len(text) >= 4:
            return date(int(text[:4]), 1, 1)
    except ValueError:
        return None
    return None


def completed_months(today: date, months_back: int) -> list[date]:
    end_month = add_months(date(today.year, today.month, 1), -1)
    start_month = add_months(end_month, -months_back + 1)
    months: list[date] = []
    cursor = start_month
    while cursor <= end_month:
        months.append(cursor)
        cursor = add_months(cursor, 1)
    return months


def month_date_range(month: date) -> tuple[date, date]:
    return month, add_months(month, 1) - timedelta(days=1)


def months_needing_fetch(
    config: dict[str, Any],
    history_key: str,
    today: date,
    months_back: int,
    backfill_months: int,
    max_backfill_per_run: int,
) -> list[date]:
    """이벤트 카운트형 소스가 이번 실행에서 조회할 월 목록.

    최근 months_back개월은 항상 다시 세고(소급 반영), 그보다 오래된 구간은
    캐시에 없는 달만 실행당 max_backfill_per_run개씩 점진적으로 백필해
    무료 API 호출 한도를 넘지 않게 합니다.
    """
    recent = completed_months(today, months_back)
    store = attach_history_store(config)
    if store is None or not recent:
        return recent
    cached_dates = {point[0] for point in store.series(history_key)}
    older = [
        month
        for month in completed_months(today, max(backfill_months, months_back))
        if month < recent[0] and month not in cached_dates
    ]
    backfill = older[-max_backfill_per_run:] if max_backfill_per_run > 0 else []
    return sorted(set(backfill + recent))


def openfda_month_range(month: date) -> tuple[str, str]:
    start_date, end_date = month_date_range(month)
    return start_date.strftime("%Y%m%d"), end_date.strftime("%Y%m%d")


def event_count_metric(
    *,
    points: list[tuple[date, float]],
    industry: str,
    name: str,
    source: str,
    source_url: str,
    frequency: str,
    group: str,
    meaning: str,
    history_key: str = "",
) -> dict[str, Any]:
    if not points:
        return make_metric(
            industry=industry,
            name=name,
            source=source,
            source_url=source_url,
            frequency=frequency,
            automation="무료 공개 API 자동 수집",
            status="error",
            note="관측값 없음",
            group=group,
            meaning=meaning,
        )
    latest_month, latest_value = points[-1]
    previous_value = points[-2][1] if len(points) > 1 else None
    yoy_value = find_yoy_value(points, latest_month)
    return make_metric(
        industry=industry,
        name=name,
        source=source,
        source_url=source_url,
        frequency=frequency,
        automation="무료 공개 API 자동 수집",
        status="ok",
        value=latest_value,
        unit="건",
        observed_at=latest_month.isoformat(),
        previous_value=previous_value,
        yoy_value=yoy_value,
        history=points,
        group=group,
        meaning=meaning,
        history_key=history_key,
    )


def normalize_lookup_text(value: str) -> str:
    return " ".join(value.lower().replace("*", "").split())


def wsts_metric_name(region: str, is_3mma: bool) -> str:
    return f"3MMA - {region}" if is_3mma else region


def collect_wsts_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    del today
    wsts_config = config.get("wsts", {})
    if not wsts_config.get("enabled", True):
        return []

    xlsx_url = wsts_config.get("download_url") or find_wsts_xlsx_url(
        str(wsts_config["page_url"]), session
    )
    response = session.get(str(xlsx_url), timeout=60)
    response.raise_for_status()
    workbook = load_workbook(BytesIO(response.content), data_only=True, read_only=True)

    regions = [str(region) for region in wsts_config.get("regions", ["Worldwide"])]
    metrics: list[dict[str, Any]] = []
    metrics.extend(
        wsts_sheet_metrics(
            workbook["Monthly Data"],
            regions,
            is_3mma=False,
            xlsx_url=str(xlsx_url),
        )
    )
    if wsts_config.get("include_3mma", True) and "3MMA" in workbook.sheetnames:
        metrics.extend(
            wsts_sheet_metrics(
                workbook["3MMA"],
                regions,
                is_3mma=True,
                xlsx_url=str(xlsx_url),
            )
        )
    return metrics


def wsts_sheet_metrics(
    sheet: Any, regions: list[str], is_3mma: bool, xlsx_url: str
) -> list[dict[str, Any]]:
    parsed = parse_wsts_sheet(sheet)
    metrics: list[dict[str, Any]] = []
    for region in regions:
        points = sorted(parsed.get(region, []), key=lambda point: point[0])
        name = wsts_metric_name(region, is_3mma)
        meaning = wsts_metric_meaning(region, is_3mma)
        if not points:
            metrics.append(
                make_metric(
                    industry="반도체",
                    name=name,
                    source="WSTS",
                    source_url=xlsx_url,
                    frequency="월간",
                    automation="무료로 안정적으로 자동화 가능",
                    status="error",
                    note="선택한 지역 데이터 없음",
                    group="판매액(WSTS)",
                    depth="전체 업황",
                    meaning=meaning,
                )
            )
            continue

        billion_points = [(observed_date, value / 1_000_000) for observed_date, value in points]
        latest_date, latest_value = billion_points[-1]
        previous_value = billion_points[-2][1] if len(billion_points) > 1 else None
        yoy_value = find_yoy_value(billion_points, latest_date)
        metrics.append(
            make_metric(
                industry="반도체",
                name=name,
                source="WSTS Historical Billings Report",
                source_url=xlsx_url,
                frequency="월간",
                automation="무료로 안정적으로 자동화 가능",
                status="ok",
                value=latest_value,
                unit="$B",
                observed_at=latest_date.isoformat(),
                previous_value=previous_value,
                yoy_value=yoy_value,
                history=billion_points,
                group="판매액(WSTS)",
                depth="전체 업황",
                meaning=meaning,
                history_key=f"wsts-{'3mma-' if is_3mma else ''}{region}",
            )
        )
    return metrics


def collect_korea_export_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    export_config = config.get("korea_exports", {})
    if not export_config.get("enabled", True):
        return []

    items = export_config.get("items", [])
    if not items:
        return []

    endpoint = str(export_config["endpoint"])
    source_url = "https://www.data.go.kr/data/15101609/openapi.do"
    service_key = os.getenv("DATA_GO_KR_SERVICE_KEY", "").strip()
    end_month = add_months(date(today.year, today.month, 1), -int(export_config.get("end_offset_months", 1)))
    months_back = int(export_config.get("months_back", 15))
    backfill_months = int(export_config.get("backfill_months", 120))

    metrics: list[dict[str, Any]] = []
    for item in items:
        name = str(item.get("name") or item.get("hs_code"))
        hs_code = str(item.get("hs_code", "")).strip()
        industry = str(item.get("industry") or infer_export_industry(hs_code))
        metric_name = f"한국 수출 {name}({hs_code})"
        history_key = f"korea-export-{hs_code}"
        # 최초 1회는 backfill_months까지 12개월 창 단위로 백필하고, 이후엔 최근 구간만 갱신합니다.
        item_months_back = (
            backfill_months
            if cached_history_last_date(config, history_key) is None
            else months_back
        )
        start_month = add_months(end_month, -item_months_back + 1)

        if not service_key:
            metrics.append(
                make_metric(
                    industry=industry,
                    name=metric_name,
                    source="관세청 품목별 수출입실적 API",
                    source_url=source_url,
                    frequency="월간",
                    automation="무료로 안정적으로 자동화 가능",
                    status="needs_key",
                    note="GitHub Secrets에 DATA_GO_KR_SERVICE_KEY 등록 필요",
                    group=str(item.get("group") or "수출"),
                    depth=str(item.get("depth") or ""),
                    meaning=str(item.get("meaning") or export_meaning(name)),
                )
            )
            continue

        try:
            records = fetch_itemtrade_records(
                session=session,
                endpoint=endpoint,
                service_key=service_key,
                hs_code=hs_code,
                start_month=start_month,
                end_month=end_month,
            )
            monthly = monthly_export_values(records)
            if not monthly:
                metrics.append(
                    make_metric(
                        industry=industry,
                        name=metric_name,
                        source="관세청 품목별 수출입실적 API",
                        source_url=source_url,
                        frequency="월간",
                        automation="무료로 안정적으로 자동화 가능",
                        status="error",
                        note=f"{month_key(start_month)}-{month_key(end_month)} 관측값 없음",
                        group=str(item.get("group") or "수출"),
                        depth=str(item.get("depth") or ""),
                        meaning=str(item.get("meaning") or export_meaning(name)),
                    )
                )
                continue

            points = sorted((observed_month, value / 1_000_000_000) for observed_month, value in monthly.items())
            latest_month, latest_value = points[-1]
            previous_value = dict(points).get(add_months(latest_month, -1))
            yoy_value = dict(points).get(add_months(latest_month, -12))
            metrics.append(
                make_metric(
                    industry=industry,
                    name=metric_name,
                    source="관세청 품목별 수출입실적 API",
                    source_url=source_url,
                    frequency="월간",
                    automation="무료로 안정적으로 자동화 가능",
                    status="ok",
                    value=latest_value,
                    unit="$B",
                    observed_at=latest_month.isoformat(),
                    previous_value=previous_value,
                    yoy_value=yoy_value,
                    history=points,
                    group=str(item.get("group") or "수출"),
                    depth=str(item.get("depth") or ""),
                    meaning=str(item.get("meaning") or export_meaning(name)),
                    history_key=history_key,
                )
            )
        except Exception as exc:  # noqa: BLE001 - one export item should not break the page.
            metrics.append(
                make_metric(
                    industry=industry,
                    name=metric_name,
                    source="관세청 품목별 수출입실적 API",
                    source_url=source_url,
                    frequency="월간",
                    automation="무료로 안정적으로 자동화 가능",
                    status="error",
                    note=str(exc),
                    group=str(item.get("group") or "수출"),
                    depth=str(item.get("depth") or ""),
                    meaning=str(item.get("meaning") or export_meaning(name)),
                )
            )
    return metrics


def monthly_export_values(records: list[dict[str, str]]) -> dict[date, float]:
    monthly: dict[date, float] = defaultdict(float)
    for record in records:
        year_text = record.get("year", "").replace("-", ".").strip()
        try:
            year, month = year_text.split(".")[:2]
            observed_month = date(int(year), int(month), 1)
        except (ValueError, TypeError):
            continue

        export_value = to_float(record.get("expDlr"))
        if export_value is not None:
            monthly[observed_month] += export_value
    return dict(monthly)


def collect_valuation_metrics(
    config: dict[str, Any], session: requests.Session, today: date
) -> list[dict[str, Any]]:
    """밸류에이션(멀티플)과 수급 과열 지표 수집기. 크롤 기반이라 전부 soft-fail합니다."""
    val_config = config.get("valuation", {})
    if not val_config.get("enabled", True):
        return []

    metrics: list[dict[str, Any]] = []

    for item in val_config.get("multpl", []):
        slug = str(item.get("slug") or "").strip()
        if not slug:
            continue
        name = str(item.get("name") or slug)
        source_url = f"https://www.multpl.com/{slug}"
        try:
            points = fetch_multpl_series(session, slug)
            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            metrics.append(
                make_metric(
                    industry=str(item.get("industry") or "매크로"),
                    name=name,
                    source="multpl.com",
                    source_url=source_url,
                    frequency="월간",
                    automation="공개 페이지 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit=str(item.get("unit") or ""),
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=find_yoy_value(points, latest_date),
                    history=points,
                    group=str(item.get("group") or "밸류에이션"),
                    meaning=str(item.get("meaning") or ""),
                    history_key=f"multpl-{slug}",
                )
            )
        except Exception as exc:  # noqa: BLE001 - keep cards independent.
            metrics.append(
                make_metric(
                    industry=str(item.get("industry") or "매크로"),
                    name=name,
                    source="multpl.com",
                    source_url=source_url,
                    frequency="월간",
                    automation="공개 페이지 자동 수집",
                    status="error",
                    note=str(exc),
                    group=str(item.get("group") or "밸류에이션"),
                    meaning=str(item.get("meaning") or ""),
                )
            )

    finra_config = val_config.get("finra_margin", {})
    if finra_config.get("enabled", True):
        finra_url = "https://www.finra.org/rules-guidance/key-topics/margin-accounts/margin-statistics"
        finra_name = str(finra_config.get("name") or "미국 신용융자 잔액(margin debt)")
        finra_meaning = str(
            finra_config.get("meaning")
            or "미국 증권사 고객의 신용융자 총액입니다. 급증하면 레버리지 과열, 급감하면 강제 청산 국면일 수 있습니다."
        )
        try:
            points = fetch_finra_margin_series(session)
            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=finra_name,
                    source="FINRA Margin Statistics",
                    source_url=finra_url,
                    frequency="월간",
                    automation="공개 페이지 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit="$B",
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=find_yoy_value(points, latest_date),
                    history=points,
                    group="수급 과열",
                    meaning=finra_meaning,
                    history_key="finra-margin-debt",
                )
            )
        except Exception as exc:  # noqa: BLE001
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=finra_name,
                    source="FINRA Margin Statistics",
                    source_url=finra_url,
                    frequency="월간",
                    automation="공개 페이지 자동 수집",
                    status="error",
                    note=str(exc),
                    group="수급 과열",
                    meaning=finra_meaning,
                )
            )

    krx_field_units = {"PER": "배", "PBR": "배", "배당수익률": "%"}
    for item in val_config.get("krx", []):
        prefix = str(item.get("name_prefix") or "코스피")
        ind_idx = str(item.get("ind_idx") or "1")
        ind_idx2 = str(item.get("ind_idx2") or "001")
        base_key = f"krx-val-{ind_idx}-{ind_idx2}"
        source_url = "https://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd?menuId=MDC0201060201"
        cached_last = cached_history_last_date(config, f"{base_key}-PER")
        if cached_last is None:
            start = parse_iso_date(item.get("backfill_start")) or date(2010, 1, 1)
        else:
            start = cached_last - timedelta(days=30)
        try:
            series_by_field = fetch_krx_valuation_series(session, ind_idx, ind_idx2, start, today)
            for field_name, points in series_by_field.items():
                if not points:
                    continue
                latest_date, latest_value = points[-1]
                previous_value = points[-2][1] if len(points) > 1 else None
                metrics.append(
                    make_metric(
                        industry="매크로",
                        name=f"{prefix} {field_name}",
                        source="KRX 정보데이터시스템",
                        source_url=source_url,
                        frequency="일간",
                        automation="공개 페이지 자동 수집",
                        status="ok",
                        value=latest_value,
                        unit=krx_field_units.get(field_name, ""),
                        observed_at=latest_date.isoformat(),
                        previous_value=previous_value,
                        yoy_value=find_yoy_value(points, latest_date),
                        history=points,
                        group="밸류에이션",
                        meaning=str(
                            item.get(f"meaning_{field_name}")
                            or krx_valuation_meaning(prefix, field_name)
                        ),
                        history_key=f"{base_key}-{field_name}",
                    )
                )
        except Exception as exc:  # noqa: BLE001 - KRX는 비공식 엔드포인트라 차단될 수 있습니다.
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=f"{prefix} PER/PBR/배당수익률",
                    source="KRX 정보데이터시스템",
                    source_url=source_url,
                    frequency="일간",
                    automation="공개 페이지 자동 수집",
                    status="error",
                    note=f"KRX 응답 실패(차단 가능): {exc}",
                    group="밸류에이션",
                )
            )

    return metrics


def krx_valuation_meaning(prefix: str, field_name: str) -> str:
    if field_name == "PER":
        return f"{prefix} 전체의 주가수익비율입니다. 과거 분포 대비 낮으면 저평가, 높으면 고평가 구간으로 봅니다."
    if field_name == "PBR":
        return (
            f"{prefix} 전체의 주가순자산비율입니다. 역사적으로 코스피 PBR 0.9 이하는 장기 저평가, "
            "1.3 이상은 고평가 구간으로 통했습니다."
        )
    return f"{prefix} 전체의 배당수익률입니다. 높을수록 배당 대비 주가가 싼 상태라는 뜻입니다."


def collect_market_flow_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
) -> list[dict[str, Any]]:
    overview_config = config.get("market_overview", {}) or {}
    flows_config = overview_config.get("flows", {}) if isinstance(overview_config, dict) else {}
    if not flows_config.get("enabled", True):
        return []

    history_config = config.get("history", {}) or {}
    history_dir = str(history_config.get("dir") or "data/history")
    keep_days = int(flows_config.get("keep_calendar_days") or 760)
    lookback_days = int(flows_config.get("lookback_calendar_days") or 760)
    max_fetch = int(flows_config.get("max_backfill_days_per_run") or 3)
    endpoint = str(flows_config.get("krx_getjson_endpoint") or "")
    source_url = "https://data.krx.co.kr/"

    metric_docs: list[tuple[str, str, dict[str, Any]]] = []
    errors: list[str] = []

    for market, label in (("kospi", "KOSPI"), ("kosdaq", "KOSDAQ")):
        path = raw_flow_snapshot_path(history_dir, market)
        document = load_raw_flow_snapshot(path, market)
        missing = missing_recent_dates(
            today=today,
            known_dates=raw_flow_known_dates(document),
            calendar_days=lookback_days,
            max_fetch=max_fetch,
        )
        for target_date in missing:
            try:
                rows = fetch_krx_stock_flow_rows(
                    session,
                    market=market,
                    start_date=target_date,
                    end_date=target_date,
                    endpoint=endpoint or "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd",
                )
                document = store_raw_flow_rows(
                    history_dir=history_dir,
                    market=market,
                    observed_at=target_date,
                    rows=rows,
                    today=today,
                    keep_calendar_days=keep_days,
                )
            except Exception as exc:  # noqa: BLE001 - KRX 정보데이터시스템은 soft-fail.
                try:
                    fallback_date, rows = fetch_krx_main_investor_flow_rows(
                        session,
                        market=market,
                        endpoint=endpoint or "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd",
                    )
                    if not rows or fallback_date is None:
                        raise RuntimeError("KRX 메인 투자자별 매매동향 응답 없음")
                    document = store_raw_flow_rows(
                        history_dir=history_dir,
                        market=market,
                        observed_at=fallback_date,
                        rows=rows,
                        today=today,
                        keep_calendar_days=keep_days,
                    )
                    break
                except Exception as fallback_exc:  # noqa: BLE001
                    errors.append(f"{label} {target_date.isoformat()}: {exc}; fallback: {fallback_exc}")
                    break
        document = load_raw_flow_snapshot(path, market)
        metric_docs.append((market, label, document))

    futures_market = "k200-futures"
    futures_path = raw_flow_snapshot_path(history_dir, futures_market)
    futures_doc = load_raw_flow_snapshot(futures_path, futures_market)
    missing = missing_recent_dates(
        today=today,
        known_dates=raw_flow_known_dates(futures_doc),
        calendar_days=lookback_days,
        max_fetch=max_fetch,
    )
    for target_date in missing:
        try:
            rows = fetch_krx_futures_flow_rows(
                session,
                start_date=target_date,
                end_date=target_date,
                endpoint=endpoint or "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd",
            )
            futures_doc = store_raw_flow_rows(
                history_dir=history_dir,
                market=futures_market,
                observed_at=target_date,
                rows=rows,
                today=today,
                keep_calendar_days=keep_days,
            )
        except Exception as exc:  # noqa: BLE001
            try:
                fallback_date, rows = fetch_krx_main_investor_flow_rows(
                    session,
                    market=futures_market,
                    endpoint=endpoint or "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd",
                )
                if not rows or fallback_date is None:
                    raise RuntimeError("KRX 메인 투자자별 매매동향 응답 없음")
                futures_doc = store_raw_flow_rows(
                    history_dir=history_dir,
                    market=futures_market,
                    observed_at=fallback_date,
                    rows=rows,
                    today=today,
                    keep_calendar_days=keep_days,
                )
                break
            except Exception as fallback_exc:  # noqa: BLE001
                errors.append(f"K200 선물 {target_date.isoformat()}: {exc}; fallback: {fallback_exc}")
                break
    futures_doc = load_raw_flow_snapshot(futures_path, futures_market)
    metric_docs.append((futures_market, "K200 선물", futures_doc))

    metrics: list[dict[str, Any]] = []
    for market, market_label, document in metric_docs:
        metrics.extend(flow_metrics_from_raw_document(market, market_label, document, source_url))

    if not metrics and errors:
        metrics.append(
            make_metric(
                industry="매크로",
                name="KRX 수급 수집 상태",
                source="KRX 정보데이터시스템",
                source_url=source_url,
                frequency="일간",
                automation="공개 JSON 자동 수집",
                status="error",
                note="; ".join(errors[:3]),
                group="수급",
                section="market",
                market_category="수급",
                meaning="KRX 투자자별 매매동향 수집 상태입니다.",
            )
        )
    return metrics


def flow_metrics_from_raw_document(
    market: str,
    market_label: str,
    document: dict[str, Any],
    source_url: str,
) -> list[dict[str, Any]]:
    metrics: list[dict[str, Any]] = []
    investors = raw_flow_investors(document)
    if not investors:
        return []

    for investor in investors:
        investor_id = investor_slug(investor)
        for measure, spec in FLOW_MEASURES.items():
            points = raw_flow_series(document, investor=investor, measure=measure)
            if not points:
                continue
            latest_date, latest_value = points[-1]
            previous_value = points[-2][1] if len(points) > 1 else None
            metric_id = f"krx-flow-{market}-{investor_id}-{measure}"
            measure_label = str(spec.get("label") or measure)
            chart_style = "flow_bars" if measure == "net" else ""
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=f"{market_label} {investor} {measure_label}",
                    source="KRX 정보데이터시스템",
                    source_url=source_url,
                    frequency="일간",
                    automation="공개 JSON 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit="억원",
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=find_yoy_value(points, latest_date),
                    history=points,
                    group=investor,
                    depth=market_label,
                    meaning=flow_metric_meaning(market_label, investor, measure_label),
                    history_key=metric_id,
                    metric_id=metric_id,
                    section="market",
                    market_category="수급",
                    chart_style=chart_style,
                    exclude_from_movers=measure != "net",
                )
            )

        net_points = raw_flow_series(document, investor=investor, measure="net")
        rolling = rolling_sum_series(net_points, 20)
        if rolling:
            latest_date, latest_value = rolling[-1]
            previous_value = rolling[-2][1] if len(rolling) > 1 else None
            metric_id = f"krx-flow-{market}-{investor_id}-net-20d"
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=f"{market_label} {investor} 20일 누적 순매수",
                    source="KRX 정보데이터시스템",
                    source_url=source_url,
                    frequency="일간",
                    automation="공개 JSON 자동 수집",
                    status="ok",
                    value=latest_value,
                    unit="억원",
                    observed_at=latest_date.isoformat(),
                    previous_value=previous_value,
                    yoy_value=find_yoy_value(rolling, latest_date),
                    history=rolling,
                    group=investor,
                    depth=market_label,
                    meaning=f"{market_label}에서 {investor}이 최근 20거래일 동안 순매수한 금액의 합계입니다. 한 주체가 시장을 계속 받치는지 확인할 때 봅니다.",
                    history_key=metric_id,
                    metric_id=metric_id,
                    section="market",
                    market_category="수급",
                    chart_style="flow_bars",
                )
            )
    return metrics


def flow_metric_meaning(market_label: str, investor: str, measure_label: str) -> str:
    investor_meanings = {
        "개인": "개인투자자의 위험 선호와 반대매매·저가매수 흐름을 볼 때 중요합니다",
        "외국인": "환율, 글로벌 자금 흐름, 한국 시장 선호가 실제 매매로 들어오는지 볼 때 중요합니다",
        "기관": "국내 기관 자금이 시장을 받치는지, 리밸런싱 압력이 있는지 확인할 때 봅니다",
        "기관합계": "국내 기관 전체의 매매 방향을 한눈에 보기 위한 합산 지표입니다",
        "기관종합": "국내 기관 전체의 매매 방향을 한눈에 보기 위한 합산 지표입니다",
        "금융투자": "증권사·금융투자 쪽의 단기 포지션과 프로그램성 매매 압력을 볼 때 중요합니다",
        "보험": "보험사 자금의 주식 비중 조절 흐름을 확인할 때 봅니다",
        "투신": "펀드 자금 유입·유출이 실제 주식 매매로 이어지는지 확인할 때 봅니다",
        "기타금융": "기타 금융기관 자금의 보조적인 매매 방향을 확인할 때 봅니다",
        "은행": "은행권 자금의 위험자산 선호 변화를 보조적으로 확인할 때 봅니다",
        "연기금": "국민연금 등 장기 자금이 시장을 받치는지, 비중 조절에 나서는지 볼 때 중요합니다",
        "사모": "사모펀드와 전문투자자 쪽의 비교적 민감한 자금 흐름을 확인할 때 봅니다",
        "기타법인": "자사주, 지주회사, 일반 법인 자금이 시장에 들어오는지 볼 때 참고합니다",
    }
    investor_note = investor_meanings.get(investor, f"{investor} 자금의 매매 방향과 시장 영향력을 확인할 때 봅니다")
    if measure_label == "순매수":
        return (
            f"{market_label}에서 {investor}이 산 금액에서 판 금액을 뺀 값입니다. "
            f"플러스면 그 주체가 시장을 순매수한 것이고, 마이너스면 순매도한 것입니다. {investor_note}."
        )
    if measure_label == "매수":
        return (
            f"{market_label}에서 {investor}이 사들인 거래대금입니다. "
            f"순매수와 함께 보면 실제 매수 강도가 커진 것인지, 매도도 같이 늘어난 단순 거래 증가인지 구분할 수 있습니다. {investor_note}."
        )
    if measure_label == "매도":
        return (
            f"{market_label}에서 {investor}이 판 거래대금입니다. "
            f"매도가 빠르게 늘면 해당 주체의 차익실현이나 위험 축소 압력이 커졌는지 확인할 수 있습니다. {investor_note}."
        )
    return (
        f"{market_label}에서 {investor}의 {measure_label} 거래대금입니다. "
        f"누가 시장을 밀고 당기는지 확인하는 수급 지표입니다. {investor_note}."
    )


def metric_by_name(metrics: list[dict[str, Any]], name: str) -> dict[str, Any] | None:
    for metric in metrics:
        if isinstance(metric, dict) and str(metric.get("name") or "") == name and metric.get("status") == "ok":
            return metric
    return None


def collect_market_derived_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
    current_metrics: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    del config
    metrics: list[dict[str, Any]] = []
    metrics.extend(collect_kimchi_premium_metric(session, today, current_metrics))
    yen_vol = yen_realized_volatility_metric(today, current_metrics)
    if yen_vol:
        metrics.append(yen_vol)
    return metrics


def collect_kimchi_premium_metric(
    session: requests.Session,
    today: date,
    current_metrics: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    btc_metric = metric_by_name(current_metrics, "비트코인")
    usdkrw_metric = metric_by_name(current_metrics, "원/달러 환율")
    btc_usd = to_float((btc_metric or {}).get("value"))
    usdkrw = to_float((usdkrw_metric or {}).get("value"))
    if btc_usd is None or usdkrw is None or btc_usd <= 0 or usdkrw <= 0:
        return []
    try:
        response = session.get(
            "https://api.upbit.com/v1/ticker",
            params={"markets": "KRW-BTC"},
            timeout=(5, 20),
        )
        response.raise_for_status()
        payload = response.json()
        item = payload[0] if isinstance(payload, list) and payload else {}
        krw_btc = to_float(item.get("trade_price")) if isinstance(item, dict) else None
        if krw_btc is None or krw_btc <= 0:
            raise ValueError("업비트 BTC 가격 없음")
        premium = (krw_btc / (btc_usd * usdkrw) - 1.0) * 100.0
        observed_at = today.isoformat()
        return [
            make_metric(
                industry="매크로",
                name="김치프리미엄",
                source="Upbit/Yahoo Finance",
                source_url="https://api.upbit.com/v1/ticker",
                frequency="장중",
                automation="무료 공개 API 자동 수집",
                status="ok",
                value=premium,
                unit="%",
                observed_at=observed_at,
                previous_value=None,
                history=[(today, premium)],
                group="크립토",
                meaning="국내 비트코인 가격이 글로벌 달러 가격을 원화로 환산한 값보다 얼마나 높은지 보여줍니다. 높아질수록 국내 개인 위험선호가 강하다는 신호로 볼 수 있습니다.",
                history_key="kimchi-premium-btc",
                section="market",
                market_category="원자재·크립토",
                also_market_category=["심리·변동성"],
                refresh_scope="intraday",
            )
        ]
    except Exception as exc:  # noqa: BLE001
        return [
            make_metric(
                industry="매크로",
                name="김치프리미엄",
                source="Upbit/Yahoo Finance",
                source_url="https://api.upbit.com/v1/ticker",
                frequency="장중",
                automation="무료 공개 API 자동 수집",
                status="error",
                note=str(exc),
                group="크립토",
                meaning="국내 비트코인 가격이 글로벌 가격보다 얼마나 높은지 보여주는 지표입니다.",
                section="market",
                market_category="원자재·크립토",
            )
        ]


def yen_realized_volatility_metric(today: date, current_metrics: list[dict[str, Any]]) -> dict[str, Any] | None:
    yen = metric_by_name(current_metrics, "엔/달러 환율")
    points = parse_stored_points((yen or {}).get("history"))
    if len(points) < 11:
        return None
    returns: list[tuple[date, float]] = []
    for (prev_date, previous), (current_date, current) in zip(points[:-1], points[1:]):
        del prev_date
        if previous <= 0 or current <= 0:
            continue
        returns.append((current_date, (current / previous - 1.0) * 100.0))
    vol_points: list[tuple[date, float]] = []
    for index in range(9, len(returns)):
        window = [value for _, value in returns[index - 9 : index + 1]]
        mean = sum(window) / len(window)
        variance = sum((value - mean) ** 2 for value in window) / len(window)
        vol_points.append((returns[index][0], variance ** 0.5))
    if not vol_points:
        return None
    latest_date, latest_value = vol_points[-1]
    previous_value = vol_points[-2][1] if len(vol_points) > 1 else None
    return make_metric(
        industry="매크로",
        name="엔 환율 10일 실현변동성",
        source="Yahoo Finance chart API",
        source_url="https://finance.yahoo.com/quote/JPY=X",
        frequency="일간",
        automation="계산 지표",
        status="ok",
        value=latest_value,
        unit="%",
        observed_at=latest_date.isoformat() or today.isoformat(),
        previous_value=previous_value,
        history=vol_points,
        group="엔캐리",
        meaning="엔/달러 환율의 최근 10거래일 변동폭입니다. 변동성이 빠르게 커지면 엔캐리 포지션이 흔들릴 가능성을 점검합니다.",
        history_key="yen-realized-volatility-10d",
        section="market",
        market_category="심리·변동성",
    )


def collect_market_sentiment_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
    current_metrics: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    sentiment_config = config.get("market_sentiment", {})
    if not sentiment_config.get("enabled", True):
        return []

    metrics: list[dict[str, Any]] = []
    store = attach_history_store(config)
    metrics.extend(collect_cnn_fear_greed_metric(config, session, today))
    try:
        metrics.extend(collect_korea_fear_greed_metrics(config, session, today, current_metrics, store))
    except Exception as exc:  # noqa: BLE001 - keep CNN sentiment visible if KRX has a temporary issue.
        for name, meaning in [
            ("코스피 공포탐욕지수", korea_fear_greed_meaning("코스피")),
            ("코스닥 공포탐욕지수", korea_fear_greed_meaning("코스닥")),
            ("VKOSPI", vkospi_meaning()),
        ]:
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=name,
                    source="KRX Open API",
                    source_url=KRX_SOURCE_URL,
                    frequency="일간",
                    automation="무료 공식 API 자동 수집",
                    status="error",
                    note=str(exc),
                    group="공포탐욕",
                    meaning=meaning,
                )
            )
    return metrics


def collect_cnn_fear_greed_metric(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
) -> list[dict[str, Any]]:
    sentiment_config = config.get("market_sentiment", {})
    cnn_config = sentiment_config.get("cnn", {})
    if not cnn_config.get("enabled", True):
        return []

    history_key = "cnn-fear-greed"
    cached_last = cached_history_last_date(config, history_key)
    if cached_last is None:
        start = parse_iso_date(cnn_config.get("backfill_start")) or date(2021, 2, 1)
    else:
        start = max(date(2021, 2, 1), cached_last - timedelta(days=45))
    source_url = str(cnn_config.get("source_url") or "https://www.cnn.com/markets/fear-and-greed")
    try:
        points, _payload = fetch_cnn_fear_greed(session, start_date=start)
        if not points:
            raise ValueError("CNN 공포탐욕 관측값 없음")
        latest_date, latest_value = points[-1]
        previous_value = points[-2][1] if len(points) > 1 else None
        return [
            make_metric(
                industry="매크로",
                name="미국 CNN 공포탐욕지수",
                source="CNN Fear & Greed Index",
                source_url=source_url,
                frequency="일간",
                automation="공개 JSON 자동 수집",
                status="ok",
                value=latest_value,
                unit="점",
                observed_at=latest_date.isoformat(),
                previous_value=previous_value,
                yoy_value=find_yoy_value(points, latest_date),
                history=points,
                group="공포탐욕",
                meaning=(
                    "CNN이 미국 주식시장의 여러 심리 지표를 합산해 발표하는 공포탐욕지수입니다. "
                    "0에 가까우면 공포, 100에 가까우면 탐욕이 강한 구간입니다."
                ),
                history_key=history_key,
            )
        ]
    except Exception as exc:  # noqa: BLE001
        return [
            make_metric(
                industry="매크로",
                name="미국 CNN 공포탐욕지수",
                source="CNN Fear & Greed Index",
                source_url=source_url,
                frequency="일간",
                automation="공개 JSON 자동 수집",
                status="error",
                note=str(exc),
                group="공포탐욕",
                meaning=(
                    "CNN이 미국 주식시장의 여러 심리 지표를 합산해 발표하는 공포탐욕지수입니다. "
                    "0에 가까우면 공포, 100에 가까우면 탐욕이 강한 구간입니다."
                ),
            )
        ]


def collect_korea_fear_greed_metrics(
    config: dict[str, Any],
    session: requests.Session,
    today: date,
    current_metrics: list[dict[str, Any]],
    store: HistoryStore | None,
) -> list[dict[str, Any]]:
    sentiment_config = config.get("market_sentiment", {})
    korea_config = sentiment_config.get("korea", {})
    if not korea_config.get("enabled", True):
        return []

    source_url = str(korea_config.get("source_url") or KRX_SOURCE_URL)
    auth_key = os.getenv("KRX_OPEN_API_KEY", "").strip() or os.getenv("KRX_API_KEY", "").strip()
    if not auth_key:
        return [
            make_metric(
                industry="매크로",
                name=name,
                source="KRX Open API",
                source_url=source_url,
                frequency="일간",
                automation="무료 공식 API 자동 수집",
                status="needs_key",
                note="GitHub Secrets에 KRX_OPEN_API_KEY 등록 필요",
                group="공포탐욕",
                meaning=meaning,
            )
            for name, meaning in [
                ("코스피 공포탐욕지수", korea_fear_greed_meaning("코스피")),
                ("코스닥 공포탐욕지수", korea_fear_greed_meaning("코스닥")),
                ("VKOSPI", vkospi_meaning()),
            ]
        ]

    history_config = config.get("history", {}) or {}
    history_dir = str(history_config.get("dir") or "data/history")
    base_url = str(korea_config.get("krx_api_base") or KRX_API_BASE)
    lookback_days = int(korea_config.get("lookback_calendar_days") or 430)
    keep_days = int(korea_config.get("keep_calendar_days") or 430)
    max_fetch = int(korea_config.get("max_backfill_days_per_run") or 8)
    high_low_window_days = int(korea_config.get("high_low_window_days") or 370)
    min_high_low_points = int(korea_config.get("min_high_low_points") or 120)

    metrics: list[dict[str, Any]] = []
    snapshots: dict[str, dict[str, Any]] = {}
    for market in ("KOSPI", "KOSDAQ"):
        snapshots[market] = collect_market_snapshot(
            session,
            auth_key=auth_key,
            base_url=base_url,
            history_dir=history_dir,
            market=market,
            today=today,
            lookback_calendar_days=lookback_days,
            max_fetch_days=max_fetch,
            keep_calendar_days=keep_days,
        )

    vkospi_key = "krx-vkospi"
    vkospi_incoming = fetch_vkospi_points(
        session,
        auth_key=auth_key,
        base_url=base_url,
        store=store,
        today=today,
        history_key=vkospi_key,
        lookback_calendar_days=lookback_days,
        max_fetch_days=max_fetch,
    )
    vkospi_points = merge_existing_and_incoming(store, vkospi_key, vkospi_incoming)
    if vkospi_points:
        latest_date, latest_value = vkospi_points[-1]
        previous_value = vkospi_points[-2][1] if len(vkospi_points) > 1 else None
        metrics.append(
            make_metric(
                industry="매크로",
                name="VKOSPI",
                source="KRX Open API",
                source_url=source_url,
                frequency="일간",
                automation="무료 공식 API 자동 수집",
                status="ok",
                value=latest_value,
                unit="",
                observed_at=latest_date.isoformat(),
                previous_value=previous_value,
                yoy_value=find_yoy_value(vkospi_points, latest_date),
                history=vkospi_points,
                group="공포탐욕",
                meaning=vkospi_meaning(),
                history_key=vkospi_key,
            )
        )

    metrics_by_history_key = {
        str(metric.get("history_key") or ""): metric
        for metric in current_metrics
        if isinstance(metric, dict) and metric.get("history_key")
    }
    index_points = {
        "KOSPI": metric_full_points(store, metrics_by_history_key.get("equity-^KS11"), "equity-^KS11"),
        "KOSDAQ": metric_full_points(store, metrics_by_history_key.get("equity-^KQ11"), "equity-^KQ11"),
    }

    for market, label in (("KOSPI", "코스피"), ("KOSDAQ", "코스닥")):
        score_data = build_korea_fear_greed_score(
            market_label=label,
            index_points=index_points[market],
            snapshot_document=snapshots[market],
            vkospi_points=vkospi_points if market == "KOSPI" else None,
            high_low_window_days=high_low_window_days,
            min_high_low_points=min_high_low_points,
        )
        if score_data is None:
            metrics.append(
                make_metric(
                    industry="매크로",
                    name=f"{label} 공포탐욕지수",
                    source="KRX Open API/Yahoo Finance",
                    source_url=source_url,
                    frequency="일간",
                    automation="무료 공식 API 자동 수집",
                    status="error",
                    note="계산에 필요한 시장 심리 구성요소가 아직 부족합니다",
                    group="공포탐욕",
                    meaning=korea_fear_greed_meaning(label),
                )
            )
            continue

        score_key = f"korea-fear-greed-{market.lower()}"
        score_points = merge_existing_and_incoming(
            store,
            score_key,
            [(score_data["observed_at"], score_data["score"])],
        )
        latest_date, latest_value = score_points[-1]
        previous_value = score_points[-2][1] if len(score_points) > 1 else None
        metrics.append(
            make_metric(
                industry="매크로",
                name=f"{label} 공포탐욕지수",
                source="KRX Open API/Yahoo Finance",
                source_url=source_url,
                frequency="일간",
                automation="무료 공식 API 자동 수집",
                status="ok",
                value=latest_value,
                unit="점",
                observed_at=latest_date.isoformat(),
                previous_value=previous_value,
                yoy_value=find_yoy_value(score_points, latest_date),
                history=score_points,
                group="공포탐욕",
                meaning=korea_fear_greed_meaning(label),
                history_key=score_key,
                history_merge="latest",
            )
        )

    return metrics


def collect_reference_metrics(config: dict[str, Any]) -> list[dict[str, Any]]:
    reference_metrics = config.get("dashboard", {}).get("reference_metrics", [])
    metrics: list[dict[str, Any]] = []
    for item in reference_metrics:
        status = str(item.get("status") or "partial")
        metrics.append(
            make_metric(
                industry=str(item.get("industry") or "매크로"),
                name=str(item.get("name") or "미정 지표"),
                source=str(item.get("source") or ""),
                source_url=str(item.get("source_url") or ""),
                frequency=str(item.get("frequency") or ""),
                automation=str(item.get("automation") or status_to_automation(status)),
                status=status,
                note=str(item.get("note") or ""),
                group=str(item.get("group") or ""),
                meaning=str(item.get("meaning") or item.get("note") or ""),
            )
        )
    return metrics
